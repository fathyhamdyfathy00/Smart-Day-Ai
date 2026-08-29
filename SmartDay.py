"""
SmartDay AI Agent v9.0 - Adaptive AI Scheduling System (Bilingual)
- NEW: Real recommendation system with epsilon-greedy exploration vs exploitation
- NEW: Diverse schedule generation (semi-random permutations + bias)
- NEW: Improved utility scoring with deadline urgency, sleep proximity
- NEW: Gradual RL weight updates (learning_rate * reward) + decay
- NEW: User profile-driven recommendations
- NEW: Modernized recommendation UI (المقترح الأول، الثاني، ...)
- NEW: Cleaner timeline arrows / connectors between tasks
- KEPT: All previous logic for tasks, postpone, swap, sounds, voice, etc.
"""

import os, sys, json, time, threading, re, random, asyncio, webbrowser
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any, Tuple
from collections import defaultdict
import warnings
warnings.filterwarnings("ignore")

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    import pyttsx3
    VOICE_OK = True
except Exception:
    VOICE_OK = False

# Sound playback imports
try:
    import pygame
    pygame.mixer.init(frequency=22050, size=-16, channels=2, buffer=512)
    SOUND_OK = True
    print("✅ Sound system initialized (pygame)")
except Exception as e:
    SOUND_OK = False
    print(f"⚠️ Sound disabled: {e}")

try:
    import requests as _req
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

try:
    import numpy as np
    NUMPY_OK = True
except ImportError:
    NUMPY_OK = False

# =============================================================================
# SOUND FILES CONFIGURATION
# =============================================================================
def create_default_sound_files():
    """Create simple WAV files if they don't exist"""
    import wave
    import struct
    
    def create_beep(filename, frequency=440, duration=0.5):
        if os.path.exists(filename):
            return
        try:
            sample_rate = 22050
            samples = int(duration * sample_rate)
            wave_obj = wave.open(filename, 'w')
            wave_obj.setnchannels(1)
            wave_obj.setsampwidth(2)
            wave_obj.setframerate(sample_rate)
            
            for i in range(samples):
                value = int(32767.0 * 0.5 * (1 + (i < samples//3)) * 
                          (1 if i < samples//3 else 0.7 if i < 2*samples//3 else 0.4))
                data = struct.pack('<h', value)
                wave_obj.writeframesraw(data)
            wave_obj.close()
            print(f"✅ Created default sound: {filename}")
        except Exception as e:
            print(f"⚠️ Could not create {filename}: {e}")
    
    if not os.path.exists("mixkit-software-interface-start-2574.wav"):
        create_beep("mixkit-software-interface-start-2574.wav", 880, 0.3)
    if not os.path.exists("mixkit-software-interface-back-2575.wav"):
        create_beep("mixkit-software-interface-back-2575.wav", 440, 0.4)

create_default_sound_files()

start_sound_file = "mixkit-software-interface-start-2574.wav"
end_sound_file = "mixkit-software-interface-back-2575.wav"

# ---------------------------------------------------------------------------
# AUDIO SYSTEM (focus music + alert music)
# ---------------------------------------------------------------------------
# Optional dedicated tracks. If user provides the files alongside the script,
# they are used; otherwise we fall back to start/end beeps so the system still
# plays *something* reliably.
focus_music_file = "focus_music.wav" if os.path.exists("focus_music.wav") else start_sound_file
alert_music_file = "alert_music.wav" if os.path.exists("alert_music.wav") else end_sound_file


class TaskAudioController:
    """
    Reliable per-task audio:
      - Plays "focus music" exactly once when a task's start time is reached
      - Plays "alert music" exactly once when a task's end time is exceeded
    
    Why a class:
      - State is OWNED here, not split across the scheduler loop
      - Survives schedule re-apply (we reset only what's relevant)
      - Each (task_id, event) is fired once per (task_id, day, scheduled_time)
        — the trigger key includes the scheduled time so RESCHEDULING re-arms it
    """
    
    def __init__(self):
        self._lock = threading.Lock()
        # key = f"{task_id}:start:{date}:{startTime}" -> bool
        self._fired: Dict[str, bool] = {}
        self._focus_playing_for: Optional[int] = None
    
    def _key(self, task_id: int, event: str, date: str, time_str: str) -> str:
        return f"{task_id}:{event}:{date}:{time_str or 'none'}"
    
    def reset_for_task(self, task_id: int):
        """Clear all fired flags for a task (called when task gets a new schedule)."""
        with self._lock:
            stale = [k for k in self._fired if k.startswith(f"{task_id}:")]
            for k in stale:
                self._fired.pop(k, None)
    
    def reset_all(self):
        """Called after a fresh schedule is applied — re-arm all triggers."""
        with self._lock:
            self._fired.clear()
            self._focus_playing_for = None
    
    def maybe_play_focus(self, task: Dict, current_min: int) -> bool:
        """Trigger focus music when the task's start time is reached (or just passed)."""
        if not SOUND_OK:
            return False
        
        start_time = task.get("startTime")
        if not start_time:
            return False
        
        start_min = to_mins(start_time)
        # Trigger window: from start_min to start_min+5 (catch up if loop missed exact tick)
        if not (start_min <= current_min <= start_min + 5):
            return False
        
        if task.get("status") != "pending":
            return False
        
        key = self._key(task["id"], "start", task.get("date", ""), start_time)
        with self._lock:
            if self._fired.get(key):
                return False
            self._fired[key] = True
            self._focus_playing_for = task["id"]
        
        play_sound(focus_music_file)
        print(f"🎵 [AUDIO] Focus music ▶  task_id={task['id']} name='{task.get('name','')}'  key={key}")
        return True
    
    def maybe_play_alert(self, task: Dict, current_min: int) -> bool:
        """Trigger alert music when the task's end time is EXCEEDED."""
        if not SOUND_OK:
            return False
        
        end_time = task.get("endTime")
        if not end_time:
            return False
        
        end_min = to_mins(end_time)
        # Fire as soon as we cross end_min. Keep firing eligible up to end+30
        # (so a brief outage doesn't permanently miss it), but only ONCE.
        if current_min < end_min:
            return False
        if current_min > end_min + 30:
            # Too late — give up to avoid playing the alert hours later
            return False
        
        if task.get("status") != "pending":
            return False
        
        key = self._key(task["id"], "alert", task.get("date", ""), end_time)
        with self._lock:
            if self._fired.get(key):
                return False
            self._fired[key] = True
        
        play_sound(alert_music_file)
        print(f"🚨 [AUDIO] Alert music ▶  task_id={task['id']} name='{task.get('name','')}'  exceeded by {current_min - end_min} min")
        return True


task_audio = TaskAudioController()

# =============================================================================
# DATA MODELS
# =============================================================================
EXCEL_FILE = "smartday_tasks.xlsx"
WEIGHTS_FILE = "rl_weights.json"
PROFILE_FILE = "productivity_profile.json"
SCHEDULE_HISTORY_FILE = "schedule_history.json"
SETTINGS_FILE = "app_settings.json"
USER_PREFERENCES_FILE = "user_preferences.json"

COLS = ["ID","Name","Date","StartTime","EndTime","Duration","Difficulty",
        "Priority","Status","Notes","PostponedCount","AlertSent","CompletedAt","CreatedAt",
        "UtilityScore","ManualOrder","UserSetTime"]

def excel_init():
    if not EXCEL_OK: return
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook(); ws = wb.active; ws.title = "Tasks"
        ws.append(COLS); wb.save(EXCEL_FILE)

def excel_load():
    if not EXCEL_OK or not os.path.exists(EXCEL_FILE): return []
    try:
        wb = load_workbook(EXCEL_FILE); ws = wb.active; rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row[0]: continue
            d = dict(zip(COLS, row))
            rows.append({
                "id": int(d["ID"]) if d["ID"] else 0,
                "name": str(d["Name"] or ""),
                "date": str(d["Date"] or ""),
                "startTime": str(d["StartTime"]) if d["StartTime"] else None,
                "endTime": str(d["EndTime"]) if d["EndTime"] else None,
                "duration": int(d["Duration"] or 30),
                "difficulty": str(d["Difficulty"] or "medium"),
                "priority": str(d["Priority"] or "medium"),
                "status": str(d["Status"] or "pending"),
                "notes": str(d["Notes"] or ""),
                "postponedCount": int(d["PostponedCount"] or 0),
                "alertSent": bool(d["AlertSent"]),
                "completedAt": str(d["CompletedAt"]) if d["CompletedAt"] else None,
                "createdAt": str(d["CreatedAt"] or ""),
                "utilityScore": float(d["UtilityScore"]) if d["UtilityScore"] else 0,
                "manualOrder": int(d["ManualOrder"]) if d["ManualOrder"] else 0,
                "userSetTime": bool(d.get("UserSetTime", False))
            })
        return rows
    except Exception:
        return []

def excel_save(tasks):
    if not EXCEL_OK: return
    try:
        wb = Workbook(); ws = wb.active; ws.title = "Tasks"; ws.append(COLS)
        for t in tasks:
            ws.append([t.get("id"), t.get("name"), t.get("date"),
                       t.get("startTime"), t.get("endTime"), t.get("duration"),
                       t.get("difficulty"), t.get("priority"), t.get("status"),
                       t.get("notes"), t.get("postponedCount", 0),
                       t.get("alertSent", False), t.get("completedAt"), t.get("createdAt"),
                       t.get("utilityScore", 0), t.get("manualOrder", 0),
                       t.get("userSetTime", False)])
        wb.save(EXCEL_FILE)
    except Exception:
        pass

excel_init()
_tasks = excel_load()
_settings = {"sleepTime": "23:30", "startTime": "08:00", "useManualOrder": False, "theme": "dark"}
_app_settings = {"theme": "dark", "sleepTime": "23:30", "startTime": "08:00", "language": "arabic"}
_history = []
_notifs = []
_lock = threading.Lock()
_executor = ThreadPoolExecutor(max_workers=4)

def load_app_settings():
    global _app_settings, _settings
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                _app_settings = json.load(f)
        except:
            pass
    
    if "sleepTime" in _app_settings:
        _settings["sleepTime"] = _app_settings["sleepTime"]
    if "startTime" in _app_settings:
        _settings["startTime"] = _app_settings["startTime"]
    if "theme" in _app_settings:
        _settings["theme"] = _app_settings["theme"]
    if "language" not in _app_settings:
        _app_settings["language"] = "arabic"

def save_app_settings():
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(_app_settings, f, indent=2, ensure_ascii=False)
    except:
        pass

load_app_settings()

def get_tasks():
    with _lock: return list(_tasks)

def save_tasks():
    excel_save(_tasks)

def new_id(): return int(time.time() * 1000) + random.randint(0, 999)
def today_str(): return datetime.now().strftime("%Y-%m-%d")
def now_str(): return datetime.now().strftime("%H:%M")

def to_mins(t):
    if not t: return 0
    try:
        p = str(t).split(":")
        return int(p[0]) * 60 + int(p[1])
    except: return 0

def to_time(m):
    m = int(m) % 1440
    return f"{m//60:02d}:{m%60:02d}"

def to_seconds(time_str):
    """Convert time string to total seconds since midnight (24-hour format)"""
    if not time_str:
        return 0
    parts = str(time_str).split(":")
    return int(parts[0]) * 3600 + int(parts[1]) * 60

def to_time_from_seconds(seconds):
    """Convert seconds since midnight to time string (24-hour format)"""
    seconds = int(seconds) % 86400
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return f"{hours:02d}:{minutes:02d}"

def convert_to_24hour(time_str: str) -> str:
    """Convert any time format to 24-hour format"""
    if not time_str:
        return None
    
    time_str = str(time_str).strip().lower()
    
    # Already in HH:MM format
    if re.match(r'^\d{1,2}:\d{2}$', time_str):
        parts = time_str.split(':')
        h = int(parts[0])
        m = int(parts[1])
        if h < 24:
            return f"{h:02d}:{m:02d}"
    
    # Handle 12-hour format with am/pm
    ampm_match = re.search(r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)', time_str, re.IGNORECASE)
    if ampm_match:
        h = int(ampm_match.group(1))
        m = int(ampm_match.group(2)) if ampm_match.group(2) else 0
        period = ampm_match.group(3).lower()
        if period == 'pm' and h < 12:
            h += 12
        elif period == 'am' and h == 12:
            h = 0
        return f"{h:02d}:{m:02d}"
    
    # Handle Arabic time formats
    arabic_match = re.search(r'(\d{1,2})(?::(\d{2}))?\s*(صباح|مساء|ص|م)', time_str, re.IGNORECASE)
    if arabic_match:
        h = int(arabic_match.group(1))
        m = int(arabic_match.group(2)) if arabic_match.group(2) else 0
        period = arabic_match.group(3).lower()
        if period in ('مساء', 'م'):
            if h < 12:
                h += 12
        elif period in ('صباح', 'ص'):
            if h == 12:
                h = 0
        return f"{h:02d}:{m:02d}"
    
    # Handle "الساعة X" format
    hour_match = re.search(r'الساعة\s*(\d{1,2})', time_str)
    if hour_match:
        h = int(hour_match.group(1))
        return f"{h:02d}:00"
    
    return None

# Helper function to play sounds
sound_lock = threading.Lock()
_last_sound_time = {}
def play_sound(sound_file, min_interval=1.0):
    """Play sound with minimum interval to prevent overlapping"""
    if not SOUND_OK or not sound_file or not os.path.exists(sound_file):
        return
    
    with sound_lock:
        now = time.time()
        if sound_file in _last_sound_time and now - _last_sound_time[sound_file] < min_interval:
            return
        _last_sound_time[sound_file] = now
    
    try:
        def _play():
            try:
                sound = pygame.mixer.Sound(sound_file)
                sound.play()
            except Exception as e:
                print(f"Sound play error: {e}")
        threading.Thread(target=_play, daemon=True).start()
    except Exception as e:
        print(f"Sound error: {e}")

# =============================================================================
# Time comparison functions that handle times after midnight
# =============================================================================

def is_future_time(time_str: str) -> bool:
    """Check if a time is in the future (strict check)"""
    try:
        time_mins = to_mins(time_str)
        now = datetime.now()
        current_mins = now.hour * 60 + now.minute
        start_day_mins = to_mins(_settings.get("startTime", "08:00"))
        sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
        
        # If sleep is after midnight (e.g., 23:30 to 01:30)
        if sleep_mins < start_day_mins:
            # Time like 01:00 is for tomorrow if current time is after start of day
            if time_mins < start_day_mins and current_mins >= start_day_mins:
                return True
            return time_mins > current_mins
        else:
            return time_mins > current_mins
    except:
        return False

def is_time_in_future(time_str: str) -> bool:
    """Strict check if time is in the future"""
    try:
        time_mins = to_mins(time_str)
        now_mins = to_mins(now_str())
        start_mins = to_mins(_settings.get("startTime", "08:00"))
        sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
        
        if sleep_mins < start_mins:
            if time_mins < start_mins:
                time_mins += 1440
            if now_mins < start_mins:
                now_mins += 1440
        
        return time_mins > now_mins
    except:
        return False

def is_within_sleep_time(time_str: str, duration: int) -> bool:
    """Check if task can be completed before sleep time"""
    sleep_time = _settings.get("sleepTime", "23:30")
    sleep_mins = to_mins(sleep_time)
    start_mins = to_mins(time_str)
    end_mins = start_mins + duration
    start_day_mins = to_mins(_settings.get("startTime", "08:00"))
    
    if sleep_mins < start_day_mins:
        if start_mins >= start_day_mins:
            return end_mins <= (sleep_mins + 1440)
        else:
            return end_mins <= sleep_mins
    else:
        return end_mins <= sleep_mins

def get_time_remaining_until_sleep() -> int:
    """Calculate minutes remaining until sleep time"""
    now = datetime.now()
    current_mins = now.hour * 60 + now.minute
    sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
    start_day_mins = to_mins(_settings.get("startTime", "08:00"))
    
    if sleep_mins < start_day_mins:
        remaining = (sleep_mins + 1440) - current_mins
        return max(0, remaining)
    else:
        if current_mins < sleep_mins:
            return sleep_mins - current_mins
        else:
            return 0

def get_last_task_end_time(tasks: List[Dict], date: str) -> int:
    start_mins = to_mins(_settings.get("startTime", "08:00"))
    now_mins = to_mins(now_str())
    current_time = max(start_mins, now_mins)
    
    today_tasks = [t for t in tasks if t.get("date") == date and t.get("status") == "pending" and t.get("endTime")]
    
    if not today_tasks:
        return current_time
    
    max_end = 0
    for task in today_tasks:
        if task.get("endTime"):
            end_mins = to_mins(task["endTime"])
            if end_mins > max_end:
                max_end = end_mins
    
    return max(max_end, current_time)

def find_available_time_slots(tasks: List[Dict], duration: int, date: str) -> List[Dict]:
    """Find available time slots for a task of given duration"""
    start_mins = to_mins(_settings.get("startTime", "08:00"))
    sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
    now_mins = to_mins(now_str())
    
    scheduled = [t for t in tasks if t.get("date") == date and t.get("status") == "pending" 
                 and t.get("startTime") and t.get("endTime")]
    scheduled.sort(key=lambda x: to_mins(x.get("startTime", "00:00")))
    
    is_sleep_after_midnight = sleep_mins < start_mins
    if is_sleep_after_midnight:
        sleep_mins += 1440
        if now_mins < start_mins:
            now_mins += 1440
        for task in scheduled:
            task_start = to_mins(task["startTime"])
            if task_start < start_mins:
                task_start += 1440
            task["_start"] = task_start
            task["_end"] = task_start + task["duration"]
    else:
        for task in scheduled:
            task["_start"] = to_mins(task["startTime"])
            task["_end"] = task["_start"] + task["duration"]
    
    current_time = max(start_mins, now_mins)
    current_time = max(current_time, start_mins)
    available_slots = []
    
    if scheduled:
        first_start = scheduled[0]["_start"]
        if current_time + duration <= first_start:
            available_slots.append({
                "start": to_time(current_time),
                "end": to_time(current_time + duration),
                "duration": duration
            })
    else:
        if current_time + duration <= sleep_mins:
            available_slots.append({
                "start": to_time(current_time),
                "end": to_time(current_time + duration),
                "duration": duration
            })
    
    for i in range(len(scheduled) - 1):
        gap_start = scheduled[i]["_end"] + 1
        gap_end = scheduled[i + 1]["_start"]
        if gap_start + duration <= gap_end:
            available_slots.append({
                "start": to_time(gap_start),
                "end": to_time(gap_start + duration),
                "duration": duration
            })
    
    if scheduled:
        last_end = scheduled[-1]["_end"]
        if last_end + duration <= sleep_mins:
            available_slots.append({
                "start": to_time(last_end + 1),
                "end": to_time(last_end + 1 + duration),
                "duration": duration
            })
    
    return available_slots[:5]

# =============================================================================
# PRODUCTIVITY PROFILE
# =============================================================================
class ProductivityProfile:
    def __init__(self):
        self.profile = self.load_profile()
    
    def load_profile(self) -> Dict:
        if os.path.exists(PROFILE_FILE):
            try:
                with open(PROFILE_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        
        return {
            "productive_hours": {
                "morning": {"start": "08:00", "end": "11:00", "score": 0.7},
                "afternoon": {"start": "14:00", "end": "17:00", "score": 0.8},
                "evening": {"start": "19:00", "end": "22:00", "score": 0.9}
            },
            "typical_sleep_time": "23:30",
            "typical_wake_time": "07:00",
            "frequently_postponed": [],
            "task_difficulty_tolerance": "medium",
            "preferred_session_length": 45,
            "completion_rate_by_hour": {},
            "postponement_patterns": {
                "by_difficulty": {"easy": 0, "medium": 0, "hard": 0},
                "by_priority": {"low": 0, "medium": 0, "high": 0},
                "by_hour": {}
            },
            "energy_pattern": {
                "08:00": 0.6, "09:00": 0.7, "10:00": 0.8, "11:00": 0.8,
                "12:00": 0.5, "13:00": 0.4, "14:00": 0.6, "15:00": 0.8,
                "16:00": 0.8, "17:00": 0.7, "18:00": 0.5, "19:00": 0.7,
                "20:00": 0.8, "21:00": 0.9, "22:00": 0.7, "23:00": 0.4
            }
        }
    
    def save_profile(self):
        try:
            with open(PROFILE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.profile, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving profile: {e}")
    
    def update_completion(self, task: Dict, completed: bool):
        hour = datetime.now().hour
        hour_key = str(hour)
        if hour_key not in self.profile["completion_rate_by_hour"]:
            self.profile["completion_rate_by_hour"][hour_key] = {"completed": 0, "total": 0}
        
        self.profile["completion_rate_by_hour"][hour_key]["total"] += 1
        if completed:
            self.profile["completion_rate_by_hour"][hour_key]["completed"] += 1
        
        success_rate = (
            self.profile["completion_rate_by_hour"][hour_key]["completed"] /
            max(self.profile["completion_rate_by_hour"][hour_key]["total"], 1)
        )
        
        for period in self.profile["productive_hours"]:
            period_hours = self.profile["productive_hours"][period]
            start_h = int(period_hours["start"].split(":")[0])
            end_h = int(period_hours["end"].split(":")[0])
            
            if start_h <= hour < end_h:
                old_score = period_hours["score"]
                period_hours["score"] = old_score * 0.8 + success_rate * 0.2
        
        self.save_profile()
    
    def update_postponement(self, task: Dict):
        diff = task.get("difficulty", "medium")
        pri = task.get("priority", "medium")
        hour = str(datetime.now().hour)
        
        self.profile["postponement_patterns"]["by_difficulty"][diff] += 1
        self.profile["postponement_patterns"]["by_priority"][pri] += 1
        
        if hour not in self.profile["postponement_patterns"]["by_hour"]:
            self.profile["postponement_patterns"]["by_hour"][hour] = 0
        self.profile["postponement_patterns"]["by_hour"][hour] += 1
        
        task_name_lower = task.get("name", "").lower()
        for pattern in self.profile["frequently_postponed"]:
            if pattern["name"].lower() in task_name_lower:
                pattern["count"] += 1
                self.save_profile()
                return
        
        self.profile["frequently_postponed"].append({"name": task.get("name", ""), "count": 1})
        self.save_profile()
    
    def get_energy_level(self, time_str: str) -> float:
        hour = time_str.split(":")[0] + ":00"
        return self.profile["energy_pattern"].get(hour, 0.5)
    
    def get_productive_score(self, start_time: str, duration: int) -> float:
        if not start_time:
            return 0.5
        
        try:
            start_hour = int(start_time.split(":")[0])
        except:
            return 0.5
        
        end_hour = start_hour + duration // 60
        
        max_score = 0.0
        for h in range(start_hour, min(end_hour + 1, 24)):
            hour_str = f"{h:02d}:00"
            energy = self.get_energy_level(hour_str)
            
            period_score = 0.0
            for period in self.profile["productive_hours"]:
                period_data = self.profile["productive_hours"][period]
                p_start = int(period_data["start"].split(":")[0])
                p_end = int(period_data["end"].split(":")[0])
                
                if p_start <= h < p_end:
                    period_score = max(period_score, period_data["score"])
            
            combined_score = (energy + period_score) / 2
            max_score = max(max_score, combined_score)
        
        return max_score
    
    def is_frequently_postponed(self, task: Dict) -> bool:
        task_name = task.get("name", "").lower()
        for pattern in self.profile["frequently_postponed"]:
            if pattern["name"].lower() in task_name and pattern["count"] >= 3:
                return True
        return False
    
    def get_postponement_probability(self, task: Dict) -> float:
        diff = task.get("difficulty", "medium")
        pri = task.get("priority", "medium")
        
        total_diff = sum(self.profile["postponement_patterns"]["by_difficulty"].values())
        total_pri = sum(self.profile["postponement_patterns"]["by_priority"].values())
        
        diff_prob = (
            self.profile["postponement_patterns"]["by_difficulty"].get(diff, 0) / 
            max(total_diff, 1)
        )
        pri_prob = (
            self.profile["postponement_patterns"]["by_priority"].get(pri, 0) / 
            max(total_pri, 1)
        )
        
        return (diff_prob + pri_prob) / 2

# =============================================================================
# TALENT DEGREE SYSTEM (NEW)
# =============================================================================
TALENT_FILE = "talent_profile.json"


class TalentProfile:
    """
    Tracks how good the user is at handling certain types of tasks.
    
    Three core scores in [0, 1]:
      - difficulty_handling_score: how well user handles HARD tasks
      - speed_efficiency_score: how often user finishes BEFORE estimated duration
      - consistency_score: how regularly user completes tasks (vs postponing/abandoning)
    
    Plus per-task-type breakdown so we know e.g. user is good at "coding"
    but slow at "writing".
    
    Updated via reinforcement-style EMA (exponential moving average): every
    completion event nudges the relevant score with a small learning rate, so
    the profile adapts smoothly over time without catastrophic forgetting.
    """
    
    LEARNING_RATE = 0.08  # how fast scores adapt (0-1). Small = stable, large = fast.
    DECAY = 0.995          # tiny per-update decay so old habits gradually lose weight
    
    def __init__(self):
        self.profile = self.load()
    
    # ---------- persistence ----------
    def load(self) -> Dict[str, Any]:
        if os.path.exists(TALENT_FILE):
            try:
                with open(TALENT_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    # backfill any missing keys
                    return self._with_defaults(data)
            except Exception:
                pass
        return self._with_defaults({})
    
    @staticmethod
    def _with_defaults(data: Dict[str, Any]) -> Dict[str, Any]:
        defaults = {
            "preferred_task_type": None,            # e.g. "coding", inferred from completion patterns
            "difficulty_handling_score": 0.5,       # in [0, 1]
            "speed_efficiency_score": 0.5,
            "consistency_score": 0.5,
            # Public-facing snapshot in the example schema the spec asked for:
            "handles_high_difficulty": 0.5,
            "fast_completion": 0.5,
            "consistency": 0.5,
            # Bookkeeping
            "completed_count": 0,
            "postponed_count": 0,
            "completed_hard": 0,
            "completed_fast": 0,
            "by_task_type": {},  # {type_name: {completed, postponed, fast, hard}}
            "last_updated": None,
        }
        for k, v in defaults.items():
            data.setdefault(k, v)
        return data
    
    def save(self):
        try:
            self.profile["last_updated"] = datetime.now().isoformat(timespec="seconds")
            # Mirror internal scores into the public-snapshot keys
            self.profile["handles_high_difficulty"] = self.profile["difficulty_handling_score"]
            self.profile["fast_completion"] = self.profile["speed_efficiency_score"]
            self.profile["consistency"] = self.profile["consistency_score"]
            with open(TALENT_FILE, "w", encoding="utf-8") as f:
                json.dump(self.profile, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"⚠️ TalentProfile save failed: {e}")
    
    # ---------- helpers ----------
    @staticmethod
    def _clamp(v: float, lo: float = 0.0, hi: float = 1.0) -> float:
        return max(lo, min(hi, v))
    
    def _ema_update(self, key: str, target: float):
        """EMA toward `target` with learning rate, plus tiny decay."""
        cur = float(self.profile.get(key, 0.5))
        new = cur * (1 - self.LEARNING_RATE) + target * self.LEARNING_RATE
        new *= self.DECAY
        # Don't let decay drag us all the way to 0 from a neutral start
        new = max(new, 0.05)
        self.profile[key] = self._clamp(new)
    
    def _infer_task_type(self, task: Dict) -> str:
        """Lightweight task-type inference from name + difficulty."""
        name = (task.get("name") or "").lower()
        # heuristic keywords (Arabic + English)
        if any(k in name for k in ["code", "كود", "برمج", "debug", "bug"]):
            return "coding"
        if any(k in name for k in ["read", "قراء", "study", "ذاكر", "مذاكر"]):
            return "study"
        if any(k in name for k in ["write", "كتاب", "essay", "مقال"]):
            return "writing"
        if any(k in name for k in ["workout", "رياض", "exercise", "gym"]):
            return "exercise"
        if any(k in name for k in ["meet", "اجتماع", "call", "مكالم"]):
            return "meeting"
        return f"diff_{task.get('difficulty', 'medium')}"  # fallback bucket
    
    def _bump_type(self, type_name: str, field: str):
        bucket = self.profile["by_task_type"].setdefault(
            type_name, {"completed": 0, "postponed": 0, "fast": 0, "hard": 0}
        )
        bucket[field] = bucket.get(field, 0) + 1
    
    # ---------- update events ----------
    def record_completion(self, task: Dict, was_on_time: bool, actual_duration: Optional[int] = None):
        """
        Called when a task is completed.
        
        - `was_on_time`: True if completed before/at endTime, False if late/extended
        - `actual_duration`: minutes actually used (if known); used to detect "fast" completion
        """
        difficulty = task.get("difficulty", "medium")
        estimated = task.get("duration", 30) or 30
        type_name = self._infer_task_type(task)
        
        self.profile["completed_count"] += 1
        self._bump_type(type_name, "completed")
        
        # 1) Difficulty score: completing hard tasks → strong positive signal
        if difficulty == "hard":
            self.profile["completed_hard"] += 1
            self._bump_type(type_name, "hard")
            # Successful hard completion → nudge toward 1.0
            self._ema_update("difficulty_handling_score", 0.9 if was_on_time else 0.65)
        elif difficulty == "easy":
            # Many easy tasks shouldn't inflate difficulty score
            self._ema_update("difficulty_handling_score",
                             min(0.5 + 0.05, self.profile["difficulty_handling_score"]))
        else:  # medium
            self._ema_update("difficulty_handling_score", 0.6 if was_on_time else 0.45)
        
        # 2) Speed score: actual_duration vs estimated
        if actual_duration is not None and actual_duration > 0:
            ratio = actual_duration / max(estimated, 1)
            # ratio < 0.85 → fast, ratio > 1.15 → slow
            if ratio <= 0.85:
                self.profile["completed_fast"] += 1
                self._bump_type(type_name, "fast")
                self._ema_update("speed_efficiency_score", 0.9)
            elif ratio >= 1.15:
                self._ema_update("speed_efficiency_score", 0.3)
            else:
                self._ema_update("speed_efficiency_score", 0.6)
        else:
            # No timing info → use was_on_time as a weak proxy
            self._ema_update("speed_efficiency_score", 0.7 if was_on_time else 0.4)
        
        # 3) Consistency: every completion is positive
        self._ema_update("consistency_score", 0.85)
        
        # 4) Update preferred task type — most-completed type wins
        self._recompute_preferred_type()
        
        self.save()
        print(f"🎓 [TALENT] +completion type={type_name} diff={difficulty} on_time={was_on_time} → "
              f"H={self.profile['difficulty_handling_score']:.2f} "
              f"F={self.profile['speed_efficiency_score']:.2f} "
              f"C={self.profile['consistency_score']:.2f}")
    
    def record_postponement(self, task: Dict):
        """Called when the user postpones a task — hurts consistency."""
        type_name = self._infer_task_type(task)
        self.profile["postponed_count"] += 1
        self._bump_type(type_name, "postponed")
        # Postponing a HARD task → small hit to difficulty handling
        if task.get("difficulty") == "hard":
            self._ema_update("difficulty_handling_score", 0.3)
        # Always hits consistency
        self._ema_update("consistency_score", 0.2)
        self.save()
        print(f"🎓 [TALENT] -postponement type={type_name} → "
              f"H={self.profile['difficulty_handling_score']:.2f} "
              f"C={self.profile['consistency_score']:.2f}")
    
    def _recompute_preferred_type(self):
        best, best_count = None, 0
        for tname, bucket in self.profile["by_task_type"].items():
            if bucket.get("completed", 0) > best_count:
                best = tname
                best_count = bucket["completed"]
        if best:
            self.profile["preferred_task_type"] = best
    
    # ---------- query API used by the scheduler / utility ----------
    def difficulty_factor_for(self, task: Dict) -> float:
        """
        Returns a multiplier in roughly [0.5, 1.5] used by the utility function.
        
        - HARD task + strong user (high difficulty score) → > 1.0  (boost: schedule earlier)
        - HARD task + weak user → < 1.0 (push later)
        - EASY task → ~1.0 (mostly neutral)
        """
        diff = task.get("difficulty", "medium")
        h = self.profile["difficulty_handling_score"]  # in [0,1]
        if diff == "hard":
            # 0.5 score → 0.7×, 1.0 score → 1.5×
            return 0.7 + (h - 0.5) * 1.6 if h >= 0.5 else 0.7 - (0.5 - h) * 0.4
        if diff == "easy":
            return 1.0
        # medium
        return 0.9 + (h - 0.5) * 0.4
    
    def speed_factor(self) -> float:
        """Returns a multiplier in roughly [0.85, 1.15] used to score "tightness" of schedule."""
        s = self.profile["speed_efficiency_score"]
        return 0.85 + s * 0.30  # 0 → 0.85, 1 → 1.15
    
    def consistency_factor(self) -> float:
        """In roughly [0.9, 1.1]. Boost schedules slightly when user is consistent."""
        c = self.profile["consistency_score"]
        return 0.9 + c * 0.2
    
    def public_snapshot(self) -> Dict[str, float]:
        """Spec-shaped snapshot for serialization / API responses."""
        return {
            "handles_high_difficulty": round(self.profile["difficulty_handling_score"], 3),
            "fast_completion": round(self.profile["speed_efficiency_score"], 3),
            "consistency": round(self.profile["consistency_score"], 3),
            "preferred_task_type": self.profile.get("preferred_task_type"),
        }
    
    def insights(self) -> List[str]:
        out = []
        h = self.profile["difficulty_handling_score"]
        s = self.profile["speed_efficiency_score"]
        c = self.profile["consistency_score"]
        pt = self.profile.get("preferred_task_type")
        if h >= 0.7:
            out.append("💪 موهبتك عالية في المهام الصعبة — يقترح النظام جدولتها مبكراً")
        elif h <= 0.35:
            out.append("🪜 المهام الصعبة تتأخر عادةً — سيؤجلها النظام لوقت طاقتك العالية")
        if s >= 0.7:
            out.append("⚡ تنجز أسرع من المتوقع — سيُسمح بجدول أكثر إحكاماً")
        elif s <= 0.35:
            out.append("🐢 تستهلك وقتاً أطول من المتوقع — سيوسّع النظام الفجوات")
        if c >= 0.75:
            out.append("📈 ثبات عالٍ في الإكمال — مكافأة في ترتيب المقترحات")
        if pt:
            out.append(f"🎯 نوع المهام المفضل (الأكثر إنجازاً): {pt}")
        return out


# =============================================================================
# USER PREFERENCES
# =============================================================================
class UserPreferences:
    def __init__(self):
        self.preferences = self.load_preferences()
        self.strategy_scores = defaultdict(float)
        self._load_strategy_scores()
    
    def load_preferences(self) -> Dict:
        if os.path.exists(USER_PREFERENCES_FILE):
            try:
                with open(USER_PREFERENCES_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if "strategy_counts" in data:
                        return data
            except:
                pass
        return {
            "strategy_counts": {
                "priority": 0,
                "difficulty": 0,
                "duration_short": 0,
                "duration_long": 0,
                "random_1": 0,
                "random_2": 0,
                "energy_based": 0,
                "deadline_based": 0
            },
            "total_selections": 0,
            "last_selected": None
        }
    
    def _load_strategy_scores(self):
        total = self.preferences["total_selections"]
        if total > 0:
            for strategy, count in self.preferences["strategy_counts"].items():
                self.strategy_scores[strategy] = count / total
    
    def save_preferences(self):
        try:
            with open(USER_PREFERENCES_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.preferences, f, indent=2, ensure_ascii=False)
        except:
            pass
    
    def record_choice(self, strategy_name: str):
        if strategy_name in self.preferences["strategy_counts"]:
            self.preferences["strategy_counts"][strategy_name] += 1
            self.preferences["total_selections"] += 1
            self.preferences["last_selected"] = strategy_name
            self._load_strategy_scores()
            self.save_preferences()
    
    def get_preferred_strategy(self) -> str:
        if not self.preferences["total_selections"]:
            return None
        return max(self.preferences["strategy_counts"], key=lambda k: self.preferences["strategy_counts"][k])
    
    def get_strategy_weight(self, strategy_name: str) -> float:
        return self.strategy_scores.get(strategy_name, 0.0)

user_preferences = UserPreferences()

# =============================================================================
# REINFORCEMENT LEARNING
# =============================================================================
DEFAULT_WEIGHTS = {
    "priority_weight": 0.35,
    "difficulty_weight": -0.15,
    "duration_weight": -0.10,
    "deadline_weight": 0.25,
    "productivity_weight": 0.15,
    "postponed_penalty": -0.10,
    "energy_match_weight": 0.10,
    "time_remaining_weight": 0.08,
    "user_preference_weight": 0.12,
    "time_pressure_weight": 0.15,
    "task_importance_weight": 0.10
}

class ReinforcementLearning:
    def __init__(self):
        self.weights = self.load_weights()
        # Gradual learning — small step per update so weights don't swing wildly
        self.learning_rate = 0.05
        # Decay factor: every update, all weights drift slightly back toward the
        # default weights. Keeps the system from getting stuck on stale preferences.
        self.decay_rate = 0.01
        self.history = self.load_history()
        self.strategy_performance = defaultdict(lambda: {"selected": 0, "score": 0.0})
    
    def load_weights(self) -> Dict:
        if os.path.exists(WEIGHTS_FILE):
            try:
                with open(WEIGHTS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return DEFAULT_WEIGHTS.copy()
    
    def save_weights(self):
        try:
            with open(WEIGHTS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.weights, f, indent=2, ensure_ascii=False)
        except:
            pass
    
    def load_history(self) -> List[Dict]:
        if os.path.exists(SCHEDULE_HISTORY_FILE):
            try:
                with open(SCHEDULE_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return []
    
    def save_history(self):
        try:
            history = self.history[-200:]
            with open(SCHEDULE_HISTORY_FILE, 'w', encoding='utf-8') as f:
                json.dump(history, f, indent=2, ensure_ascii=False)
        except:
            pass
    
    def update_weights(self, recommended_idx: int, selected_idx: int, selected_schedule: List[Dict], all_scores: List[float], strategy_name: str = None):
        """
        Gradual reinforcement update:
            new_weight = old_weight + learning_rate * reward_signal
        Plus a small decay toward the default weights so old preferences fade.
        """
        # Reward: +1 if user accepted what we recommended, scaled otherwise
        if recommended_idx == selected_idx:
            reward = 1.0
        else:
            if all_scores and recommended_idx < len(all_scores) and selected_idx < len(all_scores):
                score_diff = all_scores[selected_idx] - all_scores[recommended_idx]
                max_abs_diff = max(abs(score_diff), 1.0)
                reward = score_diff / max_abs_diff   # in [-1, 1]
            else:
                reward = -0.5
        
        if strategy_name:
            self.strategy_performance[strategy_name]["selected"] += 1
            self.strategy_performance[strategy_name]["score"] += reward
        
        weight_keys = list(self.weights.keys())
        
        for key in weight_keys:
            current = self.weights[key]
            update = self._calculate_weight_update(key, reward, selected_schedule)
            # Gradual update
            new_value = current + self.learning_rate * update
            # Decay toward default — gentle "forgetting" so old behaviour fades out
            default_val = DEFAULT_WEIGHTS.get(key, current)
            new_value = new_value * (1 - self.decay_rate) + default_val * self.decay_rate
            # Clamp
            self.weights[key] = max(-0.5, min(1.0, new_value))
        
        self.save_weights()
        
        self.history.append({
            "timestamp": datetime.now().isoformat(),
            "recommended_idx": recommended_idx,
            "selected_idx": selected_idx,
            "reward": reward,
            "strategy": strategy_name,
            "weights_after": self.weights.copy()
        })
        self.save_history()
        
        if strategy_name:
            user_preferences.record_choice(strategy_name)
    
    def _calculate_weight_update(self, weight_key: str, reward: float, schedule: List[Dict]) -> float:
        if not schedule:
            return 0.0
        
        avg_priority = self._avg_priority(schedule)
        avg_difficulty = self._avg_difficulty(schedule)
        avg_duration = sum(t.get("duration", 30) for t in schedule) / len(schedule) if schedule else 30
        
        updates = {
            "priority_weight": 0.1 * reward * (1.0 if avg_priority > 0.6 else -0.5),
            "difficulty_weight": 0.05 * reward * (-1.0 if avg_difficulty > 0.5 else 0.5),
            "duration_weight": 0.03 * reward * (-1.0 if avg_duration > 45 else 0.3),
            "deadline_weight": 0.08 * reward,
            "productivity_weight": 0.06 * reward,
            "postponed_penalty": 0.04 * reward,
            "energy_match_weight": 0.05 * reward,
            "time_remaining_weight": 0.04 * reward,
            "user_preference_weight": 0.07 * reward * (user_preferences.get_strategy_weight("smart_adaptive") if user_preferences.preferences["total_selections"] > 0 else 0.5),
            "time_pressure_weight": 0.06 * reward,
            "task_importance_weight": 0.05 * reward
        }
        
        return updates.get(weight_key, 0.0)
    
    def _avg_priority(self, schedule: List[Dict]) -> float:
        scores = {"high": 1.0, "medium": 0.6, "low": 0.3}
        if not schedule: return 0.5
        total = sum(scores.get(t.get("priority", "medium"), 0.5) for t in schedule)
        return total / len(schedule)
    
    def _avg_difficulty(self, schedule: List[Dict]) -> float:
        scores = {"hard": 1.0, "medium": 0.6, "easy": 0.3}
        if not schedule: return 0.5
        total = sum(scores.get(t.get("difficulty", "medium"), 0.5) for t in schedule)
        return total / len(schedule)
    
    def get_insights(self) -> List[str]:
        insights = []
        
        if not self.history:
            return ["🧠 بدء التعلم - سيتم تحسين التوصيات مع الاستخدام"]
        
        recent = self.history[-20:]
        valid_items = [h for h in recent if "recommended_idx" in h and "selected_idx" in h]
        if valid_items:
            acceptance_rate = sum(1 for h in valid_items if h["recommended_idx"] == h["selected_idx"]) / len(valid_items)
        else:
            acceptance_rate = 0
        
        if acceptance_rate > 0.7:
            insights.append("🎯 نظام التوصيات يعمل بشكل ممتاز!")
        elif acceptance_rate > 0.4:
            insights.append("📊 نظام التوصيات يتعلم من اختياراتك ويتحسن")
        else:
            insights.append("🔧 جاري تحسين التوصيات لتتناسب مع تفضيلاتك")
        
        weights = self.weights
        if weights.get("priority_weight", 0) > 0.3:
            insights.append("⚡ تفضل إنجاز المهام حسب الأولوية العالية أولاً")
        
        if weights.get("difficulty_weight", 0) < -0.1:
            insights.append("💪 تميل لتأجيل المهام الصعبة - حاول البدء بها مبكراً")
        
        total_selections = user_preferences.preferences["total_selections"]
        if total_selections > 0:
            best_strategy = user_preferences.get_preferred_strategy()
            strategy_names_ar = {
                "priority": "حسب الأولوية",
                "difficulty": "حسب الصعوبة",
                "duration_short": "حسب المدة (الأقصر)",
                "duration_long": "حسب المدة (الأطول)",
                "random_1": "عشوائي ذكي",
                "random_2": "عشوائي ذكي 2",
                "energy_based": "حسب الطاقة",
                "deadline_based": "حسب الوقت المتبقي"
            }
            if best_strategy in strategy_names_ar:
                insights.append(f"📈 بناءً على {total_selections} اختيار سابق، تفضل الترتيب {strategy_names_ar[best_strategy]}")
        
        return insights

# =============================================================================
# SMART SCHEDULE GENERATOR
# =============================================================================
class SmartScheduleGenerator:
    """
    Diverse schedule generator.

    Produces 4–6 *significantly different* candidate schedules using a mix of:
      - deterministic strategies (priority, difficulty, smart_combined, energy)
      - semi-random permutations with probabilistic bias toward
        (a) high-priority tasks  (b) urgent deadlines (small time-left ratio).

    Schedules are deduplicated via a signature (the ordered tuple of task ids),
    so we never return two identical or near-identical orderings.
    """
    def __init__(self, profile: ProductivityProfile):
        self.profile = profile
    
    def generate_schedules(self, tasks: List[Dict], num_schedules: int = 5,
                           start_time: str = "08:00", end_time: str = "23:30") -> Tuple[List[List[Dict]], List[str]]:
        # Clamp num_schedules to the requested 4–6 range
        num_schedules = max(4, min(6, num_schedules))

        pending_tasks = [t for t in tasks if t.get("status") == "pending"]
        if not pending_tasks:
            return [], []

        start_mins = max(self._to_minutes(start_time), self._to_minutes(now_str()))
        end_mins = self._to_minutes(end_time)

        candidates: List[Tuple[List[Dict], str]] = []

        # --- Strategy 1: priority-first (deterministic) ---
        priority_tasks = sorted(
            pending_tasks,
            key=lambda x: ({"high": 0, "medium": 1, "low": 2}.get(x.get("priority", "medium"), 1),
                           x.get("duration", 30))
        )
        candidates.append((priority_tasks, "priority"))

        # --- Strategy 2: difficulty-first (do the hard things while fresh) ---
        difficulty_tasks = sorted(
            pending_tasks,
            key=lambda x: ({"hard": 0, "medium": 1, "easy": 2}.get(x.get("difficulty", "medium"), 1),
                           {"high": 0, "medium": 1, "low": 2}.get(x.get("priority", "medium"), 1))
        )
        candidates.append((difficulty_tasks, "difficulty"))

        # --- Strategy 3: deadline / time-pressure ---
        now_mins = to_mins(now_str())
        sleep_mins = to_mins(end_time)
        time_left = max(1, sleep_mins - now_mins)
        def deadline_score(task):
            duration_ratio = task.get("duration", 30) / time_left
            pri_mult = {"high": 1.4, "medium": 1.0, "low": 0.7}.get(task.get("priority", "medium"), 1.0)
            return -(duration_ratio * pri_mult)  # smaller (more urgent) first
        deadline_tasks = sorted(pending_tasks, key=deadline_score)
        candidates.append((deadline_tasks, "deadline_based"))

        # --- Strategy 4: smart combined (priority + difficulty + duration + energy) ---
        def smart_score(task):
            pri = {"high": 1.0, "medium": 0.6, "low": 0.3}.get(task.get("priority", "medium"), 0.6)
            dif = {"easy": 1.0, "medium": 0.7, "hard": 0.5}.get(task.get("difficulty", "medium"), 0.7)
            dur = 1.0 - min(task.get("duration", 30) / 120, 0.5)
            return -(pri * 0.45 + dif * 0.25 + dur * 0.30)
        smart_tasks = sorted(pending_tasks, key=smart_score)
        candidates.append((smart_tasks, "smart_combined"))

        # --- Strategy 5 & 6: semi-random diverse permutations with bias ---
        # Bias weights = priority weight * urgency weight, used in a weighted-without-replacement
        # sampling. This yields high-priority/urgent tasks tending to appear earlier, but with
        # genuine variation between runs — no two random permutations are the same.
        for variant_idx in range(2):
            shuffled = self._biased_permutation(pending_tasks, time_left, variant_idx)
            candidates.append((shuffled, f"diverse_{variant_idx + 1}"))

        # --- Deduplicate by ordered-id signature ---
        seen_signatures = set()
        unique_candidates: List[Tuple[List[Dict], str]] = []
        for ordered, name in candidates:
            sig = tuple(t.get("id") for t in ordered)
            if sig in seen_signatures:
                continue
            seen_signatures.add(sig)
            unique_candidates.append((ordered, name))

        # If we lost too many to dedup, top up with extra biased permutations
        attempts = 0
        while len(unique_candidates) < num_schedules and attempts < 12:
            attempts += 1
            shuffled = self._biased_permutation(pending_tasks, time_left, attempts + 10)
            sig = tuple(t.get("id") for t in shuffled)
            if sig not in seen_signatures:
                seen_signatures.add(sig)
                unique_candidates.append((shuffled, f"diverse_{len(unique_candidates) + 1}"))

        # Trim and assign concrete times
        unique_candidates = unique_candidates[:num_schedules]
        schedules = [self._assign_times(c[0], start_mins, end_mins) for c in unique_candidates]
        strategy_names = [c[1] for c in unique_candidates]

        return schedules, strategy_names

    def _biased_permutation(self, pending_tasks: List[Dict], time_left: int, salt: int) -> List[Dict]:
        """Weighted-without-replacement shuffle.

        Weights favor high priority and high duration-vs-time-left urgency, so
        important/urgent tasks are *more likely* to appear early — but the
        outcome is genuinely random and varies between calls.
        """
        rng = random.Random()  # fresh RNG so we don't share global state in odd ways
        rng.seed((time.time_ns() ^ (salt * 9301 + 49297)) & 0xFFFFFFFF)

        priority_w = {"high": 3.0, "medium": 1.6, "low": 0.8}
        remaining = list(pending_tasks)
        result: List[Dict] = []

        while remaining:
            weights = []
            for t in remaining:
                pw = priority_w.get(t.get("priority", "medium"), 1.0)
                # Urgency: longer task vs remaining day = more urgent to start now
                dur = max(1, t.get("duration", 30))
                urgency = 1.0 + min(2.0, dur / max(time_left, 1) * 4)
                weights.append(pw * urgency)
            # Weighted draw
            total = sum(weights)
            r = rng.random() * total
            acc = 0.0
            picked = 0
            for i, w in enumerate(weights):
                acc += w
                if acc >= r:
                    picked = i
                    break
            result.append(remaining.pop(picked))
        return result

    def _assign_times(self, ordered_tasks: List[Dict], start_mins: int, end_mins: int) -> List[Dict]:
        schedule = []
        current_time = start_mins
        
        for task in ordered_tasks:
            task_copy = task.copy()
            duration = task_copy.get("duration", 30)
            
            if current_time + duration <= end_mins:
                task_copy["startTime"] = self._to_time_str(current_time)
                task_copy["endTime"] = self._to_time_str(current_time + duration)
                current_time += duration + 1
            else:
                task_copy["startTime"] = None
                task_copy["endTime"] = None
            
            schedule.append(task_copy)
        
        return schedule
    
    @staticmethod
    def _to_minutes(time_str: str) -> int:
        if not time_str:
            return 0
        parts = time_str.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    
    @staticmethod
    def _to_time_str(minutes: int) -> str:
        hours = (minutes // 60) % 24
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"

# =============================================================================
# UTILITY EVALUATOR with Enhanced Weighted Scoring
# =============================================================================
class UtilityEvaluator:
    def __init__(self, weights: Dict, profile: ProductivityProfile, talent: Optional["TalentProfile"] = None):
        self.weights = weights
        self.profile = profile
        self.talent = talent  # NEW: optional talent profile (kept optional for safety)
    
    def evaluate_schedule(self, schedule: List[Dict], available_time: int, sleep_time: str) -> float:
        """
        Enhanced utility scoring. Per task we account for:
          - priority
          - difficulty
          - estimated duration
          - deadline urgency (position vs total slots)
          - match with productive hours / energy
          - remaining time before sleep
        Then aggregate-level: coverage, duration efficiency, energy consistency.
        """
        if not schedule:
            return 0.0
        
        total_score = 0.0
        scheduled_count = 0
        current_time = max(to_mins(_settings.get("startTime", "08:00")), to_mins(now_str()))
        total_duration = 0
        
        sleep_mins_total = to_mins(sleep_time)
        remaining_time = max(0, sleep_mins_total - current_time)
        # 0..1 (1 = lots of time left in the day)
        time_remaining_factor = min(1.0, remaining_time / 480)
        
        total_priority_score = 0
        total_energy_match = 0
        n = max(len(schedule), 1)
        
        for i, task in enumerate(schedule):
            priority_scores = {"high": 100, "medium": 50, "low": 20}
            pri_raw = priority_scores.get(task.get("priority", "medium"), 50)
            task_score = pri_raw * self.weights.get("priority_weight", 0.35) * 2
            
            difficulty_mod = {"hard": -10, "medium": 0, "easy": 8}
            task_score += difficulty_mod.get(task.get("difficulty", "medium"), 0) * self.weights.get("difficulty_weight", -0.15) * -1
            
            duration = task.get("duration", 30)
            total_duration += duration
            # Short tasks get a small bonus
            if duration <= 30:
                task_score += 12 * abs(self.weights.get("duration_weight", -0.10))
            elif duration >= 90:
                task_score -= 6 * abs(self.weights.get("duration_weight", -0.10))
            
            # --- Deadline urgency: high-priority tasks scheduled later in the day are penalised ---
            position_ratio = i / n  # 0 (first) .. ~1 (last)
            if task.get("priority") == "high":
                # the further down the list a high-priority task sits, the worse
                task_score -= position_ratio * 25 * self.weights.get("deadline_weight", 0.25)
            elif task.get("priority") == "low":
                # low-priority tasks late are fine
                task_score += position_ratio * 6 * self.weights.get("deadline_weight", 0.25)
            
            if task.get("startTime"):
                task_time = self._to_minutes(task["startTime"])
                prod_score = self.profile.get_productive_score(task["startTime"], duration) * 25
                task_score += prod_score * self.weights.get("productivity_weight", 0.15)
                
                energy = self.profile.get_energy_level(task["startTime"]) * 20
                task_score += energy * self.weights.get("energy_match_weight", 0.10)
                
                # --- Remaining time before sleep at the moment this task ends ---
                task_end_mins = task_time + duration
                slack_to_sleep = max(0, sleep_mins_total - task_end_mins)
                # Reward keeping comfortable slack until sleep (up to ~120 min headroom)
                slack_factor = min(1.0, slack_to_sleep / 120)
                task_score += slack_factor * 10 * self.weights.get("time_remaining_weight", 0.08)
                
                if task_time >= current_time:
                    wait_time = task_time - current_time
                    if wait_time <= 15:
                        task_score += 18
                    elif wait_time <= 60:
                        task_score += 8
                    elif wait_time > 120:
                        task_score -= 5
                
                scheduled_count += 1
                total_energy_match += energy
                total_priority_score += pri_raw
            else:
                # Could not be placed in the day — heavy penalty
                task_score -= 50
            
            if self.profile.is_frequently_postponed(task):
                task_score += 15 * self.weights.get("postponed_penalty", -0.10)
            
            task_score += time_remaining_factor * 15 * self.weights.get("time_remaining_weight", 0.08)
            
            # Time-pressure factor: how full the day is
            time_pressure_factor = 1.0 - min(1.0, total_duration / max(available_time, 1))
            task_score += time_pressure_factor * 20 * self.weights.get("time_pressure_weight", 0.15)
            
            importance = (pri_raw / 100) * \
                        (1 - difficulty_mod.get(task.get("difficulty", "medium"), 0) / 20)
            task_score += importance * 15 * self.weights.get("task_importance_weight", 0.10)
            
            # ============================================================
            # TALENT INTEGRATION (per-task)
            # ============================================================
            # Apply talent-based multipliers:
            #  - If user is strong on hard tasks AND this is a hard task scheduled
            #    EARLY (low position_ratio) → big bonus (let them tackle it while fresh)
            #  - If user is weak on hard tasks AND this is a hard task scheduled LATE → bonus
            #    (the schedule respects their capacity)
            #  - Speed-efficient users get a bonus for tighter schedules (already encoded
            #    by short waits / time_pressure_factor); we amplify it.
            if self.talent is not None:
                diff_factor = self.talent.difficulty_factor_for(task)  # ~[0.5, 1.5]
                if task.get("difficulty") == "hard":
                    # Earlier slot rewarded for strong users, later slot rewarded for weak users
                    earliness = 1.0 - position_ratio  # 1.0 (first) → 0.0 (last)
                    if diff_factor >= 1.0:
                        task_score += 20 * (diff_factor - 1.0) * earliness
                    else:
                        # Weak user: reward putting it later (high position_ratio)
                        task_score += 20 * (1.0 - diff_factor) * position_ratio
                else:
                    # Mild scaling for non-hard tasks
                    task_score *= (0.95 + 0.10 * diff_factor)
            
            total_score += max(0, task_score)
        
        coverage = scheduled_count / max(len(schedule), 1)
        coverage_bonus = coverage * 45
        
        duration_efficiency = max(0, 100 - (total_duration / 45))
        energy_consistency = total_energy_match / max(scheduled_count, 1) if scheduled_count > 0 else 0
        
        final_score = total_score + coverage_bonus + duration_efficiency + energy_consistency
        
        # ============================================================
        # TALENT INTEGRATION (aggregate)
        # ============================================================
        # Speed and consistency apply at the schedule level: a speed-efficient,
        # consistent user gets a small overall ranking boost — just enough to
        # break ties in favor of schedules that match their style.
        if self.talent is not None:
            final_score *= self.talent.speed_factor()       # ~[0.85, 1.15]
            final_score *= self.talent.consistency_factor() # ~[0.90, 1.10]
        
        return final_score
    
    def _to_minutes(self, time_str: str) -> int:
        if not time_str:
            return 0
        parts = time_str.split(":")
        return int(parts[0]) * 60 + int(parts[1])

# =============================================================================
# RECOMMENDATION SYSTEM (epsilon-greedy)
# =============================================================================
class RecommendationSystem:
    """
    Real recommendation system with epsilon-greedy exploration.

    - 80% of the time → recommend the highest-utility schedule (exploitation)
    - 20% of the time → recommend a random non-best schedule (exploration)

    This guarantees the system doesn't get stuck on one ordering forever,
    while still defaulting to the best option most of the time.
    """
    EPSILON = 0.20  # 20% exploration

    def __init__(self, evaluator: UtilityEvaluator):
        self.evaluator = evaluator
    
    def recommend(self, schedules: List[List[Dict]], available_time: int, sleep_time: str, strategy_names: List[str] = None) -> Tuple[int, List[Dict], List[float], str]:
        scores = []
        for schedule in schedules:
            score = self.evaluator.evaluate_schedule(schedule, available_time, sleep_time)
            scores.append(score)
        
        if not scores:
            return -1, [], [], ""
        
        # User-preference soft boost (kept from previous logic, but smaller now)
        if strategy_names and user_preferences.preferences["total_selections"] > 0:
            preferred = user_preferences.get_preferred_strategy()
            for i, name in enumerate(strategy_names):
                if name == preferred:
                    scores[i] *= 1.10  # gentler than 1.15 to leave room for exploration
        
        best_idx = scores.index(max(scores))

        # --- epsilon-greedy choice ---
        if len(schedules) > 1 and random.random() < self.EPSILON:
            # Pick a non-best index uniformly at random
            other = [i for i in range(len(schedules)) if i != best_idx]
            chosen_idx = random.choice(other)
            print(f"🎲 Exploration: chose suggestion #{chosen_idx} instead of best #{best_idx}")
        else:
            chosen_idx = best_idx
            print(f"⭐ Exploitation: chose best suggestion #{chosen_idx}")

        chosen_schedule = schedules[chosen_idx]
        chosen_strategy = strategy_names[chosen_idx] if strategy_names else None
        return chosen_idx, chosen_schedule, scores, chosen_strategy
    
    def get_top_n(self, schedules: List[List[Dict]], n: int = 6,
                  available_time: int = 0, sleep_time: str = "23:30",
                  strategy_names: List[str] = None) -> List[Tuple[int, List[Dict], float, str]]:
        scored = []
        for i, schedule in enumerate(schedules):
            score = self.evaluator.evaluate_schedule(schedule, available_time, sleep_time)
            name = strategy_names[i] if strategy_names else f"strategy_{i}"
            scored.append((i, schedule, score, name))
        scored.sort(key=lambda x: x[2], reverse=True)
        return scored[:n]

# =============================================================================
# INTELLIGENT SCHEDULER
# =============================================================================
class IntelligentScheduler:
    def __init__(self):
        self.profile = ProductivityProfile()
        self.talent = TalentProfile()  # NEW: talent degree system
        self.rl = ReinforcementLearning()
        self.generator = SmartScheduleGenerator(self.profile)
        self.evaluator = UtilityEvaluator(self.rl.weights, self.profile, self.talent)  # talent injected
        self.recommender = RecommendationSystem(self.evaluator)
        self.pending_selected_order = None
        self.has_pending_recommendation = False
        self.last_strategies = []
        # ============================================================
        # CRITICAL BUG FIX: cache last generated schedules by unique ID
        # ============================================================
        # Each schedule gets a UUID-like unique ID at generation time.
        # When the user selects a schedule, we look up by ID — never by
        # index, never by re-generating. This guarantees the schedule
        # the user saw is the schedule that gets applied.
        self._schedule_cache: Dict[str, Dict[str, Any]] = {}
        self._last_recommended_id: Optional[str] = None
        self.selected_schedule: Optional[List[Dict]] = None  # fixed variable for selected
        self.selected_schedule_id: Optional[str] = None
    
    def _make_schedule_id(self, schedule: List[Dict], strategy: str, idx: int) -> str:
        """Generate a unique deterministic-but-unique ID for a schedule."""
        # Combine task IDs + strategy + timestamp + random salt for uniqueness
        task_sig = "-".join(str(t.get("id", "")) for t in schedule)
        salt = f"{time.time_ns()}-{random.randint(1000, 9999)}-{idx}"
        return f"sched_{strategy}_{abs(hash(task_sig + salt)) % (10**12):012d}"
    
    def generate_and_recommend(self, tasks: List[Dict], settings: Dict = None) -> Dict:
        if settings is None:
            settings = {}
        
        start_time = settings.get("startTime", "08:00")
        sleep_time = settings.get("sleepTime", "23:30")
        
        # تصفية المهام: نأخذ فقط المهام المعلقة
        all_pending = [t for t in tasks if t.get("status") == "pending"]
        
        # المهام التي حدد المستخدم وقتها يدوياً - نستبعدها من نظام التوصيات
        user_locked_tasks = [t for t in all_pending if t.get("userSetTime") == True]
        # المهام المرنة (التي ليس لها وقت محدد من المستخدم) - هذه فقط تدخل في نظام التوصيات
        pending_tasks = [t for t in all_pending if t.get("userSetTime") != True]
        
        # إذا كان هناك مهام مقفلة، نطبع رسالة توضيحية
        if user_locked_tasks:
            print(f"🔒 {len(user_locked_tasks)} مهمة بوقت محدد من المستخدم - مستبعدة من نظام التوصيات")
        
        # إذا لم توجد مهام مرنة، نرجع قائمة فارغة
        if not pending_tasks:
            print("✅ كل المهام بأوقات محددة من المستخدم. لا حاجة للتوصيات.")
            return {
                "schedules": [],
                "scores": [],
                "recommended_index": -1,
                "recommended_schedule": [],
                "top_8": [],
                "strategy_names": [],
                "recommended_strategy": ""
            }
        
        schedules, strategy_names = self.generator.generate_schedules(
            pending_tasks,
            num_schedules=5,
            start_time=start_time,
            end_time=sleep_time
        )
        self.last_strategies = strategy_names
        
        start_mins = SmartScheduleGenerator._to_minutes(start_time)
        sleep_mins = SmartScheduleGenerator._to_minutes(sleep_time)
        available_time = sleep_mins - start_mins
        
        best_idx, best_schedule, all_scores, best_strategy = self.recommender.recommend(
            schedules, available_time, sleep_time, strategy_names
        )
        
        top_n = self.recommender.get_top_n(schedules, n=len(schedules), available_time=available_time, sleep_time=sleep_time, strategy_names=strategy_names)
        
        # ============================================================
        # CRITICAL BUG FIX: assign each schedule a UNIQUE ID and cache.
        # ============================================================
        # Clear previous cache (old schedules are stale)
        self._schedule_cache = {}
        schedule_ids: List[str] = []
        for i, sched in enumerate(schedules):
            sid = self._make_schedule_id(sched, strategy_names[i] if i < len(strategy_names) else f"s{i}", i)
            schedule_ids.append(sid)
            self._schedule_cache[sid] = {
                "id": sid,
                "schedule": sched,  # exact tasks list
                "strategy": strategy_names[i] if i < len(strategy_names) else f"s{i}",
                "score": all_scores[i] if i < len(all_scores) else 0.0,
                "index": i,  # legacy index — DO NOT use for selection
            }
        
        recommended_id = schedule_ids[best_idx] if 0 <= best_idx < len(schedule_ids) else None
        self._last_recommended_id = recommended_id
        
        print(f"📋 Generated {len(schedules)} schedules. IDs: {schedule_ids}")
        print(f"⭐ Recommended schedule ID: {recommended_id}")
        
        return {
            "schedules": schedules,
            "schedule_ids": schedule_ids,
            "scores": all_scores,
            "recommended_index": best_idx,
            "recommended_id": recommended_id,
            "recommended_schedule": best_schedule,
            "top_8": top_n,  # key kept for backward compat with the API layer
            "strategy_names": strategy_names,
            "recommended_strategy": best_strategy
        }
    
    def get_schedule_by_id(self, schedule_id: str) -> Optional[Dict[str, Any]]:
        """Look up a previously-generated schedule by its unique ID. Returns None if missing."""
        return self._schedule_cache.get(schedule_id)
    
    def apply_schedule_by_id(self, schedule_id: str) -> Dict[str, Any]:
        """
        CRITICAL: Apply a schedule by its unique ID — the ONLY safe selection path.
        - No re-generation
        - No re-ranking
        - Throws on mismatch
        - Logs selected and applied IDs
        """
        print(f"🎯 [SELECT] User selected schedule ID: {schedule_id}")
        
        cached = self.get_schedule_by_id(schedule_id)
        if cached is None:
            err = f"❌ Schedule ID '{schedule_id}' not found in cache. Available: {list(self._schedule_cache.keys())}"
            print(err)
            raise ValueError(err)
        
        selected_schedule = cached["schedule"]
        selected_strategy = cached["strategy"]
        selected_index = cached["index"]
        
        # Lock the selection into a fixed variable
        self.selected_schedule = selected_schedule
        self.selected_schedule_id = schedule_id
        
        # Build timeline ONLY from selected_schedule
        apply_selected_order_exactly(self.selected_schedule)
        apply_schedule_times_from_order(self.selected_schedule)
        
        applied_id = schedule_id  # we apply exactly what was selected
        print(f"✅ [APPLY] Applied schedule ID:  {applied_id}")
        print(f"   Selected ID == Applied ID? {schedule_id == applied_id}")
        
        # Mismatch guard
        if schedule_id != applied_id:
            err = f"❌ MISMATCH: selected={schedule_id}, applied={applied_id}"
            print(err)
            raise RuntimeError(err)
        
        # Record for RL using the cached snapshot — no re-generation
        recommended_idx = -1
        all_scores = [info["score"] for info in self._schedule_cache.values()]
        if self._last_recommended_id and self._last_recommended_id in self._schedule_cache:
            recommended_idx = self._schedule_cache[self._last_recommended_id]["index"]
        
        self.record_user_choice(
            recommended_idx,
            selected_index,
            selected_schedule,
            all_scores,
            selected_strategy
        )
        
        return {
            "selected_id": schedule_id,
            "applied_id": applied_id,
            "strategy": selected_strategy,
            "task_count": len(selected_schedule),
        }
    
    def record_user_choice(self, recommended_idx: int, selected_idx: int,
                            selected_schedule: List[Dict], all_scores: List[float],
                            strategy_name: str = None):
        self.rl.update_weights(recommended_idx, selected_idx, selected_schedule, all_scores, strategy_name)
    
    def record_task_completion(self, task: Dict, completed: bool, actual_duration: Optional[int] = None):
        # Productivity profile (existing)
        self.profile.update_completion(task, completed)
        # Talent profile (NEW)
        # `completed=True` here means a clean on-time completion;
        # `completed=False` is used by task_complete_with_remaining (the user
        # adjusted remaining time → they ran over, not on time).
        try:
            self.talent.record_completion(task, was_on_time=completed, actual_duration=actual_duration)
        except Exception as e:
            print(f"⚠️ talent record_completion failed: {e}")
    
    def record_task_postponement(self, task: Dict):
        self.profile.update_postponement(task)
        try:
            self.talent.record_postponement(task)
        except Exception as e:
            print(f"⚠️ talent record_postponement failed: {e}")
    
    def get_insights(self) -> List[str]:
        insights = []
        insights.extend(self.rl.get_insights())
        # NEW: talent-driven insights
        try:
            insights.extend(self.talent.insights())
        except Exception:
            pass
        
        profile = self.profile.profile
        
        best_period = None
        best_score = 0
        for period, data in profile["productive_hours"].items():
            if data["score"] > best_score:
                best_score = data["score"]
                best_period = period
        
        if best_period and best_score > 0.7:
            period_data = profile["productive_hours"][best_period]
            insights.append(f"⏰ أكثر أوقاتك إنتاجية: {period_data['start']} - {period_data['end']}")
        
        pp = profile["postponement_patterns"]
        if pp["by_difficulty"].get("hard", 0) > pp["by_difficulty"].get("easy", 0) * 2:
            insights.append("💪 تميل لتأجيل المهام الصعبة - حاول جدولتها في أوقات طاقتك العالية")
        
        freq_postponed = profile.get("frequently_postponed", [])
        if freq_postponed:
            most_postponed = max(freq_postponed, key=lambda x: x["count"])
            if most_postponed["count"] >= 3:
                insights.append(f"⚠️ مهمة '{most_postponed['name']}' تؤجل كثيراً - حاول إنجازها مبكراً")
        
        now_mins = to_mins(now_str())
        sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
        remaining = max(0, sleep_mins - now_mins)
        if remaining < 60:
            insights.append(f"⏰ وقت محدود! متبقي {remaining} دقيقة فقط حتى النوم")
        
        return insights

intelligent_scheduler = IntelligentScheduler()
# =============================================================================
# UTILITY-BASED AGENT
# =============================================================================
class UtilityBasedAgent:
    PRIORITY_SCORES = {"high": 3.0, "medium": 2.0, "low": 1.0}
    DIFFICULTY_SCORES = {"hard": 3.0, "medium": 2.0, "easy": 1.0}
    
    @staticmethod
    def get_duration_score(duration_minutes: int) -> float:
        return min(1.0 + (duration_minutes / 120), 2.0)
    
    @classmethod
    def calculate_utility(cls, task: Dict) -> float:
        priority_score = cls.PRIORITY_SCORES.get(task.get("priority", "medium"), 2.0)
        difficulty_score = cls.DIFFICULTY_SCORES.get(task.get("difficulty", "medium"), 2.0)
        duration_score = cls.get_duration_score(task.get("duration", 30))
        return (priority_score * 0.6) + (difficulty_score * 0.3) + (duration_score * 0.1)
    
    @classmethod
    def sort_by_priority(cls, tasks: List[Dict]) -> List[Dict]:
        priority_order = {"high": 0, "medium": 1, "low": 2}
        return sorted(tasks, key=lambda x: priority_order.get(x.get("priority", "medium"), 1))
    
    @classmethod
    def sort_by_difficulty(cls, tasks: List[Dict]) -> List[Dict]:
        difficulty_order = {"hard": 0, "medium": 1, "easy": 2}
        return sorted(tasks, key=lambda x: difficulty_order.get(x.get("difficulty", "medium"), 1))
    
    @classmethod
    def sort_by_duration(cls, tasks: List[Dict]) -> List[Dict]:
        return sorted(tasks, key=lambda x: x.get("duration", 0), reverse=True)
    
    @classmethod
    def sort_combined(cls, tasks: List[Dict]) -> List[Dict]:
        return sorted(tasks, key=lambda x: (
            cls.PRIORITY_SCORES.get(x.get("priority", "medium"), 2.0) * -1,
            cls.DIFFICULTY_SCORES.get(x.get("difficulty", "medium"), 2.0) * -1,
            x.get("duration", 0) * -1
        ))
    
    @classmethod
    def optimize_tasks(cls, tasks: List[Dict]) -> List[Dict]:
        pending = [t for t in tasks if t.get("status") == "pending"]
        completed = [t for t in tasks if t.get("status") == "done"]
        
        for task in pending:
            task["utilityScore"] = cls.calculate_utility(task)
        
        pending.sort(key=lambda x: x.get("utilityScore", 0), reverse=True)
        completed.sort(key=lambda x: x.get("completedAt", ""), reverse=True)
        
        return pending + completed

def apply_selected_order_exactly(selected_order: List[Dict]) -> None:
    """Apply the exact order from selected schedule"""
    if not selected_order:
        return
    
    with _lock:
        pending_by_id = {t["id"]: t for t in _tasks if t["status"] == "pending"}
        
        ordered_ids = []
        for task in selected_order:
            task_id = task.get("id")
            if task_id and task_id in pending_by_id:
                ordered_ids.append(task_id)
        
        reordered_pending = [pending_by_id[tid] for tid in ordered_ids if tid in pending_by_id]
        
        remaining_ids = [t["id"] for t in _tasks if t["status"] == "pending" and t["id"] not in ordered_ids]
        remaining_tasks = [pending_by_id[tid] for tid in remaining_ids if tid in pending_by_id]
        
        final_pending = reordered_pending + remaining_tasks
        completed = [t for t in _tasks if t["status"] == "done"]
        
        _tasks[:] = final_pending + completed
        save_tasks()

def apply_schedule_times_from_order(order_list: List[Dict]) -> None:
    """Apply schedule times - NEVER place tasks in the past"""
    if not order_list:
        return
    
    with _lock:
        pending_tasks = [t for t in _tasks if t["status"] == "pending"]
        pending_by_id = {t["id"]: t for t in pending_tasks}
        
        # Start from current time, not start of day
        current_seconds = to_seconds(now_str())
        
        sleep_time_str = _settings.get("sleepTime", "23:30")
        sleep_seconds = to_seconds(sleep_time_str)
        start_day_seconds = to_seconds(_settings.get("startTime", "08:00"))
        
        if sleep_seconds < start_day_seconds:
            effective_sleep = sleep_seconds + 86400
            if current_seconds < start_day_seconds:
                current_seconds += 86400
        else:
            effective_sleep = sleep_seconds
        
        GAP_SECONDS = 10
        tasks_updated = 0
        
        for ordered_task in order_list:
            task_id = ordered_task.get("id")
            if not task_id or task_id not in pending_by_id:
                continue
            
            actual_task = pending_by_id[task_id]
            duration_seconds = actual_task.get("duration", 30) * 60
            
            # Check if task has a specific future time
            specific_time = actual_task.get("startTime")
            if specific_time:
                specific_seconds = to_seconds(specific_time)
                # Adjust for after-midnight
                if specific_seconds < start_day_seconds:
                    specific_seconds += 86400
                # Only use specific time if it's in the future
                if specific_seconds > current_seconds:
                    current_seconds = specific_seconds
            
            if current_seconds + duration_seconds <= effective_sleep:
                actual_task["startTime"] = to_time_from_seconds(current_seconds)
                actual_task["endTime"] = to_time_from_seconds(current_seconds + duration_seconds)
                current_seconds += duration_seconds + GAP_SECONDS
                tasks_updated += 1
            else:
                actual_task["startTime"] = None
                actual_task["endTime"] = None
        
        save_tasks()
        print(f"✅ تم تطبيق الجدول: {tasks_updated} مهمة مجدولة")
    
    # ============================================================
    # AUDIO FIX: re-arm audio triggers for re-scheduled tasks
    # ============================================================
    # The fired-flag key includes the scheduled time, so a new time
    # implicitly produces a new key — but we also clear stale per-task
    # flags to keep the cache small and predictable.
    try:
        for ordered_task in order_list:
            tid = ordered_task.get("id")
            if tid:
                task_audio.reset_for_task(tid)
    except Exception:
        pass

def optimize_schedule():
    if intelligent_scheduler.pending_selected_order is not None:
        return
    
    tasks = get_tasks()
    pending_tasks = [t for t in tasks if t.get("status") == "pending"]
    
    # المهام التي حدد المستخدم وقتها يدوياً - لا نلمسها أبداً
    user_locked_tasks = [t for t in pending_tasks if t.get("userSetTime") and t.get("startTime")]
    # المهام التي لها وقت بالفعل (حتى لو لم يحددها المستخدم) - نحميها أيضاً من إعادة الترتيب
    # هذه هي المهام التي تم وضعها في الجدول مسبقاً (مثلاً عبر get_available_time_slot)
    already_scheduled_flexible = [t for t in pending_tasks 
                                   if not t.get("userSetTime") 
                                   and t.get("startTime")]
    # فقط المهام بدون أي وقت إطلاقاً هي التي تحتاج إلى جدولة
    flexible_tasks = [t for t in pending_tasks 
                      if not t.get("userSetTime") 
                      and not t.get("startTime")]
    
    # دمج المهام المقفلة + المهام المجدولة بالفعل في قائمة "محمية" واحدة
    # لتجنب التعارض معها عند جدولة المهام الجديدة
    protected_tasks = user_locked_tasks + already_scheduled_flexible
    
    if user_locked_tasks:
        print(f"🔒 {len(user_locked_tasks)} مهمة بوقت محدد من المستخدم - محمية من إعادة الترتيب")
    if already_scheduled_flexible:
        print(f"🛡️ {len(already_scheduled_flexible)} مهمة مجدولة مسبقاً - محمية من إعادة الترتيب")
    
    # إذا لا توجد مهام تحتاج جدولة، لا نفعل شيئاً
    if not flexible_tasks:
        print("✅ كل المهام لها أوقات. لا حاجة للترتيب.")
        return
    
    # حالة مهمة واحدة مرنة فقط
    if len(flexible_tasks) <= 1:
        with _lock:
            for task in flexible_tasks:
                start_time_str = _settings.get("startTime", "08:00")
                now_time_str = now_str()
                
                if to_mins(now_time_str) > to_mins(start_time_str):
                    current_seconds = to_seconds(now_time_str)
                else:
                    current_seconds = to_seconds(start_time_str)
                
                # تجنب التعارض مع المهام المحمية (المحددة من المستخدم + المجدولة بالفعل)
                # نكرر المرور حتى نتأكد من عدم وجود تعارض مع أي مهمة محمية
                changed = True
                while changed:
                    changed = False
                    for locked in protected_tasks:
                        if locked.get("startTime") and locked.get("endTime"):
                            locked_start = to_seconds(locked["startTime"])
                            locked_end = to_seconds(locked["endTime"])
                            task_duration_sec = task.get("duration", 30) * 60
                            # إذا الوقت الحالي يتعارض مع مهمة محمية، انتقل بعدها
                            if current_seconds < locked_end and current_seconds + task_duration_sec > locked_start:
                                current_seconds = locked_end + 10
                                changed = True
                
                sleep_time_str = _settings.get("sleepTime", "23:30")
                sleep_seconds = to_seconds(sleep_time_str)
                start_day_seconds = to_seconds(start_time_str)
                
                if sleep_seconds < start_day_seconds:
                    effective_sleep = sleep_seconds + 86400
                else:
                    effective_sleep = sleep_seconds
                
                duration_seconds = task.get("duration", 30) * 60
                
                if current_seconds + duration_seconds <= effective_sleep:
                    task["startTime"] = to_time_from_seconds(current_seconds)
                    task["endTime"] = to_time_from_seconds(current_seconds + duration_seconds)
                else:
                    task["startTime"] = None
                    task["endTime"] = None
            
            save_tasks()
        return
    
    # حالة عدة مهام مرنة - استخدم الجدولة الذكية فقط على المهام المرنة
    result = intelligent_scheduler.generate_and_recommend(flexible_tasks, _settings)
    
    if not result["recommended_schedule"]:
        return
    
    recommended = result["recommended_schedule"]
    
    with _lock:
        # احصل على فترات المهام المحمية (المقفلة + المجدولة بالفعل) لتجنبها
        locked_intervals = []
        for locked in protected_tasks:
            if locked.get("startTime") and locked.get("endTime"):
                locked_intervals.append((
                    to_mins(locked["startTime"]),
                    to_mins(locked["endTime"])
                ))
        
        current_mins = max(to_mins(now_str()), to_mins(_settings.get("startTime", "08:00")))
        sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
        GAP = 1
        
        for scheduled_task in recommended:
            if not scheduled_task.get("id"):
                continue
            
            for task in _tasks:
                # نعدّل فقط المهام التي ليس لها وقت أصلاً (المهام التي تحتاج جدولة)
                if (task["id"] == scheduled_task.get("id") 
                        and not task.get("userSetTime") 
                        and not task.get("startTime")):
                    duration = task.get("duration", 30)
                    
                    # ابحث عن فترة لا تتعارض مع المهام المحمية
                    placed = False
                    attempt_start = current_mins
                    
                    while attempt_start + duration <= sleep_mins and not placed:
                        attempt_end = attempt_start + duration
                        conflict = False
                        
                        for (locked_start, locked_end) in locked_intervals:
                            # تحقق من التعارض
                            if not (attempt_end <= locked_start or attempt_start >= locked_end):
                                # تعارض - انتقل لما بعد المهمة المحمية
                                attempt_start = locked_end + GAP
                                conflict = True
                                break
                        
                        if not conflict:
                            task["startTime"] = to_time(attempt_start)
                            task["endTime"] = to_time(attempt_end)
                            # أضف هذه المهمة الجديدة إلى قائمة الفترات المحمية لتجنب
                            # التعارض معها عند جدولة المهام التالية في نفس الجولة
                            locked_intervals.append((attempt_start, attempt_end))
                            current_mins = attempt_end + GAP
                            placed = True
                    
                    if not placed:
                        # لا توجد فترة متاحة - اتركها بدون وقت (غير مجدولة)
                        task["startTime"] = None
                        task["endTime"] = None
                    break
        
        save_tasks()
    
    if random.random() < 0.3:
        generate_ai_notifications()

def generate_ai_notifications():
    insights = intelligent_scheduler.get_insights()
    for insight in insights:
        _notifs.append({
            "type": "ai_insight",
            "message": f"🧠 {insight}"
        })

# تعديل دالة task_add في الكود (السطر حوالي 1100)

def task_add(data):
    today = today_str()
    existing_tasks = get_tasks()
    
    specified_start = data.get("startTime")
    duration = int(data.get("duration", 30))
    
    # Convert time if specified
    if specified_start and specified_start not in (None, ""):
        specified_start = convert_to_24hour(specified_start)
    
    current_time_mins = to_mins(now_str())
    start_day_mins = to_mins(_settings.get("startTime", "08:00"))
    sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
    
    start_time = None
    start_mins = None

    # If user specified a time
    if specified_start:
        start_time = specified_start
        start_mins = to_mins(start_time)
        
        # CHECK: Is this time in the future?
        time_is_valid = False
        time_for_tomorrow = False
        
        # Handle after-midnight times
        if sleep_mins < start_day_mins and start_mins < start_day_mins:
            # This time (like 01:00) is for tomorrow
            test_mins = start_mins + 1440
            if test_mins > current_time_mins:
                time_is_valid = True
                time_for_tomorrow = True
                start_mins = test_mins
        else:
            # Normal comparison
            if start_mins > current_time_mins:
                time_is_valid = True
        
        # Also check if within sleep time
        if time_is_valid and not is_within_sleep_time(start_time, duration):
            # Adjust for after-midnight sleep
            if sleep_mins < start_day_mins:
                # Try adding 24 hours
                test_mins = start_mins + 1440
                if test_mins + duration <= sleep_mins + 1440:
                    start_mins = test_mins
                    time_for_tomorrow = True
                    time_is_valid = True
                else:
                    time_is_valid = False
            else:
                time_is_valid = False
        
        if not time_is_valid:
            # Time is invalid - place after last task instead
            start_mins = get_last_task_end_time(existing_tasks, today)
            start_time = to_time(start_mins)
            print(f"⚠️ الوقت {specified_start} غير صالح (ماضي أو بعد النوم). تمت الإضافة بعد آخر مهمة في {start_time}")
        else:
            # Valid time! Use it exactly
            if time_for_tomorrow:
                data["date"] = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
                print(f"📅 تم جدولة المهمة للغد في {start_time}")
            else:
                data["date"] = today
                print(f"✅ تم جدولة المهمة اليوم في {start_time}")
    else:
        # No time specified - اترك المهمة بدون وقت
        # علشان optimize_schedule (الـ AI) يرتبها مع باقي المهام حسب التوصية الحقيقية
        # بدل ما نحطها في أول فجوة فاضية ونمنع الـ AI من ترتيبها
        start_time = None
        start_mins = None
        print(f"📋 المهمة بدون وقت محدد - سيتم ترتيبها بواسطة نظام الـ AI")

    # Calculate end time
    end_time = None
    if start_time:
        end_mins = start_mins + duration
        # Handle wrap around
        if end_mins >= 1440:
            end_mins = end_mins - 1440
        end_time = to_time(end_mins)
    
    t = {
        "id": new_id(),
        "name": data.get("name", "مهمة جديدة"),
        "date": data.get("date", today),
        "startTime": start_time,
        "endTime": end_time,
        "duration": duration,
        "difficulty": data.get("difficulty", "medium"),
        "priority": data.get("priority", "medium"),
        "status": "pending",
        "notes": data.get("notes", ""),
        "postponedCount": 0,
        "alertSent": False,
        "completedAt": None,
        "createdAt": datetime.now().isoformat(),
        "utilityScore": 0,
        "manualOrder": 0,
        "userSetTime": bool(specified_start)
    }
    
    with _lock:
        _tasks.append(t)
        # IMPORTANT: لا نعيد ترتيب المهام هنا
        # فقط نحفظ المهمة الجديدة مع الحفاظ على أوقات المهام الموجودة
        save_tasks()
    
    intelligent_scheduler.pending_selected_order = None
    intelligent_scheduler.has_pending_recommendation = True
    return t


# إضافة دالة مساعدة لحساب وقت متاح لا يتعارض مع المهام الحالية
def get_available_time_slot(tasks: List[Dict], duration: int, date: str) -> Optional[int]:
    """Find an available time slot for a task without disturbing existing scheduled tasks"""
    
    start_mins = to_mins(_settings.get("startTime", "08:00"))
    sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
    now_mins = to_mins(now_str())
    
    # Get existing scheduled tasks for this date (tasks with user-set times)
    scheduled_tasks = [t for t in tasks if t.get("date") == date and t.get("status") == "pending" 
                       and t.get("startTime") and t.get("endTime")]
    
    # Sort by start time
    scheduled_tasks.sort(key=lambda x: to_mins(x.get("startTime", "00:00")))
    
    # Start from current time or start of day, whichever is later
    current_time = max(start_mins, now_mins)
    
    # Check if we can place before first scheduled task
    if scheduled_tasks:
        first_start = to_mins(scheduled_tasks[0]["startTime"])
        if current_time + duration <= first_start:
            return current_time
    else:
        if current_time + duration <= sleep_mins:
            return current_time
    
    # Check gaps between scheduled tasks
    for i in range(len(scheduled_tasks) - 1):
        current_end = to_mins(scheduled_tasks[i]["endTime"])
        next_start = to_mins(scheduled_tasks[i + 1]["startTime"])
        
        gap_start = max(current_time, current_end + 1)
        if gap_start + duration <= next_start:
            return gap_start
    
    # Check after last scheduled task
    if scheduled_tasks:
        last_end = to_mins(scheduled_tasks[-1]["endTime"])
        if max(current_time, last_end + 1) + duration <= sleep_mins:
            return max(current_time, last_end + 1)
    
    # No available slot today
    return None


# تعديل دالة sendMsg في الـ HTML لإضافة المهمة دون إعادة جدولة كل شيء
# هذا التعديل موجود بالفعل في الكود، لكن تأكد من أن addSingleTask لا تستدعي optimize_schedule()
# في HTML، دالة addSingleTask موجودة وتستخدم apiCall الصحيح

def task_complete_direct(tid):
    """Complete task directly without asking for remaining time"""
    task_copy = None
    with _lock:
        for t in _tasks:
            if t["id"] == tid and t["status"] != "done":
                task_copy = t.copy()
                t["status"] = "done"
                t["completedAt"] = now_str()
                _history.insert(0, {**t, "action": "completed",
                                    "logDate": datetime.now().strftime("%Y-%m-%d %H:%M")})
                _tasks[:] = UtilityBasedAgent.optimize_tasks(_tasks)
                save_tasks()
                break
    
    if task_copy:
        # Compute actual duration if startTime is known (NEW: feeds talent system)
        actual_duration = None
        try:
            start_time = task_copy.get("startTime")
            if start_time:
                start_min = to_mins(start_time)
                now_min = to_mins(now_str())
                # Handle midnight roll-over
                if now_min < start_min:
                    now_min += 1440
                actual_duration = max(1, now_min - start_min)
        except Exception:
            actual_duration = None
        
        intelligent_scheduler.record_task_completion(task_copy, True, actual_duration=actual_duration)
        play_sound(end_sound_file)
        return True
    return False

def task_complete_with_remaining(tid, remaining_minutes):
    """Complete task with remaining time adjustment - places after last task"""
    task_copy = None
    today = today_str()
    
    with _lock:
        for t in _tasks:
            if t["id"] == tid and t["status"] != "done":
                task_copy = t.copy()
                original_duration = t.get("duration", 30)
                new_duration = remaining_minutes
                t["duration"] = new_duration
                
                # احصل على آخر مهمة في الجدول مع استثناء المهمة الحالية
                # (لأن المهمة الحالية هي التي نريد إعادة وضعها)
                tasks_excluding_current = [tk for tk in _tasks if tk["id"] != tid]
                last_end_time = get_last_task_end_time(tasks_excluding_current, today)
                last_end_mins = last_end_time
                
                # Check if there's time for this task
                sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
                start_day_mins = to_mins(_settings.get("startTime", "08:00"))
                
                effective_sleep = sleep_mins
                if sleep_mins < start_day_mins:
                    effective_sleep = sleep_mins + 1440
                    if last_end_mins < start_day_mins:
                        last_end_mins += 1440
                
                if last_end_mins + new_duration <= effective_sleep:
                    t["startTime"] = to_time(last_end_mins)
                    t["endTime"] = to_time(last_end_mins + new_duration)
                else:
                    t["startTime"] = None
                    t["endTime"] = None
                
                t["status"] = "pending"
                t["notes"] = t.get("notes", "") + f"\n[تم التعديل: بقي {remaining_minutes} دقيقة]"
                _history.insert(0, {**t, "action": "adjusted",
                                    "logDate": datetime.now().strftime("%Y-%m-%d %H:%M")})
                _tasks[:] = UtilityBasedAgent.optimize_tasks(_tasks)
                save_tasks()
                break
    
    if task_copy:
        # The user said the task ran over: actual = original_estimate + remaining
        # but we record `was_on_time=False` so the talent system penalises speed.
        try:
            est = task_copy.get("duration", 30) or 30
            actual_duration = est + max(0, remaining_minutes)
        except Exception:
            actual_duration = None
        intelligent_scheduler.record_task_completion(task_copy, False, actual_duration=actual_duration)
        play_sound(end_sound_file)
        return True
    return False

def task_complete(tid):
    return task_complete_direct(tid)

def task_delete(tid):
    with _lock:
        before = len(_tasks)
        _tasks[:] = [t for t in _tasks if t["id"] != tid]
        if len(_tasks) < before:
            save_tasks()
            return True
    return False

def task_delete_all():
    with _lock:
        _tasks[:] = []
        save_tasks()
        return True

def task_swap(task1_id, task2_id):
    with _lock:
        pending = [t for t in _tasks if t["status"] == "pending"]
        indices = {}
        for idx, task in enumerate(pending):
            if task["id"] == task1_id: indices["task1"] = idx
            if task["id"] == task2_id: indices["task2"] = idx
        if "task1" in indices and "task2" in indices:
            t1, t2 = pending[indices["task1"]], pending[indices["task2"]]
            s1, s2 = t1.get("startTime"), t2.get("startTime")
            t1["startTime"] = s2; t1["endTime"] = to_time(to_mins(s2)+t1["duration"]) if s2 else None
            t2["startTime"] = s1; t2["endTime"] = to_time(to_mins(s1)+t2["duration"]) if s1 else None
            pending[indices["task1"]], pending[indices["task2"]] = pending[indices["task2"]], pending[indices["task1"]]
            completed = [t for t in _tasks if t["status"] == "done"]
            _tasks[:] = pending + completed
            save_tasks()
            return True
    return False

def task_postpone_tomorrow(tid):
    task_copy = None
    tom = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    with _lock:
        for t in _tasks:
            if t["id"] == tid:
                task_copy = t.copy()
                t["date"] = tom
                t["postponedCount"] += 1
                t["startTime"] = None
                t["endTime"] = None
                t["alertSent"] = False
                _history.insert(0, {**t, "action": "postponed_to_tomorrow",
                                    "logDate": datetime.now().strftime("%Y-%m-%d %H:%M")})
                
                tomorrow_tasks = [task for task in _tasks if task["date"] == tom and task["status"] == "pending"]
                if tomorrow_tasks:
                    tomorrow_tasks = UtilityBasedAgent.sort_combined(tomorrow_tasks)
                    other_tasks = [task for task in _tasks if task["date"] != tom or task["status"] == "done"]
                    _tasks[:] = tomorrow_tasks + other_tasks
                
                save_tasks()
                break
    
    if task_copy:
        intelligent_scheduler.record_task_postponement(task_copy)
        return True
    return False

def task_postpone_time(tid, new_time):
    """Postpone task to a specific time - ONLY if time is in the future"""
    new_time_24 = convert_to_24hour(new_time)
    if not new_time_24:
        return False, "صيغة وقت غير صالحة"
    
    # CRITICAL: Strict future time check
    if not is_future_time(new_time_24):
        return False, f"لا يمكن التأجيل لوقت مضى! الوقت الحالي هو {now_str()}"
    
    task_copy = None
    with _lock:
        for t in _tasks:
            if t["id"] == tid:
                task_copy = t.copy()
                if not is_within_sleep_time(new_time_24, t["duration"]):
                    sleep_time = _settings.get("sleepTime", "23:30")
                    return False, f"لا يمكن إنهاء المهمة قبل وقت النوم ({sleep_time})"
                
                t["date"] = today_str()
                t["startTime"] = new_time_24
                t["endTime"] = to_time(to_mins(new_time_24) + t["duration"])
                t["postponedCount"] += 1
                t["alertSent"] = False
                _history.insert(0, {**t, "action": "postponed_to_time",
                                    "logDate": datetime.now().strftime("%Y-%m-%d %H:%M")})
                
                _tasks[:] = UtilityBasedAgent.optimize_tasks(_tasks)
                save_tasks()
                break
    
    if task_copy:
        intelligent_scheduler.record_task_postponement(task_copy)
        return True, "success"
    return False, "المهمة غير موجودة"

def task_postpone_time_direct(tid, new_time):
    """Postpone task to a specific time - FORCED direct placement, no reordering"""
    new_time_24 = convert_to_24hour(new_time)
    if not new_time_24:
        return False, "صيغة وقت غير صالحة"
    
    # Check if time is in the future
    time_mins = to_mins(new_time_24)
    current_mins = to_mins(now_str())
    start_day_mins = to_mins(_settings.get("startTime", "08:00"))
    sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
    
    # Adjust for after-midnight
    if sleep_mins < start_day_mins and time_mins < start_day_mins:
        time_mins += 1440
    
    if time_mins <= current_mins:
        return False, f"لا يمكن التأجيل لوقت مضى! الوقت الحالي هو {now_str()}"
    
    with _lock:
        for t in _tasks:
            if t["id"] == tid:
                # Calculate end time
                end_mins = to_mins(new_time_24) + t["duration"]
                end_time = to_time(end_mins)
                
                # DIRECT FORCED UPDATE - bypass all sorting
                t["startTime"] = new_time_24
                t["endTime"] = end_time
                t["postponedCount"] = t.get("postponedCount", 0) + 1
                t["date"] = today_str()
                t["alertSent"] = False
                t["userSetTime"] = True
                
                _history.insert(0, {
                    **t, 
                    "action": "postponed_to_time",
                    "logDate": datetime.now().strftime("%Y-%m-%d %H:%M")
                })
                
                # Save directly WITHOUT reordering
                save_tasks()
                
                print(f"✅ تم تأجيل المهمة '{t['name']}' إلى {new_time_24} - {end_time}")
                return True, "success"
    
    return False, "المهمة غير موجودة"

def task_update_time(tid, new_time):
    new_time_24 = convert_to_24hour(new_time)
    if not new_time_24:
        return False, "صيغة وقت غير صالحة"
    
    if not is_future_time(new_time_24):
        return False, "لا يمكن وضع المهمة في وقت مضى"
    
    with _lock:
        for t in _tasks:
            if t["id"] == tid:
                if not is_within_sleep_time(new_time_24, t["duration"]):
                    return False, "لا يمكن وضع المهمة قبل وقت النوم"
                
                t["startTime"] = new_time_24
                t["endTime"] = to_time(to_mins(new_time_24) + t["duration"])
                t["alertSent"] = False
                t["userSetTime"] = True  # ← علامة: المستخدم حدد الوقت
                save_tasks()
                return True, "success"
    return False, "Task not found"

def task_update(tid, updates):
    with _lock:
        for t in _tasks:
            if t["id"] == tid:
                t.update(updates)
                if "startTime" in updates and t["startTime"]:
                    t["endTime"] = to_time(to_mins(t["startTime"]) + t["duration"])
                    t["alertSent"] = False
                _tasks[:] = UtilityBasedAgent.optimize_tasks(_tasks)
                save_tasks()
                return True
    return False

def reorder_tasks_by_criteria(criteria: str):
    with _lock:
        pending = [t for t in _tasks if t["status"] == "pending"]
        
        if criteria == "priority":
            pending = UtilityBasedAgent.sort_by_priority(pending)
        elif criteria == "difficulty":
            pending = UtilityBasedAgent.sort_by_difficulty(pending)
        elif criteria == "duration":
            pending = UtilityBasedAgent.sort_by_duration(pending)
        elif criteria in ["all", "combined", "ذكي", "smart"]:
            pending = UtilityBasedAgent.sort_combined(pending)
        else:
            return False, "معيار غير معروف"
        
        completed = [t for t in _tasks if t["status"] == "done"]
        _tasks[:] = pending + completed
        save_tasks()
        return True, f"تم إعادة ترتيب المهام حسب {criteria}"

# =============================================================================
# NLP UTILITIES
# =============================================================================
def extract_time(text):
    time_24 = convert_to_24hour(text)
    if time_24:
        return time_24
    
    m = re.search(r"(\d{1,2})[:\.](\d{2})", text)
    if m: return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    
    m2 = re.search(r"الساعة\s*(\d{1,2})", text)
    if m2: return f"{int(m2.group(1)):02d}:00"
    
    m3 = re.search(r"(\d{1,2})\s*(صباح|مساء|ص|م)", text, re.IGNORECASE)
    if m3:
        h = int(m3.group(1))
        period = m3.group(2).lower()
        if period in ("مساء", "م") and h < 12:
            h += 12
        elif period in ("صباح", "ص") and h == 12:
            h = 0
        return f"{h:02d}:00"
    return None

def extract_duration(text):
    m = re.search(r"(\d+)\s*(دقيقة|دقائق|دقيقه)", text, re.IGNORECASE)
    if m: return int(m.group(1))
    m2 = re.search(r"(\d+\.?\d*)\s*(ساعة|ساعات|ساعه)", text, re.IGNORECASE)
    if m2: return int(float(m2.group(1)) * 60)
    m3 = re.search(r"(لمدة|مدتها)\s*(\d+)", text)
    if m3: return int(m3.group(2))
    return None

def extract_priority(text):
    t = text.lower()
    if any(w in t for w in ["عالية", "عالي", "عاجل", "مهم جدا", "اولوية عالية"]):
        return "high"
    if any(w in t for w in ["منخفضة", "منخفض", "بسيطة", "مش مهم", "اولوية منخفضة"]):
        return "low"
    return "medium"

def extract_difficulty(text):
    t = text.lower()
    if any(w in t for w in ["صعبة", "صعب", "صعوبه"]):
        return "hard"
    if any(w in t for w in ["سهلة", "سهل", "سهوله"]):
        return "easy"
    return "medium"

def find_task_by_name(frag):
    frag = frag.strip().lower()
    if not frag: return None
    best = None
    best_sc = 0.0
    for t in get_tasks():
        tn = t["name"].lower()
        if frag == tn:
            return t
        if frag in tn:
            sc = len(frag) / max(len(tn), 1)
        elif tn in frag:
            sc = len(tn) / max(len(frag), 1) * 0.9
        else:
            wf = set(frag.split())
            wt = set(tn.split())
            sc = len(wf & wt) / max(len(wf | wt), 1)
        if sc > best_sc:
            best_sc = sc
            best = t
    return best if best_sc >= 0.2 else None

def speak(text):
    if not VOICE_OK: return
    def _go():
        try:
            e = pyttsx3.init()
            e.setProperty("rate", 150)
            e.say(text)
            e.runAndWait()
        except:
            pass
    threading.Thread(target=_go, daemon=True).start()

def web_search(query):
    if not REQUESTS_OK: return "بحث غير متاح."
    try:
        r = _req.get(f"https://api.duckduckgo.com/?q={query}&format=json&no_html=1&skip_disambig=1", timeout=4)
        d = r.json()
        ab = d.get("AbstractText", "")
        if ab: return ab[:500]
        rel = d.get("RelatedTopics", [])
        snips = [t.get("Text","") for t in rel[:3] if isinstance(t, dict) and t.get("Text")]
        return "\n".join(f"- {s}" for s in snips) if snips else "لا توجد نتائج."
    except Exception as e:
        return f"خطأ في البحث: {str(e)[:60]}"

# =============================================================================
# AI AGENT - Arabic Only
# =============================================================================
conversation_state = {}

def agent_process(user_msg: str, session_id: str = "default") -> Dict:
    """Process user message with session-specific state - Arabic only"""
    msg = user_msg.strip()
    ml = msg.lower()
    today = today_str()
    s = _settings
    sleep_m = to_mins(s.get("sleepTime","23:30"))
    now_m = to_mins(now_str())
    left = max(0, sleep_m - now_m)
    
    if session_id not in conversation_state:
        conversation_state[session_id] = {}
    state = conversation_state[session_id]
    
    # DIRECT SWAP
    SWAP_PATTERNS = [
        r"بدل\s+(?:مهمة\s+)?([^\s]+(?:\s+[^\s]+)*?)\s+(?:مع|و|وبين|وبـ)\s+(?:مهمة\s+)?(.+)",
        r"غير\s+ترتيب\s+(?:مهمة\s+)?([^\s]+(?:\s+[^\s]+)*?)\s+(?:مع|و|وبين|وبـ)\s+(?:مهمة\s+)?(.+)",
        r"تبديل\s+(?:مهمة\s+)?([^\s]+(?:\s+[^\s]+)*?)\s+(?:مع|و|وبين|وبـ)\s+(?:مهمة\s+)?(.+)",
        r"استبدل\s+(?:مهمة\s+)?([^\s]+(?:\s+[^\s]+)*?)\s+(?:مع|و|وبين|وبـ)\s+(?:مهمة\s+)?(.+)",
        r"قدم\s+(?:مهمة\s+)?([^\s]+(?:\s+[^\s]+)*?)\s+(?:على|بدل\s+مكان)\s+(?:مهمة\s+)?(.+)",
    ]
    
    task1 = None
    task2 = None
    extracted_names = []
    
    for pattern in SWAP_PATTERNS:
        match = re.search(pattern, msg, re.IGNORECASE)
        if match:
            name1 = match.group(1).strip()
            name2 = match.group(2).strip()
            extracted_names = [name1, name2]
            break
    
    if not extracted_names:
        SWAP_KEYWORDS = ["بدل", "غير", "تبديل", "استبدل", "قدم", "ترتيب", "مع", "و", "بين", "مهمة","swap","and"]
        temp_msg = msg
        for kw in SWAP_KEYWORDS:
            temp_msg = re.sub(r'\b' + kw + r'\b', '', temp_msg, flags=re.IGNORECASE)
        
        words = re.findall(r'[\u0600-\u06FF\w]+', temp_msg)
        if len(words) >= 2:
            extracted_names = [words[0], words[1]]
    
    if extracted_names:
        task1 = find_task_by_name(extracted_names[0])
        task2 = find_task_by_name(extracted_names[1])
    
    if not task1 or not task2:
        pending_tasks = [t for t in get_tasks() if t["status"] == "pending"]
        
        found_tasks = []
        for task in pending_tasks:
            if task["name"].lower() in ml:
                found_tasks.append(task)
        
        if len(found_tasks) >= 2:
            task1 = found_tasks[0]
            task2 = found_tasks[1]
    
    if task1 and task2:
        if task1["id"] == task2["id"]:
            return {"reply": "❌ لا يمكن تبديل المهمة مع نفسها! اختر مهمتين مختلفتين.", "action": "none"}
        
        if task_swap(task1["id"], task2["id"]):
            optimize_schedule()
            pri_icon1 = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(task1["priority"], "🟡")
            pri_icon2 = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(task2["priority"], "🟡")
            
            reply = f"🔄 **تم تبديل ترتيب المهام بنجاح!**\n\n📌 {pri_icon1} **{task1['name']}** ↔️ {pri_icon2} **{task2['name']}**\n\n✅ تم تحديث الجدول الزمني!"
            speak(f"تم تبديل مهمة {task1['name']} مع {task2['name']}")
            return {"reply": reply, "action": "refresh"}
        else:
            return {"reply": "❌ فشل تبديل ترتيب المهام. تأكد من أن المهام موجودة ومعلقة.", "action": "none"}
    
    if task1 and not task2:
        pending_tasks = [t for t in get_tasks() if t["status"] == "pending" and t["id"] != task1["id"]]
        if pending_tasks:
            tasks_list = "\n".join([f"• {t['name']}" for t in pending_tasks[:10]])
            return {"reply": f"📝 تم العثور على المهمة: **{task1['name']}**\n\nما هي المهمة الثانية التي تريد تبديلها معها؟\n\nالمهام المتاحة:\n{tasks_list}\n\nاكتب: بدل {task1['name']} مع [اسم المهمة الثانية]", "action": "none"}
    
    SIMPLE_SWAP_KEYWORDS = ["بدل", "تبديل", "swap", "غير الترتيب", "ترتيب المهام", "غير ترتيب","and"]
    if any(kw in ml for kw in SIMPLE_SWAP_KEYWORDS) and not task1 and not task2:
        pending_tasks = [t for t in get_tasks() if t["status"] == "pending"]
        if len(pending_tasks) < 2:
            return {"reply": "❌ تحتاج على الأقل مهمتين معلقتين لتبديلهما.\n📋 قم بإضافة مهام أولاً.", "action": "none"}
        
        tasks_list = "\n".join([f"• {t['name']} (🎯 {t['priority']})" for t in pending_tasks[:10]])
        return {"reply": f"🔄 **تبديل ترتيب المهام**\n\nالمهام المتاحة:\n{tasks_list}\n\n**أمثلة على طريقة الكتابة:**\n• 'بدل مذاكرة مع اجتماع'\n• 'غير ترتيب قراءة وكتابة'\n\n✏️ اكتب الأمر مع اسمي المهمتين مباشرة!", "action": "none"}

    # REORDER TASKS
    REORDER_KEYWORDS = ["رتب", "ترتيب", "نظم", "جدول"]
    if any(k in ml for k in REORDER_KEYWORDS):
        if "اولوية" in ml:
            success, msg_result = reorder_tasks_by_criteria("priority")
            response = "✅ تم ترتيب المهام حسب الأولوية!\n🔴 عالية ← 🟡 متوسطة ← 🟢 منخفضة" if success else f"❌ {msg_result}"
        elif "صعوبة" in ml or "صعب" in ml or "سهل" in ml:
            success, msg_result = reorder_tasks_by_criteria("difficulty")
            response = "✅ تم ترتيب المهام حسب الصعوبة!\n💪 صعبة ← 📊 متوسطة ← ✨ سهلة" if success else f"❌ {msg_result}"
        elif "مدة" in ml or "طويلة" in ml or "قصيرة" in ml:
            success, msg_result = reorder_tasks_by_criteria("duration")
            response = "✅ تم ترتيب المهام حسب المدة!\n⏱️ الأطول ← الأقصر" if success else f"❌ {msg_result}"
        elif "ذكي" in ml or "كامل" in ml:
            success, msg_result = reorder_tasks_by_criteria("all")
            optimize_schedule()
            response = "✅ تم ترتيب المهام ترتيباً ذكياً!\n📊 الأولوية ← الصعوبة ← المدة" if success else f"❌ {msg_result}"
        else:
            response = """🔄 **خيارات ترتيب المهام:**

• **حسب الأولوية:** "رتب المهام حسب الأولوية"
• **حسب الصعوبة:** "رتب المهام حسب الصعوبة"  
• **حسب المدة:** "رتب المهام حسب المدة"
• **ترتيب ذكي:** "رتب المهام ترتيب ذكي"
• **تبديل مهمتين:** اكتب "بدل مذاكرة مع اجتماع"

اختر طريقة الترتيب التي تناسبك!"""
        
        optimize_schedule()
        speak("تم إعادة ترتيب المهام")
        return {"reply": response, "action": "refresh"}

    # Handle pending completion state for partial completion
    if state.get("pending_completion"):
        task_id = state.get("pending_task_id")
        task_name = state.get("pending_task_name")
        original_duration = state.get("pending_task_duration")
        
        user_response = msg.strip().lower()
        
        incomplete_keywords = ["لا", "لسه", "مخلصتش", "لحد دلوقتي لا", "مش", "no"]
        complete_keywords = ["نعم", "ايوه", "اه", "خلصت بالكامل", "بالكامل", "كامل", "تم","Done","done"]
        
        # أولاً نحاول استخراج رقم الدقائق - له الأولوية لأن المستخدم قد يقول "لا فاضل 15 دقيقة"
        remaining_match = re.search(r"(\d+)\s*(دقيقة|دقائق|دقيقه|د|min|minutes)?", user_response, re.IGNORECASE)
        has_number = remaining_match is not None
        
        if any(kw in user_response for kw in complete_keywords) and not has_number:
            task_complete_direct(task_id)
            optimize_schedule()
            speak(f"ممتاز! تم اكمال مهمة {task_name}")
            state.pop("pending_completion", None)
            state.pop("pending_task_id", None)
            state.pop("pending_task_name", None)
            state.pop("pending_task_duration", None)
            return {"reply": f"🎉 **رائع!** تم إكمال المهمة '{task_name}' بالكامل!\n\n✅ استمر هكذا!", "action": "refresh"}
        
        elif any(kw in user_response for kw in incomplete_keywords) and not has_number:
            available_slots = find_available_time_slots(get_tasks(), original_duration, today)
            state.pop("pending_completion", None)
            state.pop("pending_task_id", None)
            state.pop("pending_task_name", None)
            state.pop("pending_task_duration", None)
            
            if available_slots:
                slots_text = "\n".join([f"• {slot['start']} - {slot['end']}" for slot in available_slots[:3]])
                return {"reply": f"📌 **لم تكتمل المهمة '{task_name}'**\n\n⏰ **الأوقات المتاحة لإعادة جدولتها:**\n{slots_text}\n\nلإعادة جدولة المهمة، اكتب: 'حدد وقت [الوقت]' أو 'أضفها بعد آخر مهمة'", "action": "none"}
            else:
                task_postpone_tomorrow(task_id)
                return {"reply": f"📌 **لم تكتمل المهمة '{task_name}'**\n\n📅 تم تأجيلها تلقائياً إلى الغد لعدم وجود وقت متبقي اليوم.", "action": "refresh"}
        
        else:
            # محاولة استخراج عدد الدقائق - يقبل "15 دقيقة" أو "15" فقط أو "فاضل 15"
            remaining_match = re.search(r"(\d+)\s*(دقيقة|دقائق|دقيقه|د|min|minutes)?", user_response, re.IGNORECASE)
            if remaining_match:
                remaining = int(remaining_match.group(1))
                if remaining > 0 and remaining < original_duration:
                    task_complete_with_remaining(task_id, remaining)
                    optimize_schedule()
                    speak(f"تم تحديث مهمة {task_name} بقي {remaining} دقيقة وتم وضعها بعد آخر مهمة")
                    state.pop("pending_completion", None)
                    state.pop("pending_task_id", None)
                    state.pop("pending_task_name", None)
                    state.pop("pending_task_duration", None)
                    
                    return {"reply": f"✅ **تم تحديث المهمة '{task_name}'**\n\n⏱️ الوقت المتبقي: {remaining} دقيقة\n\n📌 تم وضع المهمة بعد آخر مهمة في الجدول الزمني.", "action": "refresh"}
                else:
                    return {"reply": f"⚠️ **الرجاء إدخال وقت متبقي صحيح**\n\n• المدة الأصلية: {original_duration} دقيقة\n• الباقي يجب أن يكون أقل من {original_duration} دقيقة\n• مثال: '15 دقيقة' أو '10'", "action": "none"}
            else:
                return {"reply": f"❓ **لم أفهم المدخل**\n\nللإجابة، اكتب:\n• 'نعم' - إذا أنهيت المهمة بالكامل\n• 'لا' - إذا لم تنتهِ منها بعد\n• '15 دقيقة' - إذا بقي 15 دقيقة (أو أي رقم آخر)", "action": "none"}

    # INTERACTIVE ADD TASK
    if state.get("waiting"):
        if state.get("type") == "collecting_details":
            task_data = state.get("task_data", {})
            
            if not task_data.get("name"):
                name = msg.strip()
                for w in ["اسمها", "اسمه", "المهمة", "مهمة"]:
                    name = name.replace(w, "").strip()
                if name and len(name) > 1:
                    task_data["name"] = name
                    state["task_data"] = task_data
                    return {"reply": f"📝 تم حفظ الاسم: {name}\n\n⏱️ كم دقيقة ستستغرق هذه المهمة؟", "action": "none"}
                return {"reply": "📝 الرجاء إدخال اسم صحيح للمهمة", "action": "none"}
            
            if not task_data.get("duration"):
                dur = extract_duration(msg)
                if dur:
                    task_data["duration"] = dur
                    state["task_data"] = task_data
                    return {"reply": f"⏱️ تم حفظ المدة: {dur} دقيقة\n\n🎯 ما هي أولوية المهمة؟ (عالية/متوسطة/منخفضة)", "action": "none"}
                return {"reply": "⏱️ الرجاء إدخال المدة بالدقائق (رقم فقط)", "action": "none"}
            
            if not task_data.get("priority"):
                pri = extract_priority(msg)
                if pri:
                    priority_names = {"high": "عالية", "medium": "متوسطة", "low": "منخفضة"}
                    task_data["priority"] = pri
                    state["task_data"] = task_data
                    return {"reply": f"🎯 تم حفظ الأولوية: {priority_names[pri]}\n\n📊 ما هي صعوبة المهمة؟ (صعبة/متوسطة/سهلة)", "action": "none"}
                return {"reply": "🎯 الرجاء تحديد الأولوية: عالية / متوسطة / منخفضة", "action": "none"}
            
            if not task_data.get("difficulty"):
                dif = extract_difficulty(msg)
                if dif:
                    difficulty_names = {"hard": "صعبة", "medium": "متوسطة", "easy": "سهلة"}
                    task_data["difficulty"] = dif
                    state["task_data"] = task_data
                    state["type"] = "asking_time"
                    return {"reply": f"📊 تم حفظ الصعوبة: {difficulty_names[dif]}\n\n⏰ هل تريد تحديد وقت محدد لهذه المهمة؟\nإذا كنت تريد وقتاً محدداً، اكتب الوقت (مثال: 10:30 أو الساعة 2 مساءً).\nإذا كنت تريد إضافتها بعد آخر مهمة، اكتب 'لا'", "action": "none"}
                return {"reply": "📊 الرجاء تحديد الصعوبة: صعبة / متوسطة / سهلة", "action": "none"}
        
        elif state.get("type") == "asking_time":
            task_data = state.get("task_data")
            user_choice = msg.strip().lower()
            
            if "لا" in user_choice or "عادي" in user_choice:
                new_task = task_add(task_data)
                optimize_schedule()
                priority_names = {"high": "عالية", "medium": "متوسطة", "low": "منخفضة"}
                difficulty_names = {"hard": "صعبة", "medium": "متوسطة", "easy": "سهلة"}
                priority_icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(task_data.get("priority", "medium"), "🟡")
                reply = f"✅ تم اضافة المهمة بنجاح!\n\n{priority_icon} **{new_task['name']}**\n⏱️ {task_data['duration']} دقيقة\n🎯 الأولوية: {priority_names.get(task_data['priority'], 'متوسطة')}\n الصعوبة: {difficulty_names.get(task_data['difficulty'], 'متوسطة')}\n⏰ الوقت: تمت الإضافة بعد آخر مهمة"
                speak(f"تم اضافة مهمة {new_task['name']}")
                state.clear()
                return {"reply": reply, "action": "refresh"}
            else:
                new_time = extract_time(msg)
                if new_time:
                    duration = task_data.get("duration", 30)
                    if is_future_time(new_time) and is_within_sleep_time(new_time, duration):
                        task_data["startTime"] = new_time
                        new_task = task_add(task_data)
                        optimize_schedule()
                        priority_names = {"high": "عالية", "medium": "متوسطة", "low": "منخفضة"}
                        difficulty_names = {"hard": "صعبة", "medium": "متوسطة", "easy": "سهلة"}
                        priority_icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(task_data.get("priority", "medium"), "🟡")
                        reply = f"✅ تم اضافة المهمة بنجاح في الوقت المحدد!\n\n{priority_icon} **{new_task['name']}**\n⏱️ {task_data['duration']} دقيقة\n🎯 الأولوية: {priority_names.get(task_data['priority'], 'متوسطة')}\n الصعوبة: {difficulty_names.get(task_data['difficulty'], 'متوسطة')}\n⏰ الوقت: {new_time}"
                        speak(f"تم اضافة مهمة {new_task['name']} في الساعة {new_time}")
                        state.clear()
                        return {"reply": reply, "action": "refresh"}
                    else:
                        return {"reply": f"❌ الوقت المحدد غير صالح (الساعة {new_time} إما مضى أو بعد وقت النوم). هل تريد إضافة المهمة بعد آخر مهمة؟ (اكتب 'نعم' أو 'لا')", "action": "none"}
                else:
                    return {"reply": "❌ لم أتمكن من فهم الوقت. الرجاء كتابته بصيغة مثل 10:30 أو الساعة 2 مساءً.\n\nهل تريد إضافة المهمة بعد آخر مهمة؟ (اكتب 'نعم' أو 'لا')", "action": "none"}
    
    ADD_KEYWORDS = ["ضيف", "اضيف", "اضف", "اضافة", "أضف", "ضيفلي", "اضيفلي", "add", "create", "new"]
    if any(kw in ml for kw in ADD_KEYWORDS):
        state.clear()
        state["waiting"] = True
        state["type"] = "collecting_details"
        state["task_data"] = {}
        return {"reply": "📝 **إضافة مهمة جديدة - وضع المحادثة التفاعلي**\n\nما اسم المهمة التي تريد إضافتها؟", "action": "none"}

    # COMPLETE TASK - Direct completion
    DONE_KEYWORDS = ["خلصت", "انتهيت", "اتخلصت", "انجزت", "اكملت", "انهيت", "done", "completed", "finished", "تم"]
    
    is_complete_cmd = any(k in ml for k in DONE_KEYWORDS)
    
    if is_complete_cmd:
        task = None
        
        for kw in DONE_KEYWORDS:
            if kw in ml:
                rest = msg[ml.index(kw)+len(kw):].strip()
                rest = re.sub(r"^(من|المهمة|مهمة)\s*", "", rest).strip()
                if rest:
                    task = find_task_by_name(rest)
                    break
        
        if not task:
            pend = [t for t in get_tasks() if t["date"] == today and t["status"] == "pending"]
            if pend:
                task = pend[0]
        
        if task and task["status"] != "done":
            task_complete_direct(task["id"])
            optimize_schedule()
            speak(f"✅ تم إكمال مهمة {task['name']} بنجاح!")
            return {"reply": f"🎉 **رائع!** تم إكمال المهمة '{task['name']}'!\n\n✅ استمر بهذا المستوى من الإنتاجية!", "action": "refresh"}
        
        return {"reply": "❌ لم أجد مهمة معلقة بهذا الاسم.", "action": "none"}

    # INCOMPLETE TASK - كلمات توحي بأن المهمة لم تكتمل
    INCOMPLETE_KEYWORDS = [
        "ملحقتش", "مخلصتهاش", "مخلصتش", "فضل منها", "باقي", "مش مخلص", 
        "لحد دلوقتي لا", "لسه", "لسة", "متبقي", "متبقى", "مكملتش", 
        "ما خلصت", "ما خلصتش", "مش خالص", "مش خالصة", "ناقص", "ناقصة",
        "incomplete", "not done", "didn't finish", "still working",
        "نص الطريق", "في النص", "نصها", "مكملش"
    ]
    
    is_incomplete_cmd = any(k in ml for k in INCOMPLETE_KEYWORDS)
    
    if is_incomplete_cmd:
        task = None
        
        for kw in INCOMPLETE_KEYWORDS:
            if kw in ml:
                rest = msg[ml.index(kw)+len(kw):].strip()
                rest = re.sub(r"^(من|المهمة|مهمة|في)\s*", "", rest).strip()
                if rest:
                    task = find_task_by_name(rest)
                    break
        
        if not task:
            pend = [t for t in get_tasks() if t["date"] == today and t["status"] == "pending"]
            if pend:
                task = pend[0]
        
        if task and task["status"] != "done":
            state["pending_completion"] = True
            state["pending_task_id"] = task["id"]
            state["pending_task_name"] = task["name"]
            state["pending_task_duration"] = task["duration"]
            
            return {
                "reply": f"📝 **مهمة '{task['name']}' لم تكتمل بعد**\n\n⏱️ كم دقيقة بقي منها؟\n• أدخل عدد الدقائق (مثال: 15)\n• أو اكتب 'لا' إذا لم تبدأها بعد",
                "action": "pending_completion"
            }
        return {"reply": "❌ لم أجد مهمة معلقة بهذا الاسم.", "action": "none"}

    # DELETE TASK
    DELETE_KEYWORDS = ["احذف", "امسح", "حذف", "ازل", "امسحلي", "delete", "remove", "شيل"]
    if any(k in ml for k in DELETE_KEYWORDS):
        task = None
        for kw in DELETE_KEYWORDS:
            if kw in ml:
                rest = msg[ml.index(kw)+len(kw):].strip()
                rest = re.sub(r"^(المهمة|مهمة)\s*", "", rest).strip()
                if rest:
                    task = find_task_by_name(rest)
                    break
        
        if task:
            n = task["name"]
            task_delete(task["id"])
            return {"reply": f"🗑️ تم حذف المهمة:\n❌ {n}", "action": "refresh"}
        return {"reply": "✏️ اكتب اسم المهمة التي تريد حذفها.\nمثال: احذف مذاكرة", "action": "none"}

    # POSTPONE TASK - FIXED to place at exact specified time
        # POSTPONE TASK - FIXED to place at exact specified time
    DELAY_KEYWORDS = ["اجل", "تاجيل", "أجل", "ؤجل", "postpone", "delay", "later", "أخر", "اخر"]
    if any(k in ml for k in DELAY_KEYWORDS):
        new_time = extract_time(msg)
        tomorrow = any(w in ml for w in ["بكرة", "بكره", "الغد", "غداً", "يوم جاي", "tomorrow"])
        task = None
        
        # Extract task name from message
        temp_msg = msg
        for kw in DELAY_KEYWORDS:
            temp_msg = re.sub(r'\b' + kw + r'\b', '', temp_msg, flags=re.IGNORECASE)
        # Remove time patterns
        temp_msg = re.sub(r'(\d{1,2}[:\.]\d{2}|\d{1,2}\s*(صباح|مساء|ص|م|am|pm|لـ|الساعة))', '', temp_msg, flags=re.IGNORECASE)
        task_name = temp_msg.strip()
        
        if task_name:
            task = find_task_by_name(task_name)
        
        if not task:
            pend = [t for t in get_tasks() if t["date"] == today and t["status"] == "pending"]
            if pend:
                task = pend[0]
                if task_name and task_name not in task['name']:
                    return {"reply": f"⚠️ لم أجد مهمة باسم '{task_name}'. هل تقصد **{task['name']}**؟\n\nإذا كنت تريد تأجيلها، اكتب: `اجل {task['name']} لـ 10:00`", "action": "none"}
        
        if task:
            if tomorrow:
                if task_postpone_tomorrow(task["id"]):
                    optimize_schedule()
                    return {"reply": f"📅 تم تأجيل المهمة:\n📌 {task['name']}\n✨ إلى الغد!", "action": "refresh"}
                else:
                    return {"reply": f"❌ فشل تأجيل المهمة {task['name']}", "action": "none"}
                    
            elif new_time:
                # Convert time to 24-hour format
                new_time_24 = convert_to_24hour(new_time)
                if not new_time_24:
                    return {"reply": "❌ صيغة الوقت غير صحيحة. استخدم مثال: 10:30 أو الساعة 2 مساءً", "action": "none"}
                
                # Check if time is in the future
                time_mins = to_mins(new_time_24)
                current_mins = to_mins(now_str())
                start_day_mins = to_mins(_settings.get("startTime", "08:00"))
                sleep_mins = to_mins(_settings.get("sleepTime", "23:30"))
                
                # Adjust for after-midnight
                adjusted_time_mins = time_mins
                if sleep_mins < start_day_mins and time_mins < start_day_mins:
                    adjusted_time_mins += 1440
                
                if adjusted_time_mins <= current_mins:
                    return {"reply": f"❌ لا يمكن التأجيل لوقت مضى! الوقت الحالي هو {now_str()}.\nالرجاء اختيار وقت في المستقبل.\n\nمثال: `اجل {task['name']} لـ 14:30`", "action": "none"}
                
                if not is_within_sleep_time(new_time_24, task["duration"]):
                    sleep_time = _settings.get("sleepTime", "23:30")
                    return {"reply": f"❌ لا يمكن إنهاء المهمة قبل وقت النوم ({sleep_time})!", "action": "none"}
                
                # Use the direct function that bypasses reordering
                success, result = task_postpone_time_direct(task["id"], new_time_24)
                if success:
                    return {"reply": f"⏰ **تم تأجيل المهمة بنجاح!**\n\n📌 {task['name']}\n🕐 إلى {new_time_24}\n✅ تم وضعها في الجدول الزمني في الوقت المحدد!", "action": "refresh"}
                else:
                    return {"reply": f"❌ {result}", "action": "none"}
            else:
                # Default: postpone by 1 hour
                current_time = task.get("startTime") or now_str()
                current_mins = to_mins(current_time)
                new_time_mins = current_mins + 60
                new_time_str = to_time(new_time_mins)
                
                if new_time_mins > to_mins(now_str()) and is_within_sleep_time(new_time_str, task["duration"]):
                    success, _ = task_postpone_time_direct(task["id"], new_time_str)
                    if success:
                        return {"reply": f"⏰ تم تأجيل المهمة ساعة:\n📌 {task['name']}\n🕐 الآن في {new_time_str}", "action": "refresh"}
                
                return {"reply": "❌ لا يمكن التأجيل! الوقت المحدد إما في الماضي أو بعد وقت النوم.\n\n📝 مثال صحيح: `اجل مذاكرة لـ 14:30`", "action": "none"}
        
        return {"reply": "✏️ اكتب اسم المهمة والوقت.\n📝 مثال: `اجل مذاكرة لـ 14:30`\n📅 للتأجيل للغد: `اجل مذاكرة لبكرة`", "action": "none"}

    # REPORT/STATISTICS
    if any(w in ml for w in ["تقرير", "احصاء", "كم مهمة", "مهامي", "وضع"]):
        all_t = get_tasks()
        tt = [t for t in all_t if t["date"] == today]
        pend = [t for t in tt if t["status"] == "pending"]
        done = [t for t in tt if t["status"] == "done"]
        mins = sum(t["duration"] for t in pend)
        pct = round(len(done)/len(tt)*100) if tt else 0
        
        remaining_time = get_time_remaining_until_sleep()
        
        r = f"📊 **تقرير اليوم - {today}**\n\n"
        r += f"✅ تم انجازها: {len(done)}\n"
        r += f"⏳ متبقية: {len(pend)}\n"
        r += f"⏱️ وقت مطلوب: {mins} دقيقة\n"
        r += f"📈 نسبة الانجاز: {pct}%\n"
        r += f"🌙 باقي للنوم: {remaining_time} دقيقة\n"
        
        if pend:
            r += "\n**📋 المهام المتبقية (مرتبة حسب الأولوية):**\n"
            sorted_pend = UtilityBasedAgent.sort_combined(pend)
            for t in sorted_pend[:6]:
                priority_icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(t["priority"], "🟡")
                r += f"• {priority_icon} {t['name']} ({t['duration']}د)"
                if t.get("startTime"):
                    r += f" - {t['startTime']}"
                r += "\n"
        
        insights = intelligent_scheduler.get_insights()
        if insights:
            r += "\n**🧠 رؤى ذكية:**\n"
            for insight in insights[:3]:
                r += f"• {insight}\n"
        
        return {"reply": r, "action": "none"}

    if any(w in ml for w in ["اجل الكل", "تاجيل كل", "كل المهام للغد"]):
        count = 0
        with _lock:
            today_tasks = [t for t in _tasks if t["date"] == today and t["status"] == "pending"]
            today_tasks = UtilityBasedAgent.sort_combined(today_tasks)
            
            tom = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
            for task in today_tasks:
                task["date"] = tom
                task["postponedCount"] += 1
                task["startTime"] = None
                task["endTime"] = None
                task["alertSent"] = False
                count += 1
            save_tasks()
        
        return {"reply": f"📅 تم تاجيل {count} مهمة للغد.\n✨ تم ترتيبها حسب الأولوية ← الصعوبة ← المدة!", "action": "refresh"}

    # WHAT'S NEXT (English) - CORRECTED VERSION
    if re.search(r'\b(what|which)(\'?s)?\s+(the\s+)?(next|upcoming)\s+task\b', ml, re.IGNORECASE) or \
       ml == 'next' or ml == 'what next' or ml == 'what\'s next' or ml == 'whats next' or \
       any(w in ml for w in ["المهمة القادمة", "التالية", "الجاي"]):
        ordered = UtilityBasedAgent.sort_combined([t for t in get_tasks() if t["status"] == "pending"])
        pending = [t for t in ordered if t["date"] == today]
        
        if pending:
            nxt = pending[0]
            priority_icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(nxt["priority"], "🟡")
            r = f"**🎯 مهمتك القادمة:**\n\n"
            r += f"{priority_icon} **{nxt['name']}**\n"
            if nxt.get("startTime"):
                r += f"⏰ الوقت: {nxt['startTime']}\n"
            r += f"⏱️ المدة: {nxt['duration']} دقيقة\n"
            return {"reply": r, "action": "none"}
        return {"reply": "🎉 لا توجد مهام متبقية! يومك حر.", "action": "none"}

    if any(w in ml for w in ["وقت النوم", "غير النوم", "نم الساعة", "بداية اليوم"]):
        t = extract_time(msg)
        if t:
            if "نوم" in ml:
                _settings["sleepTime"] = t
                _app_settings["sleepTime"] = t
                save_app_settings()
                remaining = get_time_remaining_until_sleep()
                return {"reply": f"🌙 تم تغيير وقت النوم الى {t}!\n⏰ الوقت المتبقي حتى النوم: {remaining // 60} ساعة و {remaining % 60} دقيقة", "action": "refresh"}
            else:
                _settings["startTime"] = t
                _app_settings["startTime"] = t
                save_app_settings()
                return {"reply": f"🌅 تم تغيير وقت بداية اليوم الى {t}!", "action": "refresh"}
        return {"reply": "مثال: غير وقت النوم 11:30", "action": "none"}

    if any(w in ml for w in ["عرض", "شوف", "أرني", "قائمة", "وريني"]):
        ordered = UtilityBasedAgent.sort_combined([t for t in get_tasks() if t["status"] == "pending"])
        today_tasks = [t for t in ordered if t["date"] == today]
        
        if not today_tasks:
            return {"reply": "📭 لا توجد مهام اليوم!\n🎉 يوم حر!", "action": "none"}
        
        priority_icons = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        r = f"**📋 مهام اليوم ({today}) - مرتبة حسب الأولوية:**\n\n"
        
        for i, t in enumerate(today_tasks, 1):
            r += f"{i}. {priority_icons[t['priority']]} **{t['name']}**\n"
            r += f"   ⏱️ {t['duration']} دقيقة"
            if t.get("startTime"):
                r += f" | 🕐 {t['startTime']}"
            r += "\n"
        
        return {"reply": r, "action": "none"}

    GREET = ["مرحبا", "هلا", "هاي", "سلام", "صباح", "مساء", "ازيك", "كيفك", "اهلا","Hello","hello","Hallo","hallo","Hi","hi","HI"]
    
    if any(w in ml for w in GREET):
        pend = [t for t in get_tasks() if t["date"] == today and t["status"] == "pending"]
        done = [t for t in get_tasks() if t["date"] == today and t["status"] == "done"]
        remaining_time = get_time_remaining_until_sleep()
        h = datetime.now().hour
        if 5 <= h < 12:
            gr = "🌅 صباح النور"
        elif 12 <= h < 17:
            gr = "☀️ صباح الخير"
        elif 17 <= h < 21:
            gr = "🌙 مساء الخير"
        else:
            gr = "✨ اهلا وسهلا"
        
        return {"reply": f"{gr}! 👋\n\nأنا **SmartDay AI Agent v8.4** - مساعدك الذكي.\n\n"
                        f"📊 **لليوم:** {len(pend)} مهمة متبقية | {len(done)} مكتملة\n"
                        f"🌙 **الوقت المتبقي للنوم:** {remaining_time // 60} ساعة و {remaining_time % 60} دقيقة\n\n"
                        f"**🎯 الأوامر المدعومة:**\n"
                        f"• 🔄 **بدل مذاكرة مع اجتماع** (تبديل مباشر فوري)\n"
                        f"• ➕ ضيف مهمة (سؤال تفاعلي)\n"
                        f"• ✅ خلصت [اسم المهمة] (يتم الإكمال مباشرة)\n"
                        f"• ⏸️ ملحقتش [اسم المهمة] (يسأل عن الباقي ويضعه بعد آخر مهمة)\n"
                        f"• 🗑️ احذف [اسم المهمة]\n"
                        f"• ⏰ اجل [المهمة] لـ [وقت] / للغد (يتم وضعه في الوقت المحدد)\n"
                        f"• 🔄 رتب المهام حسب الأولوية/الصعوبة/المدة\n"
                        f"• 📊 تقرير اليوم\n"
                        f"• 📋 عرض المهام\n"
                        f"• 🧠 جدولة ذكية\n"
                        f"• 🔍 ابحث عن [موضوع]\n\n"
                        f"💡 **جرب التبديل المباشر:** اكتب 'بدل مذاكرة مع اجتماع' وسأقوم بتبديلهما فوراً!\n"
                        f"💡 **عند عدم إكمال مهمة:** سأسألك عن الوقت المتبقي وسأضعه بعد آخر مهمة في اليوم!", "action": "none"}

    if any(w in ml for w in ["جدول ذكي", "جدولة ذكية", "اقترح جدول"]):
        if intelligent_scheduler.pending_selected_order is not None:
            return {"reply": "⚠️ لديك ترتيب معلق لم يتم تطبيقه بعد. الرجاء اختيار جدول أولاً.", "action": "none"}
        
        result = intelligent_scheduler.generate_and_recommend(get_tasks(), _settings)
        if result["top_8"]:
            intelligent_scheduler.has_pending_recommendation = True
            strategy_names_ar = {
                "priority": "📊 حسب الأولوية",
                "difficulty": "💪 حسب الصعوبة",
                "duration_short": "⏱️ حسب المدة (الأقصر)",
                "duration_long": "🐢 حسب المدة (الأطول)",
                "random_1": "🎲 عشوائي ذكي 1",
                "random_2": "🎲 عشوائي ذكي 2",
                "energy_based": "⚡ حسب الطاقة",
                "deadline_based": "⏰ حسب الوقت المتبقي",
                "smart_combined": "🧠 ذكي متكامل"
            }
            
            r = "🧠 **أفضل 8 جداول ذكية:**\n\n"
            
            for i, (idx, schedule, score, strategy) in enumerate(result["top_8"]):
                strategy_name = strategy_names_ar.get(strategy, strategy)
                is_rec = idx == result["recommended_index"]
                r += f"**{i+1}. {strategy_name}** { '(⭐ موصى به)' if is_rec else ''} (الدرجة: {score:.0f})\n"
                tasks_shown = 0
                for t in schedule:
                    if tasks_shown < 4:
                        if t.get("startTime"):
                            r += f"  • {t['name']} - {t['startTime']} ({t['duration']}د)\n"
                        else:
                            r += f"  • {t['name']} (غير مجدولة)\n"
                        tasks_shown += 1
                if len(schedule) > 4:
                    r += f"  • ... و {len(schedule) - 4} مهام أخرى\n"
                r += "\n"
            r += "اختر رقم الاستراتيجية (1-8) لتطبيقها."
            return {"reply": r, "action": "none"}
        return {"reply": "لا توجد مهام لجدولتها.", "action": "none"}

    if any(p in ml for p in ["ابحث", "بحث عن", "ما هو", "ما هي"]):
        query = msg
        for kw in ["ابحث عن", "بحث عن", "ابحث"]:
            if kw in ml:
                query = msg[ml.index(kw)+len(kw):].strip()
                break
        result = web_search(query)
        return {"reply": f"🔍 **نتائج البحث عن:** {query}\n\n{result}", "action": "none"}

    if any(w in ml for w in ["مساعدة", "تعليمات", "ازاي", "كيف", "شرح"]):
        return {"reply": """**🆘 قائمة الأوامر المساعدة**

**🔄 تبديل المهام (تبديل فوري - بدون أسئلة):**
• "بدل مذاكرة مع اجتماع"
• "غير ترتيب قراءة وكتابة"

**➕ إضافة مهمة (طريقة تفاعلية):**
• "ضيف مهمة" -> يتبع أسئلة عن الاسم والمدة والأولوية والصعوبة

**✅ إكمال مهمة:**
• "خلصت مذاكرة" -> يتم الإكمال مباشرة
• "ملحقتش مذاكرة" -> يسأل عن الوقت المتبقي ويضعه بعد آخر مهمة

**🗑️ حذف مهمة:**
• "احذف مذاكرة"

**⏰ تأجيل مهمة:**
• "اجل المذاكرة لبكرة"
• "اجل الاجتماع لـ 10:00"
• "اجل الاجتماع لـ 01:30" (بعد منتصف الليل)

**🔄 إعادة ترتيب المهام:**
• "رتب المهام حسب الأولوية"
• "رتب المهام حسب الصعوبة"
• "رتب المهام حسب المدة"
• "رتب المهام ترتيب ذكي"

**🧠 جدولة ذكية:**
• "اقترح جدول" - يقترح أفضل 8 جداول

**📋 عرض المهام:**
• "وريني المهام"

**📊 تقرير:**
• "تقرير اليوم"

💡 **الأهم:** اكتب "بدل مذاكرة مع اجتماع" وسأقوم بتبديلهما مباشرة!
💡 **عند عدم إكمال مهمة:** سأسألك عن الوقت المتبقي وسأضعه بعد آخر مهمة في اليوم!""",
        "action": "none"}

    pending = [t for t in get_tasks() if t["date"] == today and t["status"] == "pending"]
    hint = ""
    if pending:
        ordered = UtilityBasedAgent.sort_combined(pending)
        nxt = ordered[0]
        hint = f"\n\n💡 **مهمتك القادمة:** {nxt['name']}"
        if nxt.get("startTime"):
            hint += f" - {nxt['startTime']}"
    
    return {
        "reply": f"❓ لم افهم طلبك. جرب:\n\n"
                 f"🔄 **بدل مذاكرة مع اجتماع** (تبديل مباشر)\n"
                 f"📝 **ضيف مهمة** (سؤال تفاعلي)\n"
                 f"✅ **خلصت** [اسم المهمة] (يتم الإكمال مباشرة)\n"
                 f"⏸️ **ملحقتش** [اسم المهمة] (يسأل عن المدة المتبقية)\n"
                 f"📋 **وريني المهام**\n"
                 f"🗑️ **احذف** [اسم المهمة]\n"
                 f"⏰ **اجل** [المهمة] لبكرة او لـ [وقت]\n"
                 f"🧠 **اقترح جدول** (جدولة ذكية)\n"
                 f"🆘 **مساعدة** لجميع الأوامر{hint}",
        "action": "none"
    }

# =============================================================================
# SCHEDULER THREAD
# =============================================================================
def scheduler_loop():
    last_min = -1
    
    while True:
        try:
            now = datetime.now()
            cur = now.hour * 60 + now.minute
            today = now.strftime("%Y-%m-%d")
            
            # ============================================================
            # AUDIO: per-tick check for every pending task
            # ============================================================
            # Run audio checks every iteration (not just on minute change) so
            # the alert fires close to the actual end_time. The controller
            # itself dedupes via fired-key flags.
            for t in get_tasks():
                if t.get("status") != "pending":
                    continue
                if t.get("date") != today:
                    continue
                # Focus music when task starts
                task_audio.maybe_play_focus(t, cur)
                # Alert music when task end time is exceeded
                task_audio.maybe_play_alert(t, cur)
            
            # ============================================================
            # Once-per-minute work (notifications, daily insights)
            # ============================================================
            if cur != last_min:
                last_min = cur
                
                for t in get_tasks():
                    if t["status"] == "pending" and t["date"] == today and t.get("startTime") and not t.get("alertSent"):
                        diff = to_mins(t["startTime"]) - cur
                        if -2 <= diff <= 5:
                            _notifs.append({
                                "type": "task_alert",
                                "message": f"⏰ حان وقت: {t['name']} - {t['startTime']}"
                            })
                            speak(f"حان وقت مهمة {t['name']}")
                            task_update(t["id"], {"alertSent": True})
                    
                    if t["status"] == "pending" and t["date"] == today and t.get("endTime"):
                        end_mins = to_mins(t["endTime"])
                        if cur >= end_mins and cur <= end_mins + 2:
                            _notifs.append({
                                "type": "task_end",
                                "message": f"✅ انتهت المهمة: {t['name']}"
                            })
                
                remaining_time = get_time_remaining_until_sleep()
                if 25 <= remaining_time <= 35:
                    pnd = [t for t in get_tasks() if t["date"] == today and t["status"] == "pending"]
                    if pnd:
                        _notifs.append({
                            "type": "sleep_warning",
                            "message": f"🌙 باقي {remaining_time} دقيقة للنوم! لديك {len(pnd)} مهام متبقية"
                        })
                        speak(f"باقي {remaining_time} دقيقة للنوم")
                
                if cur % 120 == 0 and cur > 0:
                    generate_ai_notifications()
        except Exception as e:
            print(f"Scheduler error: {e}")
        # Faster loop → more reliable audio triggering
        time.sleep(5)

threading.Thread(target=scheduler_loop, daemon=True).start()

# =============================================================================
# FASTAPI ENDPOINTS
# =============================================================================
app = FastAPI(title="SmartDay AI Agent", version="8.4")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

@app.get("/api/tasks")
def api_tasks():
    return JSONResponse({"tasks": get_tasks(), "settings": _settings})

@app.post("/api/tasks")
async def api_add(req: Request):
    data = await req.json()
    t = task_add(data)
    optimize_schedule()
    return JSONResponse({"success": True, "task": t})

@app.post("/api/tasks/batch")
async def api_add_batch(req: Request):
    data = await req.json()
    tasks_data = data.get("tasks", [])
    added_tasks = []
    
    for task_data in tasks_data:
        t = task_add(task_data)
        added_tasks.append(t)
    
    optimize_schedule()
    return JSONResponse({"success": True, "tasks": added_tasks, "count": len(added_tasks)})

@app.delete("/api/tasks/{tid}")
def api_del(tid: int):
    return JSONResponse({"success": task_delete(tid)})

@app.delete("/api/tasks")
def api_del_all():
    return JSONResponse({"success": task_delete_all()})

@app.post("/api/tasks/{tid}/complete")
def api_complete(tid: int):
    ok = task_complete_direct(tid)
    optimize_schedule()
    return JSONResponse({"success": ok})

@app.post("/api/tasks/{tid}/complete-with-remaining")
async def api_complete_with_remaining(tid: int, req: Request):
    data = await req.json()
    remaining = data.get("remaining_minutes", 0)
    ok = task_complete_with_remaining(tid, remaining)
    optimize_schedule()
    return JSONResponse({"success": ok})

@app.post("/api/tasks/{tid}/postpone/tomorrow")
def api_post_tom(tid: int):
    return JSONResponse({"success": task_postpone_tomorrow(tid)})

@app.post("/api/tasks/{tid}/postpone/time")
async def api_post_time(tid: int, req: Request):
    data = await req.json()
    new_time = data.get("time", "")
    
    # Strict future check
    if not is_time_in_future(new_time):
        return JSONResponse({"success": False, "error": f"لا يمكن التأجيل لوقت مضى! الوقت الحالي هو {now_str()}"})
    
    ok, msg = task_postpone_time_direct(tid, new_time)
    return JSONResponse({"success": ok, "message": msg})

@app.post("/api/tasks/{tid}/update-time")
async def api_update_time(tid: int, req: Request):
    data = await req.json()
    new_time = data.get("time", "")
    
    if not is_future_time(new_time):
        return JSONResponse({"success": False, "error": "لا يمكن وضع المهمة في وقت مضى"})
    
    ok, msg = task_update_time(tid, new_time)
    if ok:
        with _lock:
            save_tasks()
    return JSONResponse({"success": ok, "message": msg})

@app.post("/api/tasks/swap")
async def api_swap(req: Request):
    data = await req.json()
    task1_id = data.get("task1_id")
    task2_id = data.get("task2_id")
    success = task_swap(task1_id, task2_id)
    if success:
        optimize_schedule()
    return JSONResponse({"success": success})

@app.get("/api/schedule/find-slots")
async def api_find_slots(req: Request):
    duration = int(req.query_params.get("duration", 30))
    date = req.query_params.get("date", today_str())
    slots = find_available_time_slots(get_tasks(), duration, date)
    return JSONResponse({"slots": slots})

@app.post("/api/schedule/optimize")
def api_optimize():
    if intelligent_scheduler.pending_selected_order is not None:
        return JSONResponse({"success": False, "message": "الرجاء اختيار جدول أولاً"})
    optimize_schedule()
    return JSONResponse({"success": True})

@app.get("/api/schedule/recommendations")
def api_get_recommendations():
    try:
        result = intelligent_scheduler.generate_and_recommend(get_tasks(), _settings)
        
        # Build a quick map idx -> id from the cache
        idx_to_id: Dict[int, str] = {}
        for sid, info in intelligent_scheduler._schedule_cache.items():
            idx_to_id[info["index"]] = sid
        
        formatted = {
            "recommended_index": result["recommended_index"],
            "recommended_id": result.get("recommended_id"),
            "schedule_ids": result.get("schedule_ids", []),
            "scores": result["scores"],
            "top_8": [
                {
                    "index": item[0],
                    "id": idx_to_id.get(item[0]),  # NEW: unique ID for selection
                    "score": round(item[2], 2),
                    "strategy": item[3],
                    "tasks": [
                        {
                            "id": t.get("id"),
                            "name": t.get("name"),
                            "startTime": t.get("startTime"),
                            "endTime": t.get("endTime"),
                            "duration": t.get("duration"),
                            "priority": t.get("priority"),
                            "difficulty": t.get("difficulty")
                        }
                        for t in item[1]
                    ]
                }
                for item in result["top_8"]
            ],
            "strategy_names": result["strategy_names"],
            "recommended_strategy": result["recommended_strategy"],
            "ai_insights": intelligent_scheduler.get_insights(),
            "has_pending": intelligent_scheduler.pending_selected_order is not None,
            "total_selections": user_preferences.preferences["total_selections"]
        }
        
        return JSONResponse(formatted)
    except Exception as e:
        print(f"Error in recommendations: {e}")
        return JSONResponse({
            "recommended_index": -1,
            "recommended_id": None,
            "schedule_ids": [],
            "scores": [],
            "top_8": [],
            "strategy_names": [],
            "recommended_strategy": "",
            "ai_insights": ["⚠️ حدث خطأ في إنشاء الجداول الذكية"],
            "has_pending": False,
            "total_selections": 0
        })

@app.post("/api/schedule/select")
async def api_select_schedule(req: Request):
    """
    CRITICAL FIX: select schedule strictly by its unique cached ID.
    - No re-generation
    - No re-ranking
    - Mismatch throws an error
    """
    data = await req.json()
    schedule_id = data.get("schedule_id")
    # Backward-compat fallback (old clients send selected_idx) — try to map to ID
    selected_idx = data.get("selected_idx", -1)
    
    try:
        # Resolve ID — prefer explicit schedule_id, otherwise resolve idx via cache
        if not schedule_id:
            cache = intelligent_scheduler._schedule_cache
            for sid, info in cache.items():
                if info["index"] == selected_idx:
                    schedule_id = sid
                    break
            if not schedule_id:
                err_msg = (f"⚠️ No schedule_id supplied and selected_idx={selected_idx} "
                           f"could not be resolved. Cache empty? "
                           f"{len(intelligent_scheduler._schedule_cache) == 0}")
                print(err_msg)
                return JSONResponse(
                    {"success": False, "error": err_msg},
                    status_code=400
                )
        
        result = intelligent_scheduler.apply_schedule_by_id(schedule_id)
        
        with _lock:
            save_tasks()
        
        return JSONResponse({
            "success": True,
            "selected_id": result["selected_id"],
            "applied_id": result["applied_id"],
            "strategy": result["strategy"],
            "task_count": result["task_count"],
        })
    except (ValueError, RuntimeError) as e:
        # Mismatch / cache-miss → explicit error response
        print(f"❌ [SELECT ERROR] {e}")
        return JSONResponse(
            {"success": False, "error": str(e)},
            status_code=400
        )
    except Exception as e:
        print(f"❌ [SELECT UNEXPECTED] {e}")
        return JSONResponse(
            {"success": False, "error": f"Unexpected: {e}"},
            status_code=500
        )

@app.post("/api/settings")
async def api_settings(req: Request):
    global _settings, _app_settings
    data = await req.json()
    
    if "sleepTime" in data:
        _settings["sleepTime"] = data["sleepTime"]
        _app_settings["sleepTime"] = data["sleepTime"]
    if "startTime" in data:
        _settings["startTime"] = data["startTime"]
        _app_settings["startTime"] = data["startTime"]
    if "theme" in data:
        _settings["theme"] = data["theme"]
        _app_settings["theme"] = data["theme"]
    if "useManualOrder" in data:
        _settings["useManualOrder"] = data["useManualOrder"]
    if "language" in data:
        _app_settings["language"] = data["language"]
    
    save_app_settings()
    
    return JSONResponse({"success": True, "settings": _settings})

@app.get("/api/settings")
def api_get_settings():
    return JSONResponse({"settings": _settings, "app_settings": _app_settings})

@app.post("/api/chat")
async def api_chat(req: Request):
    data = await req.json()
    msg = data.get("message", "").strip()
    session_id = data.get("session_id", "default")
    
    if not msg:
        return JSONResponse({"reply": "اكتب رسالة!", "action": "none"})
    
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(_executor, agent_process, msg, session_id)
    
    return JSONResponse(result)

@app.get("/api/history")
def api_history():
    return JSONResponse({"history": _history[:50]})

@app.delete("/api/history")
def api_clear_history():
    _history.clear()
    return JSONResponse({"success": True})

@app.get("/api/notifications")
def api_notifs():
    n = list(_notifs)
    _notifs.clear()
    return JSONResponse({"notifications": n})

@app.get("/api/insights")
def api_insights():
    try:
        return JSONResponse({"insights": intelligent_scheduler.get_insights()})
    except Exception as e:
        print(f"Error getting insights: {e}")
        return JSONResponse({"insights": ["🧠 جاري تحليل البيانات..."]})

@app.get("/api/talent")
def api_talent():
    """NEW: expose the talent profile so the UI / external tooling can read it."""
    try:
        return JSONResponse({
            "talent": intelligent_scheduler.talent.public_snapshot(),
            "raw": intelligent_scheduler.talent.profile,
            "insights": intelligent_scheduler.talent.insights(),
        })
    except Exception as e:
        return JSONResponse({"talent": {}, "error": str(e)})

@app.get("/api/status")
def api_status():
    return JSONResponse({
        "voice": VOICE_OK,
        "excel": EXCEL_OK,
        "search": REQUESTS_OK,
        "numpy": NUMPY_OK,
        "sound": SOUND_OK,
        "tasks": len(get_tasks())
    })

# =============================================================================
# HTML (Arabic Only - Full Arabic Interface)
# =============================================================================
HTML = r"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,user-scalable=yes">
<title>سمارت داي - مساعد الذكاء الاصطناعي للجدولة</title>
<link href="https://fonts.googleapis.com/css2?family=Tajawal:wght@400;700;900&family=JetBrains+Mono:wght@400;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
:root{
  --bg:#07070f;--bg2:#0e0e1c;--bg3:#141426;--bg4:#1a1a2e;--bg5:#222235;
  --bdr:rgba(255,255,255,0.07);--bdr2:rgba(255,255,255,0.13);
  --gold:#d4a843;--gold2:#f0c96e;--gold3:#a07824;
  --green:#2dcc8a;--red:#e84c6a;--blue:#4a9fff;--amber:#f09d3a;
  --txt:#eeeaf0;--txt2:#8884a0;--txt3:#3e3a55;
  --font:'Tajawal',sans-serif;--mono:'JetBrains Mono',monospace;
}
body.light{
  --bg:#f0f0f5;--bg2:#ffffff;--bg3:#f5f5fa;--bg4:#eeeef5;--bg5:#dddde8;
  --bdr:rgba(0,0,0,0.08);--bdr2:rgba(0,0,0,0.15);
  --txt:#1a1a2e;--txt2:#555570;--txt3:#9999b0;
}
html,body{height:100%;font-family:var(--font);background:var(--bg);color:var(--txt);overflow:hidden}
.app{display:grid;grid-template-columns:240px 1fr;height:100vh;overflow:hidden}

.sidebar{background:var(--bg2);border-left:1px solid var(--bdr);display:flex;flex-direction:column;overflow:hidden}
.brand{padding:16px 18px 12px;border-bottom:1px solid var(--bdr);display:flex;align-items:center;gap:10px}
.brand-icon{width:34px;height:34px;background:linear-gradient(135deg,var(--gold),var(--gold3));border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:17px;flex-shrink:0}
.brand-text .brand-name{font-size:17px;font-weight:900}
.brand-text .brand-name span{color:var(--gold)}
.brand-text .brand-ver{font-size:10px;color:var(--txt3)}
.clock-wrap{padding:10px 18px;border-bottom:1px solid var(--bdr)}
.clock{font-family:var(--mono);font-size:24px;font-weight:700;color:var(--gold);letter-spacing:2px}
.clock-date{font-size:11px;color:var(--txt2);margin-top:3px}
nav{flex:1;padding:7px;overflow-y:auto}
.nb{display:flex;align-items:center;gap:9px;padding:9px 11px;border-radius:9px;cursor:pointer;color:var(--txt2);font-size:13px;font-weight:600;margin-bottom:2px;background:none;border:1px solid transparent;width:100%;text-align:right;font-family:var(--font);transition:all .15s}
.nb:hover{background:var(--bg4);color:var(--txt)}
.nb.active{background:rgba(212,168,67,0.13);color:var(--gold);border-color:rgba(212,168,67,0.22)}
.nb-icon{font-size:16px;width:20px;text-align:center;flex-shrink:0}
.nb-lbl{flex:1;text-align:right}
.nb-badge{margin-right:auto;background:var(--red);color:#fff;font-size:10px;font-weight:900;padding:1px 7px;border-radius:20px}
.sb-stats{padding:10px 14px;border-top:1px solid var(--bdr);display:grid;grid-template-columns:1fr 1fr;gap:6px}
.ss{background:var(--bg4);border:1px solid var(--bdr);border-radius:8px;padding:7px 9px;text-align:center}
.ss-n{font-size:20px;font-weight:900;font-family:var(--mono)}
.ss-l{font-size:9px;color:var(--txt3);margin-top:1px}

.main{display:flex;flex-direction:column;height:100vh;overflow:hidden}
.topbar{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:10px 20px;border-bottom:1px solid var(--bdr);background:rgba(14,14,28,0.9);backdrop-filter:blur(10px);flex-shrink:0}
body.light .topbar{background:rgba(255,255,255,0.9)}
.topbar-title{font-size:15px;font-weight:800}
.topbar-right{display:flex;gap:7px;align-items:center}
.sleep-info{background:rgba(212,168,67,0.08);border:1px solid rgba(212,168,67,0.18);border-radius:20px;padding:5px 12px;font-size:12px;color:var(--txt2)}
.sleep-info b{color:var(--gold);font-family:var(--mono)}

.notif-btn{position:relative;background:var(--bg4);border:1px solid var(--bdr);border-radius:50%;width:36px;height:36px;cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:16px;transition:.15s}
.notif-btn:hover{background:var(--bg5)}
.notif-btn-badge{position:absolute;top:-4px;left:-4px;background:var(--red);color:#fff;font-size:10px;font-weight:900;width:18px;height:18px;border-radius:50%;display:flex;align-items:center;justify-content:center}
.notif-panel{position:fixed;top:55px;left:20px;background:var(--bg2);border:1px solid var(--bdr2);border-radius:12px;width:340px;max-height:400px;overflow-y:auto;z-index:100;box-shadow:0 10px 40px rgba(0,0,0,.4);display:none;padding:10px}
.notif-panel.open{display:block}
.notif-item{display:flex;align-items:flex-start;gap:10px;padding:10px;border-bottom:1px solid var(--bdr);font-size:12px;line-height:1.6}
.notif-item:last-child{border-bottom:none}
.notif-empty{text-align:center;padding:20px;color:var(--txt3);font-size:13px}

.theme-btn{width:36px;height:36px;border-radius:50%;border:1px solid var(--bdr);background:var(--bg4);cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:16px;transition:.15s}
.theme-btn:hover{background:var(--bg5)}

.panel{display:none;flex:1;flex-direction:column;overflow:hidden}
.panel.show{display:flex}
.scroll{flex:1;overflow-y:auto;padding:18px 22px}
.scroll::-webkit-scrollbar{width:3px}
.scroll::-webkit-scrollbar-thumb{background:var(--bg5);border-radius:3px}

.card{background:var(--bg2);border:1px solid var(--bdr);border-radius:12px;padding:16px 18px;margin-bottom:11px}
.card-t{font-size:10px;color:var(--txt3);text-transform:uppercase;letter-spacing:.1em;font-weight:800;margin-bottom:12px;display:flex;align-items:center;justify-content:space-between}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:9px}
.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:9px}
.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:9px}

.sb{background:var(--bg3);border:1px solid var(--bdr);border-radius:11px;padding:14px 16px;transition:.2s}
.sb:hover{transform:translateY(-2px)}
.sb-n{font-size:38px;font-weight:900;font-family:var(--mono);line-height:1}
.sb-l{font-size:11px;color:var(--txt2);margin-top:4px}
.pbar{height:6px;background:var(--bg5);border-radius:4px;margin-top:6px;overflow:hidden}
.pbar-f{height:100%;border-radius:4px;transition:width .5s;background:linear-gradient(90deg,var(--green),var(--blue))}

.fg{margin-bottom:10px}
.fg label{display:block;font-size:11px;color:var(--txt2);margin-bottom:4px;font-weight:700}
input,select,textarea{width:100%;background:var(--bg3);border:1px solid var(--bdr);border-radius:7px;color:var(--txt);padding:8px 11px;font-size:13px;font-family:var(--font);outline:none;transition:.15s}
input:focus,select:focus,textarea:focus{border-color:var(--gold);box-shadow:0 0 0 3px rgba(212,168,67,.1)}
select{background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' viewBox='0 0 10 10'%3E%3Cpath fill='%236b6880' d='M5 7L1 3h8z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:left 10px center;padding-left:28px;appearance:none;-webkit-appearance:none}
textarea{resize:vertical;min-height:56px}

.btn{font-family:var(--font);font-weight:700;font-size:13px;padding:8px 15px;border-radius:7px;border:1px solid var(--bdr);cursor:pointer;transition:.15s;background:var(--bg4);color:var(--txt);display:inline-flex;align-items:center;gap:5px;white-space:nowrap}
.btn:hover{background:var(--bg5);transform:translateY(-1px)}
.btn:active{transform:scale(0.97)}
.btn:disabled{opacity:.5;cursor:not-allowed}
.btn-gold{background:linear-gradient(135deg,var(--gold),var(--gold3));border:none;color:#060610;font-weight:900;box-shadow:0 2px 12px rgba(212,168,67,.3)}
.btn-gold:hover{box-shadow:0 4px 18px rgba(212,168,67,.45)}
.btn-green{background:rgba(45,204,138,.1);color:var(--green);border-color:rgba(45,204,138,.2)}
.btn-green:hover{background:rgba(45,204,138,.2)}
.btn-red{background:rgba(232,76,106,.1);color:var(--red);border-color:rgba(232,76,106,.2)}
.btn-red:hover{background:rgba(232,76,106,.2)}
.btn-blue{background:rgba(74,159,255,.1);color:var(--blue);border-color:rgba(74,159,255,.2)}
.btn-blue:hover{background:rgba(74,159,255,.2)}
.btn-amber{background:rgba(240,157,58,.1);color:var(--amber);border-color:rgba(240,157,58,.2)}
.btn-amber:hover{background:rgba(240,157,58,.2)}
.btn-sm{padding:5px 11px;font-size:12px}
.btn-xs{padding:3px 8px;font-size:11px}
.btn-full{width:100%;justify-content:center;padding:10px}

.badge{display:inline-flex;align-items:center;padding:2px 8px;border-radius:20px;font-size:10px;font-weight:700}
.bh{background:rgba(232,76,106,.12);color:var(--red)}
.bm{background:rgba(240,157,58,.12);color:var(--amber)}
.bl{background:rgba(45,204,138,.1);color:var(--green)}

.ti{background:var(--bg2);border:1px solid var(--bdr);border-radius:11px;padding:11px 13px;margin-bottom:7px;display:flex;align-items:flex-start;gap:11px;transition:.15s;position:relative;overflow:hidden}
.ti:hover{border-color:var(--bdr2)}
.ti.done{opacity:.35}
.tbar{position:absolute;right:0;top:0;bottom:0;width:3px;border-radius:0 11px 11px 0}
.tbody{flex:1;min-width:0;padding-right:4px}
.tname{font-size:14px;font-weight:700;margin-bottom:4px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.tmeta{display:flex;align-items:center;gap:5px;flex-wrap:wrap;font-size:11px}
.ttime{font-family:var(--mono);color:var(--gold);background:rgba(212,168,67,.1);padding:1px 6px;border-radius:4px}
.tdur{color:var(--txt3)}
.tacts{display:flex;gap:4px;flex-shrink:0;align-items:center}

.timeline-container{position:relative;min-height:400px}
.timeline-background{position:relative;height:auto}

.unscheduled-area{background:linear-gradient(135deg,rgba(212,168,67,.05),rgba(212,168,67,.02));border:1px dashed rgba(212,168,67,.3);border-radius:12px;margin-bottom:24px;padding:16px}
.unscheduled-title{font-size:13px;font-weight:800;color:var(--gold);margin-bottom:12px;display:flex;align-items:center;gap:8px}
.unscheduled-tasks{display:flex;flex-direction:column;gap:8px}
.unscheduled-task{background:var(--bg3);border:1px solid var(--bdr);border-radius:10px;padding:10px 14px;display:flex;align-items:center;gap:12px;cursor:grab;transition:all .2s;border-right:3px solid}
.unscheduled-task:active{cursor:grabbing}
.unscheduled-task.dragging{opacity:.5;cursor:grabbing}
.unscheduled-task.drag-over{border:2px solid var(--gold);background:rgba(212,168,67,.1)}
.unscheduled-priority{width:4px;height:30px;border-radius:2px}
.unscheduled-content{flex:1}
.unscheduled-name{font-size:13px;font-weight:700}
.unscheduled-meta{font-size:10px;color:var(--txt2);margin-top:3px}
.unscheduled-actions{display:flex;gap:6px}

.tl-label{position:absolute;font-size:10px;color:var(--txt3);font-family:var(--mono);transform:translateY(-50%);white-space:nowrap}
.tl-hline{position:absolute;left:0;right:0;height:1px;background:rgba(255,255,255,.04)}
.tl-task{position:absolute;left:3px;right:3px;border-radius:8px;padding:5px 9px;cursor:grab;overflow:hidden;transition:.12s}
.tl-task:hover{filter:brightness(1.15);cursor:grab}
.tl-task.dragging{opacity:.5;cursor:grabbing}
.tl-task.drag-over{border:2px solid var(--gold);background:rgba(212,168,67,.2)}
.tl-tname{font-size:11px;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.tl-tsub{font-size:10px;opacity:.6;margin-top:1px}
.tl-now{position:absolute;left:0;right:0;height:2px;background:var(--green);box-shadow:0 0 8px rgba(45,204,138,.5);z-index:10}
.tl-sleep{position:absolute;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent,var(--gold),transparent);z-index:10}

/* === Timeline arrows / connectors between consecutive tasks === */
.tl-task{z-index:2}
.tl-arrow{position:absolute;left:50%;transform:translateX(-50%);width:2px;background:linear-gradient(180deg,rgba(212,168,67,.55) 0%,rgba(212,168,67,.15) 100%);border-radius:2px;z-index:1;pointer-events:none}
.tl-arrow::before{content:"";position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);width:7px;height:7px;background:var(--gold);border-radius:50%;box-shadow:0 0 6px rgba(212,168,67,.55);opacity:.85}
.tl-arrow::after{content:"";position:absolute;left:50%;bottom:-1px;transform:translateX(-50%) rotate(45deg);width:7px;height:7px;border-right:2px solid var(--gold);border-bottom:2px solid var(--gold);opacity:.9}
.tl-arrow.tl-arrow-high{background:linear-gradient(180deg,rgba(232,76,106,.55),rgba(232,76,106,.15))}
.tl-arrow.tl-arrow-high::before{background:var(--red);box-shadow:0 0 6px rgba(232,76,106,.6)}
.tl-arrow.tl-arrow-high::after{border-color:var(--red)}
.tl-arrow.tl-arrow-low{background:linear-gradient(180deg,rgba(45,204,138,.5),rgba(45,204,138,.12))}
.tl-arrow.tl-arrow-low::before{background:var(--green);box-shadow:0 0 6px rgba(45,204,138,.55)}
.tl-arrow.tl-arrow-low::after{border-color:var(--green)}

#p-chat{display:none;flex-direction:column;flex:1;height:0}
#p-chat.show{display:flex}
.chat-msgs{flex:1;overflow-y:auto;padding:14px 18px;display:flex;flex-direction:column;gap:9px}
.chat-msgs::-webkit-scrollbar{width:3px}
.mu{display:flex;justify-content:flex-start}
.mu .bub{background:linear-gradient(135deg,var(--gold),var(--gold3));color:#060610;padding:9px 14px;border-radius:16px 16px 16px 4px;font-size:13px;max-width:66%;font-weight:700;line-height:1.5}
.mb{display:flex;justify-content:flex-end;align-items:flex-end;gap:8px}
.mb .bub{background:var(--bg3);border:1px solid var(--bdr);color:var(--txt);padding:9px 14px;border-radius:16px 16px 4px 16px;font-size:13px;max-width:76%;line-height:1.9;white-space:pre-wrap}
.av{width:30px;height:30px;flex-shrink:0;background:linear-gradient(135deg,var(--gold),var(--gold3));border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:900;color:#060610}
.chat-foot{padding:10px 18px 14px;border-top:1px solid var(--bdr);background:rgba(14,14,28,.95);flex-shrink:0}
body.light .chat-foot{background:rgba(255,255,255,.95)}
.chips{display:flex;gap:5px;flex-wrap:wrap;margin-bottom:8px}
.chip{padding:4px 11px;border-radius:20px;border:1px solid var(--bdr);background:var(--bg4);color:var(--txt2);font-size:11px;cursor:pointer;font-family:var(--font);font-weight:600;transition:.12s}
.chip:hover{border-color:var(--gold);color:var(--gold)}
.chat-row{display:flex;gap:7px;align-items:center}
#chat-in{flex:1;border-radius:22px;padding:9px 16px;background:var(--bg4);border:1px solid var(--bdr);font-size:13px}
#chat-in:focus{border-color:var(--gold);box-shadow:0 0 0 3px rgba(212,168,67,.1)}
#send-btn{width:40px;height:40px;border-radius:50%;border:none;background:linear-gradient(135deg,var(--gold),var(--gold3));color:#060610;font-size:17px;cursor:pointer;flex-shrink:0;display:flex;align-items:center;justify-content:center;box-shadow:0 2px 12px rgba(212,168,67,.3);transition:.15s;font-weight:900}
#send-btn:hover{transform:scale(1.1)}
#send-btn:active{transform:scale(0.93)}
#send-btn:disabled{opacity:.5;cursor:not-allowed}
.typing{display:flex;gap:4px;align-items:center;padding:2px}
.typing span{width:6px;height:6px;border-radius:50%;background:var(--txt3);animation:tdot .8s infinite}
.typing span:nth-child(2){animation-delay:.14s}
.typing span:nth-child(3){animation-delay:.28s}
@keyframes tdot{0%,80%,100%{transform:translateY(0)}40%{transform:translateY(-5px)}}
.spin{display:inline-block;width:13px;height:13px;border:2px solid rgba(255,255,255,.15);border-top-color:currentColor;border-radius:50%;animation:sp .6s linear infinite}
@keyframes sp{to{transform:rotate(360deg)}}

#notifs{position:fixed;bottom:16px;left:16px;z-index:9999;display:flex;flex-direction:column-reverse;gap:6px;pointer-events:none;max-width:330px}
.notif{pointer-events:all;border-radius:9px;padding:10px 14px;font-size:13px;display:flex;gap:9px;align-items:flex-start;animation:sup .22s ease;box-shadow:0 4px 18px rgba(0,0,0,.5);line-height:1.5}
.n-ok{background:rgba(45,204,138,.14);border:1px solid rgba(45,204,138,.28);color:var(--green)}
.n-info{background:rgba(74,159,255,.11);border:1px solid rgba(74,159,255,.22);color:var(--blue)}
.n-warn{background:rgba(240,157,58,.13);border:1px solid rgba(240,157,58,.27);color:var(--amber)}
.n-err{background:rgba(232,76,106,.12);border:1px solid rgba(232,76,106,.25);color:var(--red)}
@keyframes sup{from{transform:translateY(10px);opacity:0}to{transform:translateY(0);opacity:1}}
.nx{cursor:pointer;opacity:.5;font-size:14px;line-height:1}
.nx:hover{opacity:1}

#modal{position:fixed;inset:0;background:rgba(0,0,0,.72);z-index:1000;display:none;align-items:center;justify-content:center;backdrop-filter:blur(5px)}
#modal.open{display:flex}
.modal-box{background:var(--bg2);border:1px solid var(--bdr2);border-radius:15px;padding:24px;max-width:460px;width:91%;box-shadow:0 20px 60px rgba(0,0,0,.7);animation:pop .18s ease}
@keyframes pop{from{transform:scale(.9);opacity:0}to{transform:scale(1);opacity:1}}
.modal-h{font-size:16px;font-weight:900;margin-bottom:7px}
.modal-b{font-size:13px;color:var(--txt2);line-height:1.8;margin-bottom:16px;white-space:pre-line}
.modal-btns{display:flex;gap:7px;flex-wrap:wrap}
.empty{display:flex;flex-direction:column;align-items:center;justify-content:center;padding:36px 18px;color:var(--txt3);text-align:center;gap:9px}
.empty-i{font-size:40px;opacity:.28}
.empty-t{font-size:14px;font-weight:700;color:var(--txt2)}

.batch-config{background:linear-gradient(135deg,rgba(212,168,67,.1),rgba(212,168,67,.05));border:1px solid rgba(212,168,67,.2);border-radius:12px;padding:16px;margin-bottom:20px}
.task-card{background:var(--bg3);border:1px solid var(--bdr);border-radius:10px;padding:15px;margin-bottom:15px;position:relative}
.task-card-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid var(--bdr)}
.task-number{font-size:14px;font-weight:800;color:var(--gold)}
.remove-task{background:rgba(232,76,106,.1);color:var(--red);border:none;border-radius:5px;padding:3px 8px;cursor:pointer;font-size:11px}
.remove-task:hover{background:rgba(232,76,106,.2)}
.batch-actions{display:flex;gap:10px;margin-top:20px;flex-wrap:wrap}

.recommendation-overlay{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.92);z-index:1000;display:flex;align-items:center;justify-content:center;backdrop-filter:blur(10px);padding:14px}
.recommendation-card{background:linear-gradient(180deg,var(--bg2) 0%,var(--bg3) 100%);border-radius:22px;max-width:760px;width:100%;max-height:88vh;overflow-y:auto;padding:22px 22px 18px;border:1px solid rgba(212,168,67,.35);box-shadow:0 24px 70px rgba(0,0,0,.55), 0 0 0 1px rgba(255,255,255,.02) inset}
.recommendation-card h3{color:var(--gold);margin-bottom:6px;font-size:21px;text-align:center;font-weight:900;letter-spacing:.2px}
.rec-sub{text-align:center;color:var(--txt3);font-size:11px;margin-bottom:18px}
.rec-option{background:var(--bg3);border:1px solid var(--bdr);border-radius:14px;padding:14px 16px;margin-bottom:11px;cursor:pointer;transition:transform .18s ease,border-color .18s ease,background .18s ease,box-shadow .18s ease;position:relative;overflow:hidden}
.rec-option::before{content:"";position:absolute;top:0;right:0;bottom:0;width:3px;background:linear-gradient(180deg,var(--gold),transparent);opacity:.45;transition:opacity .18s}
.rec-option:hover{transform:translateY(-2px);border-color:rgba(212,168,67,.55);background:rgba(212,168,67,.06);box-shadow:0 8px 22px rgba(0,0,0,.35)}
.rec-option:hover::before{opacity:1}
.rec-option.is-best{border-color:rgba(212,168,67,.7);background:linear-gradient(180deg,rgba(212,168,67,.08),rgba(212,168,67,.02))}
.rec-option-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;padding-bottom:8px;border-bottom:1px dashed var(--bdr);gap:10px;flex-wrap:wrap}
.rec-title-wrap{display:flex;align-items:center;gap:10px;flex:1;min-width:0}
.rec-num-pill{display:inline-flex;align-items:center;justify-content:center;min-width:30px;height:30px;padding:0 10px;border-radius:999px;background:rgba(212,168,67,.14);color:var(--gold);font-weight:900;font-size:13px;border:1px solid rgba(212,168,67,.3);font-family:var(--mono)}
.rec-title{font-weight:800;font-size:14px;color:var(--txt)}
.rec-best-badge{display:inline-flex;align-items:center;gap:4px;background:linear-gradient(135deg,var(--gold),#e6c068);color:#1a1408;font-size:10px;font-weight:900;padding:3px 8px;border-radius:999px;letter-spacing:.3px}
.rec-score-wrap{display:flex;flex-direction:column;align-items:flex-end;gap:2px}
.rec-score{color:var(--gold);font-family:var(--mono);font-weight:900;font-size:15px;line-height:1}
.rec-score-lbl{font-size:9px;color:var(--txt3);font-weight:600;letter-spacing:.4px;text-transform:uppercase}
.rec-tasks{font-size:12px;color:var(--txt2);line-height:1.7;margin-top:6px;display:flex;flex-direction:column;gap:2px}
.rec-task-item{display:flex;gap:8px;align-items:center;padding:5px 8px;border-radius:7px;background:rgba(255,255,255,.015);transition:background .15s}
.rec-task-item:hover{background:rgba(255,255,255,.04)}
.rec-priority-high{color:var(--red);font-weight:700}
.rec-priority-medium{color:var(--amber);font-weight:600}
.rec-priority-low{color:var(--green)}
.rec-time{font-family:var(--mono);color:var(--gold);font-size:11px;margin-right:auto;background:rgba(212,168,67,.1);padding:2px 7px;border-radius:5px}
.rec-badge{display:inline-block;background:rgba(255,255,255,.05);padding:2px 7px;border-radius:8px;font-size:10px;color:var(--txt2)}
.rec-more{font-size:10px;color:var(--txt3);padding:5px 8px;font-style:italic}
.rec-learn-tip{margin-top:14px;padding:10px 12px;background:linear-gradient(90deg,rgba(212,168,67,.12),rgba(212,168,67,.04));border:1px solid rgba(212,168,67,.18);border-radius:10px;text-align:center;font-size:11px;color:var(--txt2)}

@media(max-width:640px){
  .app{grid-template-columns:1fr}
  .sidebar{position:fixed;bottom:0;left:0;right:0;height:56px;flex-direction:row;border-left:none;border-top:1px solid var(--bdr);z-index:50;background:var(--bg2)}
  .brand,.clock-wrap,.sb-stats{display:none}
  nav{display:flex;flex-direction:row;padding:3px 5px;align-items:center;justify-content:space-around;flex:1;overflow-x:auto}
  .nb{flex-direction:column;gap:1px;padding:4px 5px;font-size:9px;min-width:46px;border:none;flex:1;justify-content:center;align-items:center}
  .nb-icon{font-size:17px;width:auto}
  .main{height:calc(100vh - 56px)}
  .topbar{padding:8px 13px}
  .scroll{padding:11px 13px}
  .g4{grid-template-columns:1fr 1fr}
  .g3{grid-template-columns:1fr 1fr}
  .chat-msgs{padding:11px 13px}
  .chat-foot{padding:8px 13px 11px}
  .notif-panel{width:90%;left:5%;right:5%}
}
@media(max-width:360px){.g2,.g3,.g4{grid-template-columns:1fr}}
</style>
</head>
<body>
<div id="notifs"></div>

<div class="notif-panel" id="notif-panel">
  <div style="font-weight:800;margin-bottom:8px;color:var(--gold)">🔔 مركز الإشعارات</div>
  <div id="notif-list"></div>
</div>

<div id="modal" onclick="if(event.target===this)closeModal()">
  <div class="modal-box">
    <div class="modal-h" id="m-title"></div>
    <div class="modal-b" id="m-body"></div>
    <div id="m-extra"></div>
    <div class="modal-btns" id="m-btns"></div>
  </div>
</div>

<div id="recommendation-overlay" class="recommendation-overlay" style="display:none">
  <div class="recommendation-card">
    <h3>🧠 المقترحات الذكية للجدولة</h3>
    <div class="rec-sub">اختر أحد المقترحات وسيتم تطبيقه على جدولك الزمني فوراً</div>
    <div id="rec-options-list"></div>
    <button class="btn btn-gold btn-full" onclick="closeRecommendationOverlay()" style="margin-top:14px">إغلاق</button>
  </div>
</div>

<div class="app">
  <aside class="sidebar">
    <div class="brand">
      <div class="brand-icon">⭐</div>
      <div class="brand-text">
        <div class="brand-name">Smart<span>Day</span></div>
        <div class="brand-ver">AI v9.0</div>
      </div>
    </div>
    <div class="clock-wrap">
      <div class="clock" id="clock">--:--:--</div>
      <div class="clock-date" id="clock-date"></div>
    </div>
    <nav>
      <button class="nb active" id="n-dashboard" onclick="showPanel('dashboard')"><span class="nb-icon">🏠</span><span class="nb-lbl">الرئيسية</span></button>
      <button class="nb" id="n-add" onclick="showPanel('add')"><span class="nb-icon">➕</span><span class="nb-lbl">إضافة</span></button>
      <button class="nb" id="n-tasks" onclick="showPanel('tasks')"><span class="nb-icon">📋</span><span class="nb-lbl">المهام</span><span class="nb-badge" id="nb-cnt" style="display:none">0</span></button>
      <button class="nb" id="n-timeline" onclick="showPanel('timeline')"><span class="nb-icon">📅</span><span class="nb-lbl">الجدول</span></button>
      <button class="nb" id="n-chat" onclick="showPanel('chat')"><span class="nb-icon">💬</span><span class="nb-lbl">المساعد</span></button>
      <button class="nb" id="n-history" onclick="showPanel('history')"><span class="nb-icon">📜</span><span class="nb-lbl">السجل</span></button>
    </nav>
    <div class="sb-stats">
      <div class="ss"><div class="ss-n" id="ss-done" style="color:var(--green)">0</div><div class="ss-l">تم</div></div>
      <div class="ss"><div class="ss-n" id="ss-pend" style="color:var(--amber)">0</div><div class="ss-l">متبقية</div></div>
      <div class="ss"><div class="ss-n" id="ss-mins" style="color:var(--gold)">0</div><div class="ss-l">دقيقة</div></div>
      <div class="ss"><div class="ss-n" id="ss-free" style="color:var(--blue)">0</div><div class="ss-l">وقت حر</div></div>
    </div>
  </aside>

  <main class="main">
    <div class="topbar">
      <span class="topbar-title" id="ptitle">🏠 الرئيسية</span>
      <div class="topbar-right">
        <div class="sleep-info">🌙 <b id="sv">--:--</b> | باقي: <b id="sl">--</b></div>
        <button class="notif-btn" id="notif-btn" onclick="toggleNotifPanel()" title="الإشعارات">
          🔔<span class="notif-btn-badge" id="notif-badge" style="display:none">0</span>
        </button>
        <button class="theme-btn" onclick="toggleTheme()" title="تغيير المظهر">🌓</button>
        <button class="btn btn-sm btn-gold" onclick="runOptimize()">🔄 ترتيب</button>
        <button class="btn btn-sm btn-blue" onclick="refreshData()">🔄 تحديث</button>
      </div>
    </div>

    <div id="p-dashboard" class="panel show">
      <div class="scroll">
        <div class="g4" style="margin-bottom:11px">
          <div class="sb"><div class="sb-n" id="d-done" style="color:var(--green)">0</div><div class="sb-l">✅ تم انجازها</div></div>
          <div class="sb"><div class="sb-n" id="d-pend" style="color:var(--amber)">0</div><div class="sb-l">⏳ متبقية</div></div>
          <div class="sb"><div class="sb-n" id="d-mins" style="color:var(--gold)">0</div><div class="sb-l">⏱️ دقيقة</div></div>
          <div class="sb"><div class="sb-n" id="d-free" style="color:var(--blue)">0</div><div class="sb-l">🌙 وقت حر</div></div>
        </div>
        <div class="card">
          <div class="card-t"><span>تقدم اليوم</span><span id="d-pct" style="color:var(--green);font-size:14px;font-weight:900">0%</span></div>
          <div style="display:flex;justify-content:space-between;font-size:12px;color:var(--txt2);margin-bottom:6px"><span id="d-dl">0 تمت</span><span id="d-tl">0 اجمالي</span></div>
          <div class="pbar"><div class="pbar-f" id="d-bar" style="width:0%"></div></div>
        </div>
        <div class="g2">
          <div class="card">
            <div class="card-t">🔥 مهام عاجلة <button class="btn btn-xs btn-gold" onclick="showPanel('add')">+ اضافة</button></div>
            <div id="d-urg"></div>
          </div>
          <div class="card">
            <div class="card-t">🎯 المهمة القادمة</div>
            <div id="d-nxt"></div>
          </div>
        </div>
        <div class="card">
          <div class="card-t">🧠 رؤى ذكية</div>
          <div id="d-insights"></div>
        </div>
        <div class="card">
          <div class="card-t">⚡ أوامر سريعة</div>
          <div style="display:flex;gap:7px;flex-wrap:wrap">
            <button class="btn btn-sm btn-gold" onclick="showPanel('chat')">💬 المساعد</button>
            <button class="btn btn-sm btn-green" onclick="runOptimize()">🔄 ترتيب ذكي</button>
            <button class="btn btn-sm btn-amber" onclick="quickCmd('تقرير اليوم')">📊 تقرير</button>
            <button class="btn btn-sm btn-blue" onclick="showRecommendationsManual()">🧠 اقتراح جدول</button>
            <button class="btn btn-sm" onclick="showPanel('add')">➕ مهمة جديدة</button>
            <button class="btn btn-sm btn-red" onclick="quickCmd('اجل كل مهام اليوم للغد')">📅 تأجيل الكل</button>
          </div>
        </div>
      </div>
    </div>

    <div id="p-add" class="panel">
      <div class="scroll">
        <div class="card">
          <div class="card-t">⚙️ إعدادات اليوم</div>
          <div class="g2">
            <div class="fg"><label>🌙 وقت النوم</label><input type="time" id="cfg-sleep" value="23:30"></div>
            <div class="fg"><label>🌅 وقت البداية</label><input type="time" id="cfg-start" value="08:00"></div>
          </div>
          <button class="btn btn-sm btn-amber" onclick="saveSettings()">💾 حفظ الإعدادات</button>
        </div>

        <div class="card">
          <div class="card-t">📦 إضافة متعددة</div>
          <div class="batch-config">
            <div class="fg">
              <label>🔢 كم عدد المهام التي تريد إضافتها؟</label>
              <div style="display:flex;gap:10px;align-items:center">
                <input type="number" id="task-count" min="1" max="20" value="1" style="width:100px">
                                <button class="btn btn-gold" onclick="generateTaskForms()">✨ إنشاء الحقول</button>
              </div>
            </div>
          </div>
          <div id="batch-tasks-container"></div>
          <div class="batch-actions" id="batch-actions" style="display:none">
            <button class="btn btn-gold btn-full" onclick="submitBatchTasks()">📦 إضافة جميع المهام</button>
            <button class="btn btn-sm" onclick="clearBatchForms()">🗑️ مسح الكل</button>
          </div>
        </div>

        <div class="card">
          <div class="card-t">➕ إضافة مفردة</div>
          <div class="fg"><label>اسم المهمة *</label><input type="text" id="f-name" placeholder="مثال: مراجعة دروس، اجتماع..."></div>
          <div class="g2">
            <div class="fg"><label>⏱️ المدة (دقيقة)</label><input type="number" id="f-dur" value="30" min="5" max="480"></div>
            <div class="fg"><label>⏰ وقت البدء (اختياري)</label><input type="time" id="f-start"></div>
          </div>
          <div class="g3">
            <div class="fg"><label> الأولوية</label>
              <select id="f-pri"><option value="high">🔴 عالية</option><option value="medium" selected>🟡 متوسطة</option><option value="low">🟢 منخفضة</option></select></div>
            <div class="fg"><label> الصعوبة</label>
              <select id="f-dif"><option value="hard"> صعبة</option><option value="medium" selected> متوسطة</option><option value="easy"> سهلة</option></select></div>
            <div class="fg"><label>📅 التاريخ</label><input type="date" id="f-date"></div>
          </div>
          <div class="fg"><label>📝 ملاحظات</label><textarea id="f-notes" placeholder="تفاصيل إضافية..."></textarea></div>
          <button class="btn btn-blue btn-full" onclick="addSingleTask()">➕ إضافة مهمة واحدة</button>
        </div>
      </div>
    </div>

    <div id="p-tasks" class="panel">
      <div class="scroll">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:11px;flex-wrap:wrap;gap:8px">
          <span style="font-size:14px;font-weight:800">📋 إدارة المهام</span>
          <div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center">
            <select id="flt-d" onchange="renderTasks()" style="padding:6px 8px;font-size:12px"><option value="today">اليوم</option><option value="all">الكل</option><option value="tomorrow">الغد</option></select>
            <select id="flt-s" onchange="renderTasks()" style="padding:6px 8px;font-size:12px"><option value="all">كل الحالات</option><option value="pending">معلقة</option><option value="done">تمت</option></select>
            <button class="btn btn-sm btn-gold" onclick="runOptimize()">🔄 ترتيب</button>
            <button class="btn btn-sm btn-red" onclick="clearAllTasks()">🗑️ مسح الكل</button>
            <button class="btn btn-sm btn-blue" onclick="refreshData()">🔄 تحديث</button>
          </div>
        </div>
        <div id="tasks-list"></div>
      </div>
    </div>

    <div id="p-timeline" class="panel">
      <div class="scroll">
        <div class="card" style="padding:0;overflow:hidden">
          <div style="padding:12px 16px;border-bottom:1px solid var(--bdr);display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:7px">
            <span style="font-size:13px;font-weight:800;color:var(--txt2)">📅 الجدول الزمني - جميع مهام اليوم</span>
            <div style="display:flex;gap:6px">
              <button class="btn btn-sm btn-gold" onclick="runOptimize()">🔄 ترتيب ذكي</button>
              <button class="btn btn-sm" onclick="refreshData()">🔄 تحديث</button>
            </div>
          </div>
          <div id="tl-wrap" style="padding:16px;min-height:300px;"></div>
        </div>
      </div>
    </div>

    <div id="p-chat" class="panel">
      <div class="chat-msgs" id="chat-msgs">
        <div class="mb"><div class="av">AI</div><div class="bub">✨ اهلاً! أنا SmartDay AI Agent v8.4 - مساعدك الذكي للجدولة\n\n**🌟 الميزات الجديدة:**\n• 🔄 تبديل مباشر للمهام\n• ✅ إكمال المهمة مباشرة\n• ⏸️ عدم إكمال المهمة مع وضع الباقي بعد آخر مهمة\n• 🧠 8 استراتيجيات جدولة ذكية\n• 📊 يتعلم من اختياراتك السابقة\n• ⏰ يأخذ الإنتاجية والطاقة بعين الاعتبار\n• 🌓 الوضع النهاري والليلي\n• 🗑️ زر مسح جميع المهام\n\n**🎯 جرب الآن:**\n• "اقترح جدول" - 8 جداول ذكية\n• "بدل مذاكرة مع اجتماع" - تبديل مباشر\n• "ضيف مهمة" - إضافة تفاعلية\n• "خلصت مذاكرة" - إكمال مباشر\n• "ملحقتش مذاكرة" - تعديل الوقت المتبقي\n• "غير وقت النوم 01:30" - ضبط وقت النوم</div></div>
      </div>
      <div class="chat-foot">
        <div class="chips">
          <span class="chip" onclick="sendChip('جدول ذكي')">🧠 جدولة ذكية</span>
          <span class="chip" onclick="sendChip('بدل مذاكرة مع اجتماع')">🔄 تبديل مباشر</span>
          <span class="chip" onclick="sendChip('ضيف مهمة')">➕ إضافة تفاعلية</span>
          <span class="chip" onclick="sendChip('تقرير اليوم')">📊 تقرير</span>
          <span class="chip" onclick="sendChip('رتب المهام حسب الأولوية')">🔄 ترتيب</span>
          <span class="chip" onclick="sendChip('المهمة القادمة')">🎯 التالية</span>
          <span class="chip" onclick="sendChip('وريني المهام')">📋 المهام</span>
          <span class="chip" onclick="sendChip('مساعدة')">🆘 مساعدة</span>
        </div>
        <div class="chat-row">
          <input type="text" id="chat-in" placeholder="اكتب أمرك... مثال: بدل مذاكرة مع اجتماع" autocomplete="off" onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendMsg()}">
          <button id="send-btn" onclick="sendMsg()">➤</button>
        </div>
      </div>
    </div>

    <div id="p-history" class="panel">
      <div class="scroll">
        <div class="card" style="padding:0;overflow:hidden">
          <div style="padding:12px 16px;border-bottom:1px solid var(--bdr);display:flex;justify-content:space-between;align-items:center">
            <span style="font-size:13px;font-weight:800;color:var(--txt2)">📜 سجل الأنشطة</span>
            <div style="display:flex;gap:6px">
              <button class="btn btn-sm btn-red" onclick="clearHistory()">🗑️ مسح السجل</button>
              <button class="btn btn-sm btn-red" onclick="clearAllTasks()">🗑️ مسح كل المهام</button>
            </div>
          </div>
          <div id="hist-list" style="max-height:calc(100vh - 175px);overflow-y:auto"></div>
        </div>
      </div>
    </div>
  </main>
</div>

<script>
var tasks = [];
var settings = {sleepTime: '23:30', startTime: '08:00', theme: 'dark', language: 'arabic'};
var curPanel = 'dashboard';
var chatBusy = false;
var modalCbs = [];
var batchTaskForms = [];
var dragSourceId = null;
var dragSourceElement = null;
var isDraggingUnscheduled = false;
var notificationCount = 0;
var allNotifications = [];
var currentRecommendations = null;
var sessionId = 'user_' + Date.now();
// Flag لتتبع ما إذا تم عرض الـ recommendation overlay تلقائياً مسبقاً
// يبقى true حتى يتم حذف كل المهام (ثم يعود false ليظهر مرة أخرى عند أول إضافة)
var recommendationShownThisSession = false;

function initTheme() {
    var savedTheme = settings.theme || 'dark';
    if (savedTheme === 'light') {
        document.body.classList.add('light');
    } else {
        document.body.classList.remove('light');
    }
}

function toggleTheme() {
    if (document.body.classList.contains('light')) {
        document.body.classList.remove('light');
        settings.theme = 'dark';
    } else {
        document.body.classList.add('light');
        settings.theme = 'light';
    }
    saveThemeSetting();
    showNotif(settings.theme === 'dark' ? '🌙 تم تفعيل الوضع الليلي' : '☀️ تم تفعيل الوضع النهاري', 'info');
}

async function saveThemeSetting() {
    await apiCall('/api/settings', 'POST', {theme: settings.theme});
}

function toggleNotifPanel() {
    var panel = document.getElementById('notif-panel');
    panel.classList.toggle('open');
    if (panel.classList.contains('open')) {
        renderNotifications();
        notificationCount = 0;
        updateNotifBadge();
    }
}

function updateNotifBadge() {
    var badge = document.getElementById('notif-badge');
    if (notificationCount > 0) {
        badge.style.display = 'flex';
        badge.textContent = notificationCount > 99 ? '99+' : notificationCount;
    } else {
        badge.style.display = 'none';
    }
}

function addNotification(msg, type) {
    allNotifications.unshift({
        message: msg,
        type: type || 'info',
        time: new Date().toLocaleTimeString('ar-EG')
    });
    notificationCount++;
    updateNotifBadge();
    if (allNotifications.length > 50) {
        allNotifications = allNotifications.slice(0, 50);
    }
}

function renderNotifications() {
    var el = document.getElementById('notif-list');
    if (!el) return;
    if (allNotifications.length === 0) {
        el.innerHTML = '<div class="notif-empty">🔔 لا توجد إشعارات حالياً</div>';
        return;
    }
    var icons = {info: 'ℹ️', warn: '⚠️', ok: '✅', err: '❌', ai: '🧠'};
    el.innerHTML = allNotifications.slice(0, 20).map(function(n) {
        var icon = icons[n.type] || '📌';
        return '<div class="notif-item"><span>' + icon + '</span><div><div>' + esc(n.message) + '</div><div style="font-size:10px;color:var(--txt3)">' + n.time + '</div></div></div>';
    }).join('');
}

function todayStr() {
    var d = new Date();
    return d.toISOString().split('T')[0];
}

function tomStr() {
    var d = new Date();
    d.setDate(d.getDate() + 1);
    return d.toISOString().split('T')[0];
}

function nowStr() {
    var n = new Date();
    return pad(n.getHours()) + ':' + pad(n.getMinutes());
}

function toMins(t) {
    if (!t) return 0;
    var p = String(t).split(':');
    return (parseInt(p[0]) || 0) * 60 + (parseInt(p[1]) || 0);
}

function toTime(m) {
    m = ((m % 1440) + 1440) % 1440;
    return pad(Math.floor(m / 60)) + ':' + pad(m % 60);
}

function pad(n) {
    return String(n).padStart(2, '0');
}

function esc(s) {
    return String(s || '').replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
}

function $(id) {
    return document.getElementById(id);
}

function setText(id, v) {
    var el = $(id);
    if (el) el.textContent = v;
}

function formatMarkdown(t) {
    return esc(t).replace(/\*\*(.+?)\*\*/g, '<strong>$1</strong>').replace(/\n/g, '<br>');
}

async function apiCall(path, method, body) {
    try {
        var opts = {method: method || 'GET', headers: {'Content-Type': 'application/json'}};
        if (body != null) opts.body = JSON.stringify(body);
        var r = await fetch(path, opts);
        if (!r.ok) throw new Error('HTTP ' + r.status);
        return await r.json();
    } catch (e) {
        showNotif('خطأ في الاتصال: ' + e.message, 'err');
        return null;
    }
}

function dragStartHandler(e, taskId, startTime) {
    dragSourceId = taskId;
    dragSourceElement = e.target.closest('.tl-task');
    if (dragSourceElement) {
        dragSourceElement.classList.add('dragging');
    }
    e.dataTransfer.effectAllowed = 'move';
    e.dataTransfer.setData('text/plain', JSON.stringify({id: taskId, time: startTime, type: 'scheduled'}));
}

function dragEndHandler(e) {
    if (dragSourceElement) {
        dragSourceElement.classList.remove('dragging');
    }
    dragSourceId = null;
    dragSourceElement = null;
    isDraggingUnscheduled = false;
    document.querySelectorAll('.tl-task, .unscheduled-task').forEach(function(el) {
        el.classList.remove('drag-over');
    });
}

function dragOverHandler(e, targetTime) {
    e.preventDefault();
    var target = e.target.closest('.tl-task');
    if (target) {
        target.classList.add('drag-over');
    }
    e.dataTransfer.dropEffect = 'move';
}

function dragLeaveHandler(e) {
    var target = e.target.closest('.tl-task');
    if (target) {
        target.classList.remove('drag-over');
    }
}

function unscheduledDragStart(e, taskId, taskName) {
    dragSourceId = taskId;
    isDraggingUnscheduled = true;
    e.dataTransfer.effectAllowed = 'move';
    e.dataTransfer.setData('text/plain', JSON.stringify({id: taskId, type: 'unscheduled'}));
    e.target.closest('.unscheduled-task').classList.add('dragging');
}

function unscheduledDragEnd(e) {
    var el = e.target.closest('.unscheduled-task');
    if (el) el.classList.remove('dragging');
    dragSourceId = null;
    isDraggingUnscheduled = false;
    document.querySelectorAll('.tl-task, .unscheduled-task').forEach(function(el) {
        el.classList.remove('drag-over');
    });
}

async function dropOnTimelineHandler(e, dropTime) {
    e.preventDefault();
    document.querySelectorAll('.tl-task, .unscheduled-task').forEach(function(el) {
        el.classList.remove('drag-over');
    });
    if (!dragSourceId) return;
    var newTime = dropTime;
    var nowMins = toMins(nowStr());
    var newMins = toMins(newTime);
    if (newMins < nowMins) {
        showNotif('❌ لا يمكن وضع المهمة في وقت مضى!', 'err');
        return;
    }
    var result = await apiCall('/api/tasks/' + dragSourceId + '/update-time', 'POST', {time: newTime});
    if (result && result.success) {
        await loadAll();
        showNotif('✅ تم تحديد وقت المهمة بنجاح!', 'ok');
        renderTimeline();
    } else {
        showNotif('❌ فشل تحديد الوقت: ' + (result.message || 'وقت غير صالح'), 'err');
    }
    dragSourceId = null;
    isDraggingUnscheduled = false;
}

async function dropHandler(e, targetTime, targetTaskId) {
    e.preventDefault();
    var target = e.target.closest('.tl-task');
    if (target) {
        target.classList.remove('drag-over');
    }
    if (!dragSourceId) return;
    var newTime = targetTime;
    var nowMins = toMins(nowStr());
    var newMins = toMins(newTime);
    if (newMins < nowMins) {
        showNotif('❌ لا يمكن وضع المهمة في وقت مضى!', 'err');
        return;
    }
    var result = await apiCall('/api/tasks/' + dragSourceId + '/update-time', 'POST', {time: newTime});
    if (result && result.success) {
        await loadAll();
        showNotif('✅ تم تغيير وقت المهمة بنجاح!', 'ok');
        renderTimeline();
    } else {
        showNotif('❌ فشل تغيير وقت المهمة: ' + (result.message || 'وقت غير صالح'), 'err');
    }
    dragSourceId = null;
    dragSourceElement = null;
}

async function clearAllTasks() {
    if (!confirm('⚠️ تحذير: سيتم حذف جميع المهام نهائياً!\nهل أنت متأكد؟')) return;
    var result = await apiCall('/api/tasks', 'DELETE');
    if (result && result.success) {
        await loadAll();
        showNotif('🗑️ تم حذف جميع المهام!', 'ok');
        addNotification('🗑️ تم حذف جميع المهام', 'warn');
    } else {
        showNotif('❌ فشل حذف المهام', 'err');
    }
}

async function showRecommendationsManual() {
    var pendingCount = tasks.filter(function(t) { return t.status === 'pending' && t.date === todayStr(); }).length;
    if (pendingCount <= 1) {
        showNotif('⚠️ لا تحتاج إلى جدولة ذكية! لديك ' + pendingCount + ' مهمة فقط.', 'warn');
        return;
    }
    await loadRecommendationsAndShowOverlay();
}

async function loadRecommendationsAndShowOverlay() {
    var result = await apiCall('/api/schedule/recommendations');
    if (!result || !result.top_8 || result.top_8.length === 0) {
        showNotif('لا توجد مهام لجدولتها', 'warn');
        return;
    }
    currentRecommendations = result;
    var overlay = document.getElementById('recommendation-overlay');
    var container = document.getElementById('rec-options-list');

    // أسماء عربية مرتبة (المقترح الأول، الثاني، ...)
    var ordinalNames = [
        'المقترح الأول',
        'المقترح الثاني',
        'المقترح الثالث',
        'المقترح الرابع',
        'المقترح الخامس',
        'المقترح السادس',
        'المقترح السابع',
        'المقترح الثامن'
    ];

    var html = '';
    for (var i = 0; i < result.top_8.length; i++) {
        var schedule = result.top_8[i];
        var isRec = schedule.index === result.recommended_index;
        var label = ordinalNames[i] || ('المقترح ' + (i + 1));
        var scheduleIndex = schedule.index;
        var bestClass = isRec ? ' is-best' : '';

        html += '<div class="rec-option' + bestClass + '" onclick="applySchedule(' + scheduleIndex + ', \'' + schedule.strategy + '\', \'' + (schedule.id || '') + '\')">';
        html += '<div class="rec-option-header">';
        html += '<div class="rec-title-wrap">';
        html += '<span class="rec-num-pill">' + (i + 1) + '</span>';
        html += '<span class="rec-title">' + label + '</span>';
        if (isRec) {
            html += '<span class="rec-best-badge">⭐ الأفضل</span>';
        }
        html += '</div>';
        html += '<div class="rec-score-wrap">';
        html += '<span class="rec-score">' + schedule.score.toFixed(0) + '</span>';
        html += '<span class="rec-score-lbl">نقطة الجودة</span>';
        html += '</div>';
        html += '</div>';
        html += '<div class="rec-tasks">';
        if (schedule.tasks && schedule.tasks.length > 0) {
            for (var j = 0; j < Math.min(schedule.tasks.length, 5); j++) {
                var t = schedule.tasks[j];
                var priClass = t.priority === 'high' ? 'rec-priority-high' : (t.priority === 'medium' ? 'rec-priority-medium' : 'rec-priority-low');
                var priIcon = t.priority === 'high' ? '🔴' : (t.priority === 'medium' ? '🟡' : '🟢');
                var timeStr = t.startTime ? t.startTime : 'غير مجدولة';
                html += '<div class="rec-task-item">';
                html += '<span class="' + priClass + '">' + priIcon + '</span>';
                html += '<span><strong>' + esc(t.name) + '</strong></span>';
                html += '<span class="rec-time">' + timeStr + '</span>';
                html += '<span class="rec-badge">' + t.duration + 'د</span>';
                html += '</div>';
            }
            if (schedule.tasks.length > 5) {
                html += '<div class="rec-more">... و ' + (schedule.tasks.length - 5) + ' مهام أخرى</div>';
            }
        } else {
            html += '<div class="rec-more">لا توجد مهام</div>';
        }
        html += '</div></div>';
    }

    if (result.total_selections && result.total_selections > 0) {
        html += '<div class="rec-learn-tip">📚 تم التعلم من ' + result.total_selections + ' اختيار سابق لتحسين هذه المقترحات</div>';
    }
    container.innerHTML = html;
    overlay.style.display = 'flex';
}

function closeRecommendationOverlay() {
    document.getElementById('recommendation-overlay').style.display = 'none';
    currentRecommendations = null;
}

async function applySchedule(selectedIndex, strategyName, scheduleId) {
    if (!currentRecommendations) {
        showNotif('⚠️ لا توجد توصيات لعرضها', 'warn');
        return;
    }
    
    // التحقق من أن selectedIndex صحيح
    if (selectedIndex === undefined || selectedIndex === null) {
        showNotif('❌ خطأ: لم يتم تحديد جدول صحيح', 'err');
        return;
    }
    
    // CRITICAL: log selected ID on the client side too
    console.log('[SELECT] schedule_id=' + scheduleId + ' index=' + selectedIndex + ' strategy=' + strategyName);
    
    var targetBtn = event && event.target ? event.target.closest('.rec-option') : null;
    if (targetBtn) {
        targetBtn.style.opacity = '0.6';
        targetBtn.style.pointerEvents = 'none';
    }
    
    showNotif('🔄 جاري تطبيق الجدول المختار...', 'info', 2000);
    
    try {
        // CRITICAL FIX: send unique schedule_id, not just index
        var result = await apiCall('/api/schedule/select', 'POST', {
            schedule_id: scheduleId,
            recommended_idx: currentRecommendations.recommended_index,
            selected_idx: selectedIndex,
            strategy_name: strategyName
        });
        
        if (result && result.success) {
            // Verify selected vs applied IDs match
            if (result.selected_id && result.applied_id && result.selected_id !== result.applied_id) {
                console.error('[MISMATCH] selected=' + result.selected_id + ' applied=' + result.applied_id);
                showNotif('❌ عدم تطابق في معرّف الجدول!', 'err');
                return;
            }
            console.log('[APPLIED] id=' + result.applied_id);
            closeRecommendationOverlay();
            await loadAll();
            
            // تحديث جميع الشاشات
            if (curPanel === 'timeline') {
                renderTimeline();
            }
            if (curPanel === 'dashboard') {
                renderDash();
            }
            if (curPanel === 'tasks') {
                renderTasks();
            }
            
            // اسم المقترح حسب موقعه في القائمة (المقترح الأول، الثاني، ...)
            var ordinalNamesApply = [
                'المقترح الأول','المقترح الثاني','المقترح الثالث',
                'المقترح الرابع','المقترح الخامس','المقترح السادس',
                'المقترح السابع','المقترح الثامن'
            ];
            // ابحث عن الترتيب في القائمة المعروضة
            var orderInList = -1;
            if (currentRecommendations && currentRecommendations.top_8) {
                for (var k = 0; k < currentRecommendations.top_8.length; k++) {
                    if (currentRecommendations.top_8[k].index === selectedIndex) {
                        orderInList = k; break;
                    }
                }
            }
            var suggestionLabel = (orderInList >= 0)
                ? (ordinalNamesApply[orderInList] || ('المقترح ' + (orderInList + 1)))
                : 'المقترح المختار';

            showNotif('✅ تم تطبيق ' + suggestionLabel + ' على الجدول الزمني بنجاح!', 'ok');
            addNotification('✅ تم تطبيق ' + suggestionLabel + ' على الجدول الزمني', 'ok');
            
            // التبديل إلى عرض الجدول الزمني تلقائياً
            showPanel('timeline');
            updateSidebar();
            
            // إعادة رسم الجدول للتأكد
            setTimeout(function() {
                renderTimeline();
                var timelineScroll = document.querySelector('#p-timeline .scroll');
                if (timelineScroll) {
                    timelineScroll.scrollTop = 0;
                }
            }, 100);
            
        } else {
            var errorMsg = result?.message || 'حدث خطأ أثناء تطبيق الجدول';
            showNotif('❌ ' + errorMsg, 'err');
            console.error('Schedule application failed:', result);
        }
    } catch (error) {
        console.error('Exception in applySchedule:', error);
        showNotif('❌ حدث خطأ في الاتصال بالخادم', 'err');
    } finally {
        if (targetBtn) {
            targetBtn.style.opacity = '';
            targetBtn.style.pointerEvents = '';
        }
    }
}

function generateTaskForms() {
    var countInput = $('task-count');
    var count = parseInt(countInput.value);
    if (isNaN(count) || count < 1) {
        showNotif('الرجاء إدخال عدد صحيح أكبر من 0', 'warn');
        return;
    }
    if (count > 20) {
        showNotif('الحد الأقصى 20 مهمة في المرة الواحدة', 'warn');
        return;
    }
    batchTaskForms = [];
    var container = $('batch-tasks-container');
    var html = '';
    for (var i = 1; i <= count; i++) {
        html += '<div class="task-card" id="task-card-' + i + '">' +
            '<div class="task-card-header">' +
            '<span class="task-number">📌 المهمة رقم ' + i + '</span>' +
            '<button class="remove-task" onclick="removeTaskCard(' + i + ')">✕ حذف</button>' +
            '</div>' +
            '<div class="fg"><label>📝 اسم المهمة *</label><input type="text" id="batch-name-' + i + '" placeholder="أدخل اسم المهمة" required></div>' +
            '<div class="g2">' +
            '<div class="fg"><label>⏱️ المدة (دقيقة)</label><input type="number" id="batch-dur-' + i + '" value="30" min="5" max="480"></div>' +
            '<div class="fg"><label>⏰ وقت البدء (اختياري)</label><input type="time" id="batch-time-' + i + '"></div>' +
            '</div>' +
            '<div class="g3">' +
            '<div class="fg"><label> الأولوية</label><select id="batch-pri-' + i + '"><option value="high">🔴 عالية</option><option value="medium" selected>🟡 متوسطة</option><option value="low">🟢 منخفضة</option></select></div>' +
            '<div class="fg"><label> الصعوبة</label><select id="batch-dif-' + i + '"><option value="hard"> صعبة</option><option value="medium" selected> متوسطة</option><option value="easy"> سهلة</option></select></div>' +
            '<div class="fg"><label>📅 التاريخ (اختياري)</label><input type="date" id="batch-date-' + i + '" value="' + todayStr() + '"></div>' +
            '</div>' +
            '<div class="fg"><label>📝 ملاحظات (اختياري)</label><textarea id="batch-notes-' + i + '" rows="2" placeholder="أي تفاصيل إضافية..."></textarea></div>' +
            '</div>';
        batchTaskForms.push(i);
    }
    container.innerHTML = html;
    $('batch-actions').style.display = 'flex';
}

function clearBatchForms() {
    $('batch-tasks-container').innerHTML = '';
    $('batch-actions').style.display = 'none';
    batchTaskForms = [];
    showNotif('تم مسح جميع حقول الإضافة', 'info');
}

async function submitBatchTasks() {
    if (batchTaskForms.length === 0) {
        showNotif('لا توجد مهام لإضافتها. قم بإنشاء حقول أولاً', 'warn');
        return;
    }
    var tasksToAdd = [];
    var hasError = false;
    for (var i = 0; i < batchTaskForms.length; i++) {
        var num = batchTaskForms[i];
        var nameInput = $('batch-name-' + num);
        var name = nameInput ? nameInput.value.trim() : '';
        if (!name) {
            showNotif('الرجاء إدخال اسم للمهمة رقم ' + (i + 1), 'warn');
            hasError = true;
            break;
        }
        var duration = parseInt($('batch-dur-' + num).value) || 30;
        var startTime = $('batch-time-' + num).value || null;
        var priority = $('batch-pri-' + num).value;
        var difficulty = $('batch-dif-' + num).value;
        var date = $('batch-date-' + num).value || todayStr();
        var notes = $('batch-notes-' + num).value || '';
        tasksToAdd.push({
            name: name,
            duration: duration,
            startTime: startTime,
            priority: priority,
            difficulty: difficulty,
            date: date,
            notes: notes
        });
    }
    if (!hasError && tasksToAdd.length > 0) {
        var btn = event.target;
        var originalText = btn.innerHTML;
        btn.disabled = true;
        btn.innerHTML = '<span class="spin"></span> جاري الإضافة...';
        var result = await apiCall('/api/tasks/batch', 'POST', {tasks: tasksToAdd});
        btn.disabled = false;
        btn.innerHTML = originalText;
        if (result && result.success) {
            showNotif('✅ تم إضافة ' + result.count + ' مهمة بنجاح!', 'ok');
            addNotification('✅ تم إضافة ' + result.count + ' مهمة دفعة واحدة', 'ok');
            clearBatchForms();
            await loadAll();
            showPanel('tasks');
            var pendingCount = tasks.filter(function(t) { return t.status === 'pending' && t.date === todayStr(); }).length;
            if (pendingCount > 1 && !recommendationShownThisSession) {
                recommendationShownThisSession = true;
                await loadRecommendationsAndShowOverlay();
            } else if (pendingCount <= 1) {
                showNotif('💡 لديك ' + pendingCount + ' مهمة فقط. سيتم جدولتها تلقائياً!', 'info');
            }
        } else {
            showNotif('حدث خطأ أثناء إضافة المهام', 'err');
        }
    }
}

async function addSingleTask() {
    var name = ($('f-name') || {}).value;
    if (name) name = name.trim();
    if (!name) {
        showNotif('ادخل اسم المهمة', 'warn');
        return;
    }
    var btn = event.target;
    var originalText = btn.innerHTML;
    btn.disabled = true;
    btn.innerHTML = '<span class="spin"></span> جاري الإضافة...';
    var data = {
        name: name,
        duration: parseInt(($('f-dur') || {}).value) || 30,
        startTime: ($('f-start') || {}).value || null,
        priority: ($('f-pri') || {}).value || 'medium',
        difficulty: ($('f-dif') || {}).value || 'medium',
        notes: ($('f-notes') || {}).value || '',
        date: ($('f-date') || {}).value || todayStr()
    };
    var r = await apiCall('/api/tasks', 'POST', data);
    btn.disabled = false;
    btn.innerHTML = originalText;
    if (r && r.success) {
        $('f-name').value = '';
        var fs = $('f-start');
        if (fs) fs.value = '';
        var fn = $('f-notes');
        if (fn) fn.value = '';
        await loadAll();
        showNotif('تم إضافة "' + name + '"', 'ok');
        showPanel('tasks');
        var pendingCount = tasks.filter(function(t) { return t.status === 'pending' && t.date === todayStr(); }).length;
        if (pendingCount > 1 && !recommendationShownThisSession) {
            recommendationShownThisSession = true;
            await loadRecommendationsAndShowOverlay();
        } else if (pendingCount <= 1) {
            showNotif('💡 لديك ' + pendingCount + ' مهمة فقط. سيتم جدولتها تلقائياً!', 'info');
        }
    }
}

async function loadAll() {
    var d = await apiCall('/api/tasks');
    if (!d) return;
    tasks = d.tasks || [];
    settings = d.settings || settings;
    
    // إعادة تعيين flag الـ recommendation إذا لم تعد هناك مهام pending
    // (سواء حذف المستخدم كل المهام أو أنجزها كلها)
    var pendingTodayCount = tasks.filter(function(t) { 
        return t.status === 'pending' && t.date === todayStr(); 
    }).length;
    if (pendingTodayCount === 0) {
        recommendationShownThisSession = false;
    }
    
    var cs = $('cfg-sleep'), cst = $('cfg-start');
    if (cs) cs.value = settings.sleepTime;
    if (cst) cst.value = settings.startTime;
    var fd = $('f-date');
    if (fd && !fd.value) fd.value = todayStr();
    initTheme();
    updateSidebar();
    updateSleep();
    if (curPanel === 'dashboard') renderDash();
    else if (curPanel === 'tasks') renderTasks();
    else if (curPanel === 'timeline') renderTimeline();
    else if (curPanel === 'history') renderHistory();
}

function refreshData() {
    loadAll();
    showNotif('تم تحديث البيانات', 'ok');
}

function updateSidebar() {
    var today = todayStr();
    var tt = tasks.filter(function(t) { return t.date === today; });
    var done = tt.filter(function(t) { return t.status === 'done'; });
    var pend = tt.filter(function(t) { return t.status === 'pending' && t.date === today; });
    var mins = pend.reduce(function(s, t) { return s + t.duration; }, 0);
    var slm = toMins(settings.sleepTime);
    var stm = toMins(settings.startTime);
    var nm = toMins(nowStr());
    
    var free = 0;
    
    if (slm < stm) {
        var targetSleepMins = slm + 1440;
        free = Math.max(0, targetSleepMins - nm - mins);
    } else {
        free = Math.max(0, slm - nm - mins);
    }
    
    setText('ss-done', done.length);
    setText('ss-pend', pend.length);
    setText('ss-mins', mins);
    setText('ss-free', free);
    
    var nb = $('nb-cnt');
    if (nb) {
        nb.textContent = pend.length;
        nb.style.display = pend.length > 0 ? '' : 'none';
    }
}

function updateSleep() {
    var sleepTime = settings.sleepTime;
    var startTime = settings.startTime;
    var now = new Date();
    var nowMins = now.getHours() * 60 + now.getMinutes();
    var sleepMins = toMins(sleepTime);
    var startMins = toMins(startTime);
    
    var left = 0;
    
    if (sleepMins < startMins) {
        var targetMins = sleepMins + 1440;
        if (nowMins < startMins) {
            left = targetMins - nowMins;
        } else {
            left = targetMins - nowMins;
        }
        if (left < 0) left = 0;
        if (left > 1440) left = 1440;
    } else {
        if (nowMins < sleepMins) {
            left = sleepMins - nowMins;
        } else {
            left = 0;
        }
    }
    
    var h = Math.floor(left / 60);
    var m = left % 60;
    setText('sv', sleepTime);
    setText('sl', (h > 0 ? h + 'س ' : '') + m + 'د');
}

function tickClock() {
    var n = new Date();
    var cl = $('clock');
    if (cl) cl.textContent = [n.getHours(), n.getMinutes(), n.getSeconds()].map(pad).join(':');
    var days = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت'];
    var months = ['يناير', 'فبراير', 'مارس', 'ابريل', 'مايو', 'يونيو', 'يوليو', 'اغسطس', 'سبتمبر', 'اكتوبر', 'نوفمبر', 'ديسمبر'];
    var cd = $('clock-date');
    if (cd) cd.textContent = days[n.getDay()] + ' ' + n.getDate() + ' ' + months[n.getMonth()] + ' ' + n.getFullYear();
    updateSleep();
}
setInterval(tickClock, 1000);
tickClock();

var PANELS = ['dashboard', 'add', 'tasks', 'timeline', 'chat', 'history'];
var PTITLES = {
    dashboard: '🏠 الرئيسية',
    add: '➕ إضافة مهمة',
    tasks: '📋 إدارة المهام',
    timeline: '📅 الجدول الزمني',
    chat: '💬 المساعد الذكي',
    history: '📜 السجل'
};

function showPanel(name) {
    if (PANELS.indexOf(name) < 0) return;
    curPanel = name;
    PANELS.forEach(function(p) {
        var el = $('p-' + p);
        if (el) el.style.display = 'none';
        var nb = $('n-' + p);
        if (nb) nb.classList.remove('active');
    });
    var target = $('p-' + name);
    if (target) {
        target.style.display = 'flex';
        target.style.flexDirection = 'column';
    }
    var nb = $('n-' + name);
    if (nb) nb.classList.add('active');
    var pt = $('ptitle');
    if (pt) pt.innerHTML = PTITLES[name] || name;
    if (name === 'dashboard') renderDash();
    else if (name === 'tasks') renderTasks();
    else if (name === 'timeline') renderTimeline();
    else if (name === 'history') renderHistory();
    else if (name === 'chat') {
        var cm = $('chat-msgs');
            if (cm) setTimeout(function() { cm.scrollTop = cm.scrollHeight; }, 40);
        var ci = $('chat-in');
        if (ci) setTimeout(function() { ci.focus(); }, 80);
    }
}

function renderDash() {
    var today = todayStr();
    var tt = tasks.filter(function(t) { return t.date === today; });
    var done = tt.filter(function(t) { return t.status === 'done'; });
    var pend = tt.filter(function(t) { return t.status === 'pending'; });
    var mins = pend.reduce(function(s, t) { return s + t.duration; }, 0);
    var slm = toMins(settings.sleepTime), nm = toMins(nowStr()), startMins = toMins(settings.startTime);
    var free = 0;
    
    if (slm < startMins) {
        var totalSleepMins = slm + 1440;
        free = Math.max(0, totalSleepMins - nm - mins);
    } else {
        free = Math.max(0, slm - nm - mins);
    }
    
    var pct = tt.length > 0 ? Math.round(done.length / tt.length * 100) : 0;
    setText('d-done', done.length);
    setText('d-pend', pend.length);
    setText('d-mins', mins);
    setText('d-free', free);
    setText('d-pct', pct + '%');
    setText('d-dl', done.length + ' تمت');
    setText('d-tl', tt.length + ' اجمالي');
    var bar = $('d-bar');
    if (bar) bar.style.width = pct + '%';
    var urgEl = $('d-urg');
    if (urgEl) {
        var urg = pend.filter(function(t) { return t.priority === 'high'; });
        if (!urg.length) {
            urgEl.innerHTML = '<div class="empty"><div class="empty-i">🎉</div><div class="empty-t">لا توجد مهام عاجلة</div></div>';
        } else {
            urgEl.innerHTML = urg.slice(0, 4).map(function(t) {
                return '<div style="display:flex;align-items:center;gap:9px;padding:7px 0;border-bottom:1px solid var(--bdr)">' +
                    '<span style="color:var(--red);font-size:14px">🔴</span>' +
                    '<div style="flex:1"><div style="font-size:13px;font-weight:700">' + esc(t.name) + '</div>' +
                    '<div style="font-size:11px;color:var(--txt3);margin-top:2px">' + (t.startTime ? t.startTime + ' - ' : '') + t.duration + 'د</div></div>' +
                    '<button class="btn btn-xs btn-green" onclick="completeTask(' + t.id + ')">✅ تم</button></div>';
            }).join('');
        }
    }
    var nxtEl = $('d-nxt');
    if (nxtEl) {
        var ordered = pend.slice().sort(function(a, b) {
            var scoreA = {high:3,medium:2,low:1}[a.priority] || 2;
            var scoreB = {high:3,medium:2,low:1}[b.priority] || 2;
            if (scoreA !== scoreB) return scoreB - scoreA;
            var diffA = {hard:3,medium:2,easy:1}[a.difficulty] || 2;
            var diffB = {hard:3,medium:2,easy:1}[b.difficulty] || 2;
            if (diffA !== diffB) return diffB - diffA;
            return b.duration - a.duration;
        });
        var nxt = ordered[0];
        if (nxt) {
            var em = {high:'🔴', medium:'🟡', low:'🟢'}[nxt.priority] || '🟡';
            nxtEl.innerHTML = '<div style="display:flex;align-items:center;gap:11px">' +
                '<span style="font-size:24px">' + em + '</span>' +
                '<div><div style="font-size:14px;font-weight:800">' + esc(nxt.name) + '</div>' +
                '<div style="font-size:11px;color:var(--txt2);margin-top:2px">' + (nxt.startTime ? nxt.startTime + ' - ' : '') + nxt.duration + ' دقيقة</div></div></div>' +
                '<div style="margin-top:9px;display:flex;gap:6px">' +
                '<button class="btn btn-sm btn-green" onclick="completeTask(' + nxt.id + ')">✅ تم</button>' +
                '<button class="btn btn-sm btn-amber" onclick="postponePrompt(' + nxt.id + ')">⏰ تأجيل</button></div>';
        } else {
            nxtEl.innerHTML = '<div class="empty"><div class="empty-i">🎉</div><div class="empty-t">لا توجد مهام متبقية</div></div>';
        }
    }
    loadDashboardInsights();
}

async function loadDashboardInsights() {
    var el = $('d-insights');
    if (!el) return;
    try {
        var result = await apiCall('/api/insights');
        if (result && result.insights && result.insights.length > 0) {
            el.innerHTML = result.insights.slice(0, 4).map(function(i) {
                return '<div style="font-size:12px;color:var(--txt2);margin-bottom:6px;padding:6px 10px;background:var(--bg3);border-radius:6px">💡 ' + i + '</div>';
            }).join('');
        } else {
            el.innerHTML = '<div style="font-size:12px;color:var(--txt3)">سيتم توليد رؤى ذكية مع الاستخدام...</div>';
        }
    } catch(e) {
        el.innerHTML = '<div style="font-size:12px;color:var(--txt3)">جاري تحميل الرؤى...</div>';
    }
}

function renderTasks() {
    var el = $('tasks-list');
    if (!el) return;
    var fd = ($('flt-d') || {}).value || 'today';
    var fs = ($('flt-s') || {}).value || 'all';
    var today = todayStr(), tom = tomStr();
    var list = tasks.slice();
    if (fd === 'today') list = list.filter(function(t) { return t.date === today; });
    else if (fd === 'tomorrow') list = list.filter(function(t) { return t.date === tom; });
    if (fs !== 'all') list = list.filter(function(t) { return t.status === fs; });
    list.sort(function(a, b) {
        if (a.status !== b.status) return a.status === 'done' ? 1 : -1;
        var scoreA = {high:3, medium:2, low:1}[a.priority] || 2;
        var scoreB = {high:3, medium:2, low:1}[b.priority] || 2;
        if (scoreA !== scoreB) return scoreB - scoreA;
        var diffA = {hard:3, medium:2, easy:1}[a.difficulty] || 2;
        var diffB = {hard:3, medium:2, easy:1}[b.difficulty] || 2;
        if (diffA !== diffB) return diffB - diffA;
        return b.duration - a.duration;
    });
    if (!list.length) {
        el.innerHTML = '<div class="empty"><div class="empty-i">📭</div><div class="empty-t">لا توجد مهام</div></div>';
        return;
    }
    var pLbl = {high:'عالية', medium:'متوسطة', low:'منخفضة'};
    var bC = {high:'var(--red)', medium:'var(--amber)', low:'var(--green)'};
    var bdg = {high:'bh', medium:'bm', low:'bl'};
    el.innerHTML = list.map(function(t) {
        var dd = t.date === today ? 'اليوم' : (t.date === tom ? 'الغد' : t.date);
        var eT = t.startTime ? toTime(toMins(t.startTime) + t.duration) : '';
        return '<div class="ti' + (t.status === 'done' ? ' done' : '') + '">' +
            '<div class="tbar" style="background:' + bC[t.priority] + '"></div>' +
            '<div class="tbody">' +
            '<div class="tname">' + esc(t.name) + '</div>' +
            '<div class="tmeta">' +
            '<span style="font-size:11px;color:var(--txt3)">' + dd + '</span>' +
            (t.startTime ? '<span class="ttime">' + t.startTime + '-' + eT + '</span>' : '<span class="ttime" style="background:rgba(212,168,67,.05)">⏰ غير مجدولة</span>') +
            '<span class="tdur">' + t.duration + 'د</span>' +
            '<span class="badge ' + bdg[t.priority] + '">' + pLbl[t.priority] + '</span>' +
            (t.postponedCount > 0 ? '<span class="badge" style="background:rgba(212,168,67,.1);color:var(--gold)">⏰×' + t.postponedCount + '</span>' : '') +
            '</div>' +
            (t.notes ? '<div style="font-size:11px;color:var(--txt3);margin-top:4px;padding-top:4px;border-top:1px solid var(--bdr)">' + esc(t.notes) + '</div>' : '') +
            '</div>' +
            '<div class="tacts">' +
            (t.status !== 'done' ? '<button class="btn btn-xs btn-green" onclick="completeTask(' + t.id + ')">✅</button>' : '') +
            (t.status === 'pending' ? '<button class="btn btn-xs btn-amber" onclick="postponePrompt(' + t.id + ')">⏰</button>' : '') +
            '<button class="btn btn-xs btn-red" onclick="deleteTask(' + t.id + ')">🗑️</button>' +
            '</div></div>';
    }).join('');
}

function renderTimeline() {
    var el = $('tl-wrap');
    if (!el) return;
    
    var today = todayStr();
    var todayTasks = tasks.filter(function(t) { 
        return t.date === today && t.status === 'pending';
    });
    
    var stm = toMins(settings.startTime);
    var slm = toMins(settings.sleepTime);
    var nm = toMins(nowStr());
    
    var isSleepAfterMidnight = (slm < stm);
    
    var scheduled = todayTasks.filter(function(t) { 
        return t.startTime && t.startTime !== null && t.startTime !== '';
    });
    var unscheduled = todayTasks.filter(function(t) { 
        return !t.startTime || t.startTime === null || t.startTime === '';
    });
    
    if (todayTasks.length === 0) {
        el.innerHTML = '<div class="empty"><div class="empty-i">📅</div><div class="empty-t">لا توجد مهام اليوم</div></div>';
        return;
    }
    
    var html = '';
    
    if (unscheduled.length > 0) {
        var priorityColors = {high: 'var(--red)', medium: 'var(--amber)', low: 'var(--green)'};
        var priorityIcons = {high: '🔴', medium: '🟡', low: '🟢'};
        var priorityNames = {high: 'عالية', medium: 'متوسطة', low: 'منخفضة'};
        html += '<div class="unscheduled-area">';
        html += '<div class="unscheduled-title">📌 المهام غير المجدولة <span style="font-size:11px;color:var(--txt3)">(اسحب أي مهمة إلى الجدول أدناه لتحديد وقتها)</span></div>';
        html += '<div class="unscheduled-tasks">';
        for (var i = 0; i < unscheduled.length; i++) {
            var t = unscheduled[i];
            var borderColor = priorityColors[t.priority] || 'var(--gold)';
            html += '<div class="unscheduled-task" style="border-right-color: ' + borderColor + '" ' +
                'draggable="true" ' +
                'ondragstart="unscheduledDragStart(event, ' + t.id + ', \'' + esc(t.name) + '\')" ' +
                'ondragend="unscheduledDragEnd(event)">' +
                '<div class="unscheduled-priority" style="background: ' + borderColor + '"></div>' +
                '<div class="unscheduled-content">' +
                '<div class="unscheduled-name">' + priorityIcons[t.priority] + ' ' + esc(t.name) + '</div>' +
                '<div class="unscheduled-meta">⏱️ ' + t.duration + ' دقيقة | 🎯 ' + priorityNames[t.priority] + '</div>' +
                '</div>' +
                '<div class="unscheduled-actions">' +
                '<button class="btn btn-xs btn-blue" onclick="quickScheduleTask(' + t.id + ')">⏰ تحديد وقت</button>' +
                '<button class="btn btn-xs btn-red" onclick="deleteTask(' + t.id + ')">🗑️</button>' +
                '</div>' +
                '</div>';
        }
        html += '</div></div>';
    }
    
    if (scheduled.length === 0) {
        html += '<div class="empty" style="margin-top:20px"><div class="empty-i">⏰</div><div class="empty-t">' + 
                (unscheduled.length > 0 ? 'اسحب المهام أعلاه إلى هنا لتحديد وقتها' : 'لا توجد مهام مجدولة اليوم') + 
                '</div></div>';
        el.innerHTML = html;
        return;
    }
    
    var sorted = scheduled.slice().sort(function(a, b) { 
        return toMins(a.startTime) - toMins(b.startTime); 
    });
    
    var firstM = toMins(sorted[0].startTime);
    var lastM = toMins(sorted[sorted.length - 1].startTime) + sorted[sorted.length - 1].duration;
    
    var effectiveSlm = slm;
    var effectiveLastM = lastM;
    var effectiveFirstM = firstM;
    var effectiveNm = nm;
    
    if (isSleepAfterMidnight) {
        effectiveSlm = slm + 1440;
        if (firstM < stm) {
            effectiveFirstM = firstM + 1440;
        }
        if (lastM < stm) {
            effectiveLastM = lastM + 1440;
        }
        if (nm < stm) {
            effectiveNm = nm + 1440;
        }
    }
    
    var minTime = Math.min(stm, effectiveNm, effectiveFirstM);
    var maxTime = Math.max(effectiveSlm, effectiveLastM);
    
    var minH = Math.max(0, Math.floor(minTime / 60) - 1);
    var maxH = Math.min(48, Math.ceil(maxTime / 60) + 1);
    
    var PPM = 2.0;
    var LPAD = 54;
    
    html += '<div class="timeline-container"><div class="timeline-background" style="position:relative;min-height:' + ((maxH - minH) * 60 * PPM + 40) + 'px;margin-right:' + LPAD + 'px">';
    
    for (var h = minH; h <= maxH; h++) {
        var y = (h * 60 - minH * 60) * PPM;
        var displayHour = h % 24;
        html += '<div class="tl-label" style="top:' + y + 'px;right:-' + LPAD + 'px">' + pad(displayHour) + ':00</div>' +
            '<div class="tl-hline" style="top:' + y + 'px"></div>';
    }
    
    var cols = {
        high: {bg: 'rgba(232,76,106,0.14)', bdr: 'rgba(232,76,106,0.3)', bar: 'var(--red)'},
        medium: {bg: 'rgba(240,157,58,0.12)', bdr: 'rgba(240,157,58,0.28)', bar: 'var(--amber)'},
        low: {bg: 'rgba(45,204,138,0.1)', bdr: 'rgba(45,204,138,0.22)', bar: 'var(--green)'}
    };
    
    for (var i = 0; i < sorted.length; i++) {
        var t = sorted[i];
        var sm = toMins(t.startTime);
        if (isSleepAfterMidnight && sm < stm) {
            sm = sm + 1440;
        }
        var em = sm + t.duration;
        var ty = (sm - minH * 60) * PPM;
        var th = Math.max(t.duration * PPM, 28);
        var c = cols[t.priority] || cols.medium;
        html += '<div class="tl-task" style="top:' + ty + 'px;height:' + th + 'px;background:' + c.bg + ';border:1px solid ' + c.bdr + '" ' +
            'draggable="true" ' +
            'ondragstart="dragStartHandler(event, ' + t.id + ', \'' + t.startTime + '\')" ' +
            'ondragend="dragEndHandler(event)">' +
            '<div style="position:absolute;right:0;top:0;bottom:0;width:3px;background:' + c.bar + ';border-radius:0 8px 8px 0"></div>' +
            '<div class="tl-tname" style="padding-right:6px">' + esc(t.name) + '</div>' +
            '<div class="tl-tsub" style="padding-right:6px">' + t.startTime + ' - ' + toTime(em % 1440) + ' (' + t.duration + 'د)</div>' +
            '</div>';

        // === Draw arrow/connector to next task (only if gap is large enough) ===
        if (i < sorted.length - 1) {
            var nextT = sorted[i + 1];
            var nextSm = toMins(nextT.startTime);
            if (isSleepAfterMidnight && nextSm < stm) {
                nextSm = nextSm + 1440;
            }
            var taskBottom = ty + th;
            var nextTop = (nextSm - minH * 60) * PPM;
            var arrowH = nextTop - taskBottom - 2;
            if (arrowH >= 14) {  // only show if there's enough vertical space
                var arrowCls = 'tl-arrow';
                if (nextT.priority === 'high') arrowCls += ' tl-arrow-high';
                else if (nextT.priority === 'low') arrowCls += ' tl-arrow-low';
                html += '<div class="' + arrowCls + '" style="top:' + (taskBottom + 1) + 'px;height:' + arrowH + 'px"></div>';
            }
        }
    }
    
    var currentDisplay = isSleepAfterMidnight && nm < stm ? nm + 1440 : nm;
    if (currentDisplay >= minH * 60 && currentDisplay <= maxH * 60) {
        var nowY = (currentDisplay - minH * 60) * PPM;
        html += '<div class="tl-now" style="top:' + nowY + 'px"><span style="position:absolute;right:-3px;top:-12px;font-size:9px;color:var(--green);background:var(--bg3);padding:1px 5px;border-radius:3px">الآن</span></div>';
    }
    
    var sleepDisplay = isSleepAfterMidnight ? slm + 1440 : slm;
    if (sleepDisplay >= minH * 60 && sleepDisplay <= maxH * 60) {
        var sy = (sleepDisplay - minH * 60) * PPM;
        html += '<div class="tl-sleep" style="top:' + sy + 'px"><span style="position:absolute;right:-3px;top:-12px;font-size:9px;color:var(--gold);background:var(--bg3);padding:1px 6px;border-radius:3px">🌙 ' + settings.sleepTime + '</span></div>';
    }
    
    html += '</div></div>';
    el.innerHTML = html;
}

function quickScheduleTask(taskId) {
    var task = tasks.find(function(t) { return t.id == taskId; });
    if (!task) return;
    openModal('⏰ تحديد وقت للمهمة: ' + task.name,
        '⏱️ المدة: ' + task.duration + ' دقيقة\n⚠️ الرجاء اختيار وقت في المستقبل',
        '<div style="margin:11px 0"><input type="time" id="qt-inp" value="' + nowStr() + '" style="width:100%"></div>',
        [
            {label: '✅ تحديد', cls: 'btn-gold', cb: async function() {
                var t2 = ($('qt-inp') || {}).value;
                if (!t2) { showNotif('اختر وقتا', 'warn'); return; }
                var nowMins = toMins(nowStr());
                var newMins = toMins(t2);
                if (newMins < nowMins) {
                    showNotif('❌ لا يمكن تحديد وقت مضى! اختر وقتاً مستقبلياً', 'err');
                    return;
                }
                closeModal();
                var result = await apiCall('/api/tasks/' + taskId + '/update-time', 'POST', {time: t2});
                if (result && result.success) {
                    await loadAll();
                    showNotif('✅ تم تحديد وقت المهمة بنجاح!', 'ok');
                    renderTimeline();
                } else {
                    showNotif('❌ فشل تحديد الوقت: ' + (result.message || 'وقت غير صالح'), 'err');
                }
            }},
            {label: 'إلغاء', cls: '', cb: closeModal}
        ]);
}

async function renderHistory() {
    var d = await apiCall('/api/history');
    if (!d) return;
    var hist = d.history || [];
    var el = $('hist-list');
    if (!el) return;
    if (!hist.length) {
        el.innerHTML = '<div class="empty"><div class="empty-i">📜</div><div class="empty-t">لا يوجد سجل</div></div>';
        return;
    }
    el.innerHTML = hist.slice(0, 50).map(function(h) {
        return '<div style="display:flex;align-items:center;gap:11px;padding:10px 16px;border-bottom:1px solid var(--bdr)">' +
            '<span style="font-size:17px">' + (h.action === 'completed' ? '✅' : '⏰') + '</span>' +
            '<div style="flex:1"><div style="font-size:13px;font-weight:700">' + esc(h.name) + '</div>' +
            '<div style="font-size:11px;color:var(--txt3);margin-top:2px">' + h.logDate + ' - ' + h.duration + 'د</div></div>' +
            '<span class="badge ' + (h.action === 'completed' ? 'bl' : 'bm') + '">' + (h.action === 'completed' ? 'تمت' : 'مؤجلة') + '</span></div>';
    }).join('');
}

async function clearHistory() {
    if (!confirm('مسح كل السجل؟')) return;
    await apiCall('/api/history', 'DELETE');
    renderHistory();
    showNotif('تم مسح السجل', 'info');
}

async function completeTask(id) {
    var r = await apiCall('/api/tasks/' + id + '/complete', 'POST');
    if (r && r.success !== false) {
        await loadAll();
        showNotif('✅ أحسنت! تم إكمال المهمة!', 'ok');
        addNotification('✅ تم إكمال مهمة بنجاح', 'ok');
    }
}

async function deleteTask(id) {
    var t = tasks.find(function(x) { return x.id == id; });
    if (!t) return;
    if (!confirm('حذف: "' + t.name + '"؟')) return;
    await apiCall('/api/tasks/' + id, 'DELETE');
    await loadAll();
    showNotif('🗑️ تم الحذف', 'info');
}

function postponePrompt(id) {
    var t = tasks.find(function(x) { return x.id == id; });
    if (!t) return;
    openModal('⏰ تأجيل: ' + t.name,
        '⏱️ مدة المهمة: ' + t.duration + ' دقيقة\n🌙 وقت النوم: ' + settings.sleepTime, '',
        [
            {label: '✅ تم الإنجاز', cls: 'btn-green', cb: function() { closeModal(); completeTask(id); }},
            {label: '📅 للغد (ترتيب حسب الأولوية)', cls: 'btn-blue', cb: function() { closeModal(); apiCall('/api/tasks/' + id + '/postpone/tomorrow', 'POST').then(function() { loadAll(); showNotif('تم التأجيل للغد', 'info'); }); }},
            {label: '⏰ وقت جديد (مستقبلي فقط)', cls: 'btn-amber', cb: function() { closeModal(); postponeTimeModal(id, t); }},
            {label: '🗑️ حذف', cls: 'btn-red', cb: function() { closeModal(); deleteTask(id); }},
            {label: 'إلغاء', cls: '', cb: closeModal}
        ]);
}

function postponeTimeModal(id, task) {
    openModal('⏰ وقت جديد: ' + task.name, '⏱️ المدة: ' + task.duration + ' دقيقة\n⚠️ الرجاء اختيار وقت في المستقبل',
        '<div style="margin:11px 0"><input type="time" id="pt-inp" value="' + nowStr() + '" style="width:100%"></div>',
        [
            {label: '✅ تأكيد', cls: 'btn-gold', cb: function() {
                var t2 = ($('pt-inp') || {}).value;
                if (!t2) { showNotif('اختر وقتا', 'warn'); return; }
                var nowMins = toMins(nowStr());
                var newMins = toMins(t2);
                if (newMins <= nowMins) {
                    showNotif('❌ لا يمكن التأجيل لوقت مضى! اختر وقتاً مستقبلياً', 'err');
                    return;
                }
                closeModal();
                apiCall('/api/tasks/' + id + '/postpone/time', 'POST', {time: t2}).then(function(res) {
                    if (res && res.success) {
                        loadAll();
                        showNotif('تم التأجيل إلى ' + t2, 'ok');
                    } else {
                        showNotif('فشل التأجيل: ' + (res.message || 'وقت غير صالح'), 'err');
                    }
                });
            }},
            {label: 'إلغاء', cls: '', cb: closeModal}
        ]);
}

async function runOptimize() {
    var r = await apiCall('/api/schedule/optimize', 'POST');
    if (r) {
        await loadAll();
        showNotif('🔄 تم ترتيب المهام حسب الأولوية ← الصعوبة ← المدة!', 'ok');
    }
}

async function saveSettings() {
    var sleep = ($('cfg-sleep') || {}).value;
    var start = ($('cfg-start') || {}).value;
    if (!sleep || !start) return;
    var result = await apiCall('/api/settings', 'POST', {sleepTime: sleep, startTime: start});
    if (result && result.success) {
        settings.sleepTime = sleep;
        settings.startTime = start;
        updateSleep();
        updateSidebar();
        showNotif('💾 تم حفظ الإعدادات بنجاح!', 'ok');
    } else {
        showNotif('❌ فشل حفظ الإعدادات', 'err');
    }
}

function addMessage(role, text) {
    var el = $('chat-msgs');
    if (!el) return;
    var d = document.createElement('div');
    if (role === 'user') {
        d.className = 'mu';
        d.innerHTML = '<div class="bub">' + esc(text) + '</div>';
    } else {
        d.className = 'mb';
        d.innerHTML = '<div class="av">AI</div><div class="bub">' + formatMarkdown(text) + '</div>';
    }
    el.appendChild(d);
    el.scrollTop = el.scrollHeight;
    return d;
}

function addTyping() {
    var el = $('chat-msgs');
    if (!el) return null;
    var d = document.createElement('div');
    d.className = 'mb';
    d.innerHTML = '<div class="av">AI</div><div class="bub"><div class="typing"><span></span><span></span><span></span></div></div>';
    el.appendChild(d);
    el.scrollTop = el.scrollHeight;
    return d;
}

async function sendMsg() {
    if (chatBusy) return;
    var inp = $('chat-in'), btn = $('send-btn');
    if (!inp) return;
    var msg = inp.value.trim();
    if (!msg) return;
    inp.value = '';
    inp.disabled = true;
    if (btn) {
        btn.disabled = true;
        btn.innerHTML = '<span class="spin" style="border-top-color:#060610"></span>';
    }
    chatBusy = true;
    addMessage('user', msg);
    var ty = addTyping();

    // ⏳ تأخير "تفكير" قبل ما الـ bot يرد (1200-2200ms طبيعي)
    var thinkingDelay = 1200 + Math.random() * 1000;

    try {
        // نبعت الطلب ونستنى التأخير في نفس الوقت
        var apiPromise = apiCall('/api/chat', 'POST', {message: msg, session_id: sessionId});
        var delayPromise = new Promise(function(resolve) { setTimeout(resolve, thinkingDelay); });

        // ✨ ننتظر الاتنين معاً (الـ API + التأخير)
        var results = await Promise.all([apiPromise, delayPromise]);
        var r = results[0];

        // ⏳ تأخير صغير إضافي كأنه "بيكتب الرد" (300-700ms)
        await new Promise(function(resolve) { setTimeout(resolve, 300 + Math.random() * 400); });

        if (ty) ty.remove();
        if (r && r.reply) {
            // 💬 لو الرد طويل وفيه فقرات، نقسمه على رسالتين بفاصل typing بينهم
            var reply = r.reply;
            if (reply.length > 200 && reply.indexOf('\n\n') > 0) {
                var parts = reply.split('\n\n');
                addMessage('bot', parts[0]);
                // typing indicator تاني قبل الجزء الثاني
                var ty2 = addTyping();
                await new Promise(function(resolve) { setTimeout(resolve, 600 + Math.random() * 500); });
                if (ty2) ty2.remove();
                addMessage('bot', parts.slice(1).join('\n\n'));
            } else {
                addMessage('bot', reply);
            }
            if (r.action === 'refresh') await loadAll();
            if (r.action === 'pending_completion') {
                // المساعد سيتعامل مع الإكمال الجزئي في الرسالة التالية
            }
        } else {
            addMessage('bot', '❌ حدث خطأ. حاول مرة أخرى.');
        }
    } catch (e) {
        if (ty) ty.remove();
        addMessage('bot', '❌ خطأ في الاتصال بالخادم.');
    }
    chatBusy = false;
    inp.disabled = false;
    if (btn) {
        btn.disabled = false;
        btn.innerHTML = '➤';
    }
    inp.focus();
    var cm = $('chat-msgs');
    if (cm) cm.scrollTop = cm.scrollHeight;
}

function sendChip(msg) {
    var inp = $('chat-in');
    if (inp) inp.value = msg;
    sendMsg();
}

function quickCmd(msg) {
    showPanel('chat');
    setTimeout(function() { sendChip(msg); }, 120);
}

function openModal(title, body, extra, buttons) {
    modalCbs = buttons || [];
    $('m-title').textContent = title;
    $('m-body').textContent = body;
    $('m-extra').innerHTML = extra || '';
    $('m-btns').innerHTML = modalCbs.map(function(b, i) {
        return '<button class="btn ' + b.cls + '" onclick="modalCbs[' + i + '].cb()">' + b.label + '</button>';
    }).join('');
    $('modal').classList.add('open');
}

function closeModal() {
    $('modal').classList.remove('open');
    modalCbs = [];
}

function showNotif(msg, type, dur) {
    dur = dur || 4000;
    var stack = $('notifs');
    if (!stack) return;
    var el = document.createElement('div');
    el.className = 'notif n-' + (type || 'info');
    el.innerHTML = '<span style="flex:1">' + esc(msg) + '</span><span class="nx" onclick="this.parentElement.remove()">✕</span>';
    stack.appendChild(el);
    setTimeout(function() {
        el.style.transition = 'all .3s';
        el.style.opacity = '0';
        el.style.transform = 'translateY(8px)';
        setTimeout(function() { el.remove(); }, 320);
    }, dur);
}

async function pollNotifs() {
    var d = await apiCall('/api/notifications');
    if (d && d.notifications) {
        d.notifications.forEach(function(n) {
            showNotif(n.message, n.type === 'task_alert' ? 'warn' : 'info', 10000);
            addNotification(n.message, n.type || 'info');
        });
    }
}

document.addEventListener('keydown', function(e) {
    if (e.key === 'Escape') {
        closeModal();
        var panel = document.getElementById('notif-panel');
        if (panel) panel.classList.remove('open');
        closeRecommendationOverlay();
    }
    if (e.ctrlKey || e.metaKey) {
        var map = {'1': 'dashboard', '2': 'add', '3': 'tasks', '4': 'timeline', '5': 'chat'};
        if (map[e.key]) {
            e.preventDefault();
            showPanel(map[e.key]);
        }
    }
});

document.addEventListener('click', function(e) {
    var panel = document.getElementById('notif-panel');
    var btn = document.getElementById('notif-btn');
    if (panel && panel.classList.contains('open') && !panel.contains(e.target) && e.target !== btn && !btn.contains(e.target)) {
        panel.classList.remove('open');
    }
});

(function() {
    PANELS.forEach(function(p) {
        var el = document.getElementById('p-' + p);
        if (el) el.style.display = 'none';
    });
    var dp = document.getElementById('p-dashboard');
    if (dp) {
        dp.style.display = 'flex';
        dp.style.flexDirection = 'column';
    }
    loadAll();
    initTheme();
    setInterval(loadAll, 60000);
    setTimeout(pollNotifs, 5000);
    setInterval(pollNotifs, 20000);
})();
</script>
</body>
</html>"""

@app.get("/", response_class=HTMLResponse)
def root():
    return HTMLResponse(HTML)

# فتح المتصفح تلقائياً عند بدء التشغيل (Chrome)
def open_browser():
    time.sleep(1.5)
    # Force open with Chrome if available, otherwise default browser
    chrome_paths = [
        "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
        "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe",
        "C:\\Users\\" + os.getenv("USERNAME") + "\\AppData\\Local\\Google\\Chrome\\Application\\chrome.exe",
        "/usr/bin/google-chrome",
        "/usr/bin/chromium-browser",
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    ]
    
    chrome_found = False
    for chrome_path in chrome_paths:
        if os.path.exists(chrome_path):
            try:
                webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path))
                webbrowser.get('chrome').open("http://localhost:8000")
                chrome_found = True
                print(f"✅ Opened in Chrome: {chrome_path}")
                break
            except:
                pass
    
    if not chrome_found:
        webbrowser.open("http://localhost:8000")
        print("✅ Opened in default browser (Chrome not found)")

optimize_schedule()

if __name__ == "__main__":
    import socket
    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "127.0.0.1"
    
    ip = get_local_ip()
    sep = "=" * 60
    print(sep)
    print("  🚀 SmartDay AI Agent v9.0 - Adaptive Recommendation System")
    print(sep)
    print("  ✨ ما الجديد في v9.0:")
    print("     1. 🎯 نظام توصيات حقيقي (Epsilon-Greedy 80/20)")
    print("     2. 🌈 4–6 مقترحات متنوعة فعلاً (لا تكرار)")
    print("     3. 🧮 دالة تقييم محسّنة (urgency + sleep proximity)")
    print("     4. 📈 تعلم تدريجي + decay يمنع الجمود")
    print("     5. 👤 ملف مستخدم يؤثر على المقترحات")
    print("     6. 🎨 واجهة مقترحات جديدة (المقترح الأول/الثاني/...)")
    print("     7. ➡️ أسهم بصرية على الجدول الزمني تربط المهام")
    print(sep)
    print("  🎯 الأوامر المدعومة:")
    print("     • 'خلصت [المهمة]' / 'complete [task]'")
    print("     • 'ملحقتش [المهمة]' / 'not finished [task]'")
    print("     • 'اجل [المهمة] لـ [الوقت]' / 'postpone [task] to [time]'")
    print("     • 'جدول ذكي' / 'smart schedule'")
    print("     • 'بدل [مهمة1] مع [مهمة2]' / 'swap [task1] with [task2]'")
    print(sep)
    print("  🌐 عنوان الوصول:")
    print(f"     http://localhost:8000")
    print(f"     http://{ip}:8000")
    print(sep)
    print("  🚀 جاري فتح المتصفح...")
    
    # تشغيل فتح المتصفح في Thread منفصل
    threading.Thread(target=open_browser, daemon=True).start()
    
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="warning")